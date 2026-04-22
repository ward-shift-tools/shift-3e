#!/usr/bin/env python3
"""
看護師勤務表自動作成ツール v5.0
PuLP（混合整数計画法）による最適化

使い方:
  ■ ローカルExcel
    python3 shift_scheduler.py --init              → 入力テンプレート生成
    python3 shift_scheduler.py                     → 勤務表作成

  ■ Google スプレッドシート
    python3 shift_scheduler.py --init-gsheet       → テンプレート生成
    python3 shift_scheduler.py --gsheet URL_or_ID  → 読み込み → 結果書き戻し

v5.0 変更点:
  - 前月繰越対応（前月末の夜勤/明けを引き継ぎ）
  - 希望休上限7日、希望はソフト制約（未達時アラート）
  - 祝日対応（設定シートで日付指定）
  - 週N勤務対応（日勤限定パートタイム）
"""

import argparse
import calendar
import os
import re
import sys
from datetime import date

try:
    import jpholiday
except ImportError:
    print("jpholidayが必要です: pip3 install jpholiday")
    sys.exit(1)

try:
    import pulp
except ImportError:
    print("PuLPが必要です: pip3 install pulp")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxlが必要です: pip3 install openpyxl")
    sys.exit(1)

# ============================================================
# 定数
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, "勤務表_入力.xlsx")
CRED_FILE = os.path.join(BASE_DIR, "credentials.json")

D = "日"; N = "夜"; A = "明"; O = "休"; R = "研"; V = "暇"
# 拡張勤務形態
E  = "早"   # 早出: 7:00〜16:00 (8h) 日勤系
L  = "遅"   # 遅出: 12:00〜21:00 (8h) 日勤系
ST = "短"   # 時短: 8:45〜16:00 (6.25h) 育児・介護対応
LD = "長"   # 長日勤: 8:45〜21:00 (12h) 翌日休推奨
SN = "準"   # 短夜勤: 17:00〜翌5:00 (12h) 夜勤系・翌日明け
I  = "委"   # 委員会: 勤務日としてカウントされるが時間中に離席→A/ABバックアップ必須

SHIFTS = [D, N, A, O, R, V, E, L, ST, LD, SN, I]
# 日勤系シフト（日勤人数カウントに含む）
DAY_SHIFTS = [D, E, L, ST, LD, I]
# リーダー判定用の純粋な日勤系（委=離席枠は含まない）
PURE_DAY_SHIFTS = [D, E, L, ST, LD]
# 夜勤系シフト（夜勤人数カウントに含む）
NIGHT_SHIFTS = [N, SN]
# 夜勤系ごとの時間数（72h計算用）
NIGHT_HOURS_MAP = {N: 16, SN: 12}  # デフォルト値 / 実際は設定値で上書き

VALID_REQUEST = {D, N, O, R, E, L, ST, LD, SN, I, "夜不", "休暇", "明休", "遅希"}
# 3E クラス制度（ユニット能力） + リーダーフラグ分離
CLS_ER  = "ER可"         # ER/HCU/病棟配置可（ERに入れる唯一のクラス）
CLS_HCU = "HCU可"        # HCU/病棟配置可
CLS_WD  = "病棟可"       # 病棟のみ
VALID_CLASSES = {CLS_ER, CLS_HCU, CLS_WD}
# 後方互換エイリアス（Excel出力・旧コード用。実体は同じ）
CLS_ERL = CLS_ER
CLS_LDR = CLS_HCU
TIER_A = CLS_ER; TIER_AB = CLS_HCU; TIER_B = CLS_ER
TIER_CP = CLS_HCU; TIER_C = CLS_WD
VALID_TIERS = VALID_CLASSES
MAX_REQUEST_DAYS = 7

# ユニット定数
UNIT_WD   = "病棟"
UNIT_HCU  = "HCU"
UNIT_ER   = "ER"
UNIT_LEAD = "共リーダー"
DAY_UNITS = [UNIT_WD, UNIT_HCU, UNIT_ER, UNIT_LEAD]
NIGHT_UNITS = [UNIT_WD, UNIT_HCU, UNIT_LEAD]

# openpyxl styles
_F = Font; _PF = PatternFill; _A = Alignment; _B = Border; _S = Side
FONT    = _F(color="000000", size=10)
FONT_H  = _F(bold=True, size=11)
FONT_T  = _F(bold=True, size=14)
FONT_N  = _F(color="FFFFFF", bold=True, size=10)
FONT_R  = _F(bold=True, color="FF0000", size=10)
FONT_RS = _F(color="FF0000", size=9)
CTR     = _A(horizontal="center", vertical="center")
BDR     = _B(left=_S(style="thin"), right=_S(style="thin"),
             top=_S(style="thin"), bottom=_S(style="thin"))
BDR_REQ = _B(left=_S(style="medium", color="FF6600"),
             right=_S(style="medium", color="FF6600"),
             top=_S(style="medium", color="FF6600"),
             bottom=_S(style="medium", color="FF6600"))
FH  = _PF(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FS  = {
    D:  _PF(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),  # 日勤: 白
    N:  _PF(start_color="4472C4", end_color="4472C4", fill_type="solid"),  # 夜勤: 青
    A:  _PF(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # 明け: 薄黄
    O:  _PF(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # 休み: 薄緑
    R:  _PF(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid"),  # 研修: 薄紫
    V:  _PF(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),  # 休暇: 薄赤
    "早": _PF(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),  # 早出: 水色
    "遅": _PF(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid"),  # 遅出: 薄橙
    "短": _PF(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid"),  # 時短: 薄緑系
    "長": _PF(start_color="FFF5CC", end_color="FFF5CC", fill_type="solid"),  # 長日勤: 薄黄緑
    "準": _PF(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),  # 短夜勤: 濃青
    "委": _PF(start_color="FFE699", end_color="FFE699", fill_type="solid"),  # 委員会: 金色系
}
FT = {
    CLS_ER:  _PF(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    CLS_HCU: _PF(start_color="FFF4E0", end_color="FFF4E0", fill_type="solid"),
    CLS_WD:  _PF(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid"),
}

# SAMPLE: (名前, クラス, リーダー可, ERリーダー可, 遅出可, 週勤務, 前月末)
SAMPLE_STAFF = [
    ("田中",  CLS_ER,  True,  True,  True,  "", ""),
    ("佐藤",  CLS_ER,  True,  True,  False, "", ""),
    ("鈴木",  CLS_ER,  True,  False, True,  "", ""),
    ("高橋",  CLS_HCU, True,  False, False, "", ""),
    ("山田",  CLS_ER,  False, False, True,  "", ""),
    ("中村",  CLS_ER,  False, False, False, "", ""),
    ("小林",  CLS_HCU, False, False, True,  "", ""),
    ("加藤",  CLS_HCU, False, False, False, "", ""),
    ("吉田",  CLS_WD,  False, False, False, "", ""),
    ("山口",  CLS_WD,  False, False, False, "", ""),
]

SETTINGS_DEF = [
    ("対象年",                2026, "西暦4桁"),
    ("対象月",                5,    "1〜12"),
    ("公休日数",              "",   "空欄=自動(土日祝) 数値=手動指定"),
    ("病棟最低人数_平日",     4,    "平日の病棟最低配置人数"),
    ("HCU最低人数_平日",      2,    "平日のHCU最低配置人数"),
    ("ER最低人数_平日",       3,    "平日のER最低配置人数（1名はERリーダー必須）"),
    ("病棟最低人数_休日",     4,    "土日祝の病棟最低配置人数"),
    ("HCU最低人数_休日",      2,    "土日祝のHCU最低配置人数"),
    ("休日日勤下限",          8,    "土日祝の日勤系(D/遅/早等)合計の下限"),
    ("休日日勤上限",          8,    "土日祝の日勤系(D/遅/早等)合計の上限"),
    ("平日日勤上限",          14,   "平日の日勤系(D/遅/早等)合計の上限（偏り防止）"),
    ("夜勤上限",              5,    "月あたり最大夜勤回数"),
    ("夜勤推奨",              4,    "月あたり推奨夜勤回数"),
    ("最大連勤日数",          5,    "超過禁止"),
    ("推奨連勤日数",          4,    "超過ペナルティ"),
    ("計算時間上限（秒）",    120,  ""),
    ("祝日",                  "",   "日付をカンマ区切り 例: 3,4,5"),
]
SETTINGS_KEYS = [
    "year", "month", "public_off_override",
    "min_ward_wd", "min_hcu_wd", "min_er_wd",
    "min_ward_hd", "min_hcu_hd",
    "min_day_staff_hd", "max_day_staff_hd",
    "max_day_staff_wd",
    "max_night", "pref_night",
    "max_consecutive", "pref_consecutive",
    "solver_time_limit", "holidays",
]


# ============================================================
# データ構造
# ============================================================
class Staff:
    def __init__(self, name, cls, is_leader=False, is_er_leader=False,
                 can_late=False, weekly_days=None, prev_month="",
                 night_min=None, night_max=None, consec_max=None,
                 work_days=None, no_holiday=False, no_weekend=False):
        self.name = name
        self.cls = cls                  # ユニット能力 (CLS_ER/CLS_HCU/CLS_WD)
        self.tier = cls                 # 後方互換用エイリアス
        self.is_leader = is_leader or is_er_leader    # リーダー枠可
        self.is_er_leader = is_er_leader              # ERリーダー必須枠可
        self.can_late = can_late        # True=遅出勤務可能
        self.weekly_days = weekly_days  # None=フルタイム, int=週N日
        self.prev_month = prev_month    # ""=通常, "夜"=前月末夜勤, "明"=前月末明け
        self.night_min = night_min      # None=設定値を使用, int=個別指定
        self.night_max = night_max      # None=設定値を使用, int=個別指定 (0=夜勤完全禁止)
        self.consec_max = consec_max    # None=設定値を使用, int=個別の連勤上限
        self.work_days = work_days      # None=全曜日, set of ints (0=月..6=日)
        self.no_holiday = no_holiday    # True=祝日勤務不可
        self.no_weekend = no_weekend    # True=土日勤務不可
        # ICU互換フィールド（3Eでは未使用、デフォルト値を維持）
        self.dedicated = False
        self.short_time = False
        self.night_training = False
        self.night_training_max = None
        self.new_hire = False
        self.new_hire_graduation_day = None
    @property
    def is_parttime(self):
        return self.weekly_days is not None


# ============================================================
# 入力テンプレート生成
# ============================================================

def create_template():
    wb = Workbook()
    # --- スタッフ一覧 ---
    ws1 = wb.active
    ws1.title = "スタッフ一覧"
    headers = ["名前", "Tier", "夜勤専従", "週勤務", "前月末", "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可", "夜勤研修", "研修夜勤回数", "新人", "新人卒業日"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font = FONT_H; c.fill = FH; c.border = BDR; c.alignment = CTR
    for i, (name, tier, ded, weekly, prev) in enumerate(SAMPLE_STAFF, 2):
        ws1.cell(row=i, column=1, value=name).border = BDR
        ws1.cell(row=i, column=2, value=tier).border = BDR
        ws1.cell(row=i, column=2).alignment = CTR
        ws1.cell(row=i, column=3, value=ded).border = BDR
        ws1.cell(row=i, column=3).alignment = CTR
        ws1.cell(row=i, column=4, value=weekly).border = BDR
        ws1.cell(row=i, column=4).alignment = CTR
        ws1.cell(row=i, column=5, value=prev).border = BDR
        ws1.cell(row=i, column=5).alignment = CTR
        ws1.cell(row=i, column=6).border = BDR
        ws1.cell(row=i, column=6).alignment = CTR
        ws1.cell(row=i, column=7).border = BDR
        ws1.cell(row=i, column=7).alignment = CTR
        ws1.cell(row=i, column=8).border = BDR
        ws1.cell(row=i, column=8).alignment = CTR
        ws1.cell(row=i, column=9).border = BDR
        ws1.cell(row=i, column=9).alignment = CTR
        ws1.cell(row=i, column=10).border = BDR
        ws1.cell(row=i, column=10).alignment = CTR

    dv = DataValidation(type="list", formula1='"A,AB,B,C+,C"', allow_blank=False)
    ws1.add_data_validation(dv); dv.add("B2:B100")
    dv2 = DataValidation(type="list", formula1='"○,"', allow_blank=True)
    ws1.add_data_validation(dv2); dv2.add("C2:C100")
    dv3 = DataValidation(type="list", formula1='"夜,明,"', allow_blank=True)
    ws1.add_data_validation(dv3); dv3.add("E2:E100")

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 6
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 8
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 10
    ws1.column_dimensions["I"].width = 12
    ws1.column_dimensions["J"].width = 10

    # 説明
    desc_col = 12
    ws1.cell(row=1, column=desc_col, value="列の説明").font = FONT_H
    descs = [
        "Tier: A=リーダー, AB=サブリーダー, B=メンバー1, C+=C既卒, C=メンバー2",
        "夜勤専従: ○=専従（夜明夜明サイクル対応）",
        "週勤務: 空欄=フルタイム, 数字=週N日勤務（日勤限定・夜勤なし）",
        "前月末: 空欄=通常, 夜=前月末夜勤→当月1日明け, 明=前月末明け→当月1日休",
        "夜勤Min/Max: 空欄=設定シートの値（4±1自動）, 数字=個別指定",
        "連勤Max: 空欄=設定シートの値, 数字=個別の連続勤務上限日数",
        "勤務曜日: 空欄=全曜日, 月火水木金土日から指定 例: 月火木",
        "祝日不可: ○=祝日は勤務しない",
        f"希望休は「勤務希望」シートに入力（1人最大{MAX_REQUEST_DAYS}日）",
    ]
    for i, d in enumerate(descs, 2):
        ws1.cell(row=i, column=desc_col, value=d).font = Font(size=9, color="666666")
    ws1.column_dimensions[get_column_letter(desc_col)].width = 65

    # --- 勤務希望 ---
    ws2 = wb.create_sheet("勤務希望")
    ws2.cell(row=1, column=1, value="勤務希望入力").font = FONT_T
    ws2.cell(row=2, column=1,
             value=f"日/夜/休/研/夜不/休暇/明休（空欄=希望なし）希望は最大{MAX_REQUEST_DAYS}日 ※夜不・休暇・明休は制限外").font = Font(
        size=10, color="666666")
    ws2.cell(row=4, column=1, value="名前").font = FONT_H
    ws2.cell(row=4, column=1).fill = FH; ws2.cell(row=4, column=1).border = BDR
    for d in range(1, 32):
        c = ws2.cell(row=4, column=d + 1, value=d)
        c.font = FONT_H; c.fill = FH; c.alignment = CTR; c.border = BDR
    for i, (name, *_) in enumerate(SAMPLE_STAFF, 5):
        ws2.cell(row=i, column=1, value=name).border = BDR
        for d in range(1, 32):
            ws2.cell(row=i, column=d + 1).border = BDR
            ws2.cell(row=i, column=d + 1).alignment = CTR
    dvr = DataValidation(type="list", formula1='"日,夜,休,研,夜不,休暇,明休"', allow_blank=True)
    ws2.add_data_validation(dvr); dvr.add("B5:AF100")
    ws2.column_dimensions["A"].width = 14
    for d in range(1, 32):
        ws2.column_dimensions[get_column_letter(d + 1)].width = 4

    # --- 設定 ---
    ws3 = wb.create_sheet("設定")
    ws3.cell(row=1, column=1, value="勤務表設定").font = FONT_T
    for col, h in enumerate(["項目", "値", "説明"], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = FONT_H; c.fill = FH; c.border = BDR
    for i, (label, val, desc) in enumerate(SETTINGS_DEF, 4):
        ws3.cell(row=i, column=1, value=label).border = BDR
        ws3.cell(row=i, column=2, value=val).border = BDR
        ws3.cell(row=i, column=2).alignment = CTR
        ws3.cell(row=i, column=3, value=desc).font = Font(size=9, color="666666")
        ws3.cell(row=i, column=3).border = BDR
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 35

    wb.save(INPUT_FILE)
    print(f"✅ テンプレート生成: {INPUT_FILE}")
    print(f"  → 記入後: python3 {os.path.basename(__file__)}")


# ============================================================
# ローカルExcel読み込み
# ============================================================

def _is_truthy(val):
    """○/◯/TRUE/1/任意の非ゼロ数値をTrueとして判定"""
    s = str(val).strip() if val is not None else ""
    if s in ("○", "◯", "TRUE", "True", "1"):
        return True
    try:
        return float(s) != 0
    except (ValueError, TypeError):
        return False

def _to_int(val):
    """数値文字列/float → int、無効値 → None"""
    s = str(val).strip() if val is not None else ""
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None

def _parse_staff_list(rows):
    """行データからStaffリストを生成 (rows: list of lists, header skip済み)
    3E カラム配列 (13列):
      0: 名前
      1: クラス (ER可/HCU可/病棟可)
      2: リーダー可 (◯/空)
      3: ERリーダー可 (◯/空) ※ONなら自動でリーダー可も真扱い
      4: 遅出可 (◯/空)
      5: 週勤務 (パート週N日、空=フルタイム)
      6: 前月末 (夜/明/空)
      7: 夜勤Min
      8: 夜勤Max (0=夜勤完全禁止)
      9: 連勤Max
      10: 勤務曜日 ("月火木" など)
      11: 祝日不可 (◯/空)
      12: 土日不可 (◯/空)
    """
    staff = []
    for row in rows:
        if len(row) < 2:
            continue
        name = str(row[0]).strip()
        if not name:
            break
        cls = str(row[1]).strip()
        if cls not in VALID_CLASSES:
            print(f"⚠ '{name}' クラス '{cls}' 不正 → スキップ")
            continue
        is_ldr   = _is_truthy(row[2]) if len(row) > 2 else False
        is_erl   = _is_truthy(row[3]) if len(row) > 3 else False
        if is_erl and cls != CLS_ER:
            print(f"⚠ '{name}' ERリーダー可はクラス=ER可のみ指定可 → ERリーダー可を無効化")
            is_erl = False
        can_late = _is_truthy(row[4]) if len(row) > 4 else False
        weekly   = _to_int(row[5])    if len(row) > 5 else None
        prev     = str(row[6]).strip() if len(row) > 6 else ""
        if prev not in ("夜", "明", ""):
            print(f"⚠ '{name}' 前月末 '{prev}' 不正（夜/明/空欄）→ 無視")
            prev = ""
        n_min = _to_int(row[7]) if len(row) > 7 else None
        n_max = _to_int(row[8]) if len(row) > 8 else None
        c_max = _to_int(row[9]) if len(row) > 9 else None
        wd_str = str(row[10]).strip() if len(row) > 10 else ""
        wd_map = {"月":0, "火":1, "水":2, "木":3, "金":4, "土":5, "日":6}
        work_days = None
        if wd_str:
            work_days = set()
            for ch in wd_str:
                if ch in wd_map:
                    work_days.add(wd_map[ch])
            if not work_days:
                work_days = None
        no_hol = _is_truthy(row[11]) if len(row) > 11 else False
        no_we  = _is_truthy(row[12]) if len(row) > 12 else False
        staff.append(Staff(name, cls, is_leader=is_ldr, is_er_leader=is_erl,
                           can_late=can_late, weekly_days=weekly, prev_month=prev,
                           night_min=n_min, night_max=n_max, consec_max=c_max,
                           work_days=work_days, no_holiday=no_hol, no_weekend=no_we))
    if not staff:
        print("✗ スタッフが0人です")
        sys.exit(1)
    return staff


def _parse_settings(rows):
    """設定シートの行データからdict生成 (rows: list of lists, row index 3〜)"""
    settings = {}
    for i, key in enumerate(SETTINGS_KEYS):
        row_idx = i  # rows[0] = 設定row4
        if row_idx < len(rows) and len(rows[row_idx]) > 1:
            val = str(rows[row_idx][1]).strip()
            if key == "holidays":
                settings[key] = val
            else:
                try:
                    settings[key] = int(val) if val else None
                except ValueError:
                    settings[key] = None
        else:
            settings[key] = None
    return settings


def _parse_holidays(holiday_str):
    """カンマ区切り文字列 → 日付setを返す（手動追加分）"""
    if not holiday_str:
        return set()
    result = set()
    for part in str(holiday_str).split(","):
        part = part.strip()
        if part.isdigit():
            result.add(int(part))
    return result


def _get_holidays_and_days_off(year, month):
    """日本の祝日を自動取得し、土日祝の公休日数を返す
    Returns:
        holidays: set of day numbers that are national holidays (平日祝日のみ)
        weekends: set of day numbers that are 土日
        public_off: 公休日数 (土日 + 平日祝日)
    """
    num_days = calendar.monthrange(year, month)[1]
    weekends = set()
    holidays = set()
    for day in range(1, num_days + 1):
        dt = date(year, month, day)
        wd = dt.weekday()
        if wd >= 5:  # 土日
            weekends.add(day)
        if jpholiday.is_holiday(dt) and wd < 5:  # 平日の祝日のみ
            holidays.add(day)
    public_off = len(weekends) + len(holidays)
    return holidays, weekends, public_off


def _parse_requests(rows, staff_names, num_days):
    """勤務希望シート → {name: {day: shift}} （上限チェック付き）
    夜不・休暇はシフト希望とは別枠（日数制限に含めない）"""
    requests = {}
    for row in rows:
        if not row:
            continue
        name = str(row[0]).strip()
        if not name:
            break
        if name not in staff_names:
            print(f"⚠ 希望: '{name}' はスタッフ一覧に存在しません → スキップ")
            continue
        reqs = {}
        extra = {}  # 夜不・休暇・明休は別カウント（日数制限外）
        for d in range(1, min(num_days + 1, len(row))):
            val = str(row[d]).strip() if d < len(row) else ""
            if val in ("夜不", "休暇", "明休"):
                extra[d] = val
            elif val in VALID_REQUEST:
                reqs[d] = val
        if len(reqs) > MAX_REQUEST_DAYS:
            print(f"⚠ {name}: 希望{len(reqs)}日 > 上限{MAX_REQUEST_DAYS}日 → 先頭{MAX_REQUEST_DAYS}日のみ採用")
            kept = dict(list(sorted(reqs.items()))[:MAX_REQUEST_DAYS])
            reqs = kept
        reqs.update(extra)  # 夜不・休暇を合流（制限外）
        if reqs:
            requests[name] = reqs
    return requests


def load_input():
    if not os.path.exists(INPUT_FILE):
        print(f"✗ {INPUT_FILE} が見つかりません → --init でテンプレート生成")
        sys.exit(1)
    wb = load_workbook(INPUT_FILE, data_only=True)

    # スタッフ
    ws1 = wb["スタッフ一覧"]
    staff_rows = []
    r = 2
    while True:
        name = ws1.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            break
        staff_rows.append([
            ("" if ws1.cell(row=r, column=c).value is None else ws1.cell(row=r, column=c).value) for c in range(1, 11)
        ])
        r += 1
    staff_list = _parse_staff_list(staff_rows)

    # 設定
    ws3 = wb["設定"]
    setting_rows = []
    for r in range(4, 4 + len(SETTINGS_KEYS)):
        setting_rows.append([("" if ws3.cell(row=r, column=c).value is None else ws3.cell(row=r, column=c).value) for c in range(1, 4)])
    settings = _parse_settings(setting_rows)

    year = settings.get("year") or 2026
    month = settings.get("month") or 5
    num_days = calendar.monthrange(year, month)[1]

    # 希望
    ws2 = wb["勤務希望"]
    req_rows = []
    r = 5
    while True:
        name = ws2.cell(row=r, column=1).value
        if name is None or str(name).strip() == "":
            break
        row_data = [("" if ws2.cell(row=r, column=c).value is None else ws2.cell(row=r, column=c).value) for c in range(1, num_days + 2)]
        req_rows.append(row_data)
        r += 1
    staff_names = [s.name for s in staff_list]
    requests = _parse_requests(req_rows, staff_names, num_days)

    wb.close()
    return staff_list, requests, settings


# ============================================================
# Google スプレッドシート
# ============================================================

def _get_gc():
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("✗ pip3 install gspread google-auth")
        sys.exit(1)
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    # 1) Streamlit Secrets（Cloud用）
    try:
        import streamlit as _st
        if "gcp_service_account" in _st.secrets:
            creds = Credentials.from_service_account_info(
                dict(_st.secrets["gcp_service_account"]), scopes=scopes)
            return gspread.authorize(creds)
    except Exception:
        pass
    # 2) ローカルファイル
    if os.path.exists(CRED_FILE):
        creds = Credentials.from_service_account_file(CRED_FILE, scopes=scopes)
        return gspread.authorize(creds)
    print(f"✗ Google認証情報が見つかりません")
    print("\nセットアップ (ローカル):")
    print("1. Google Cloud Console → プロジェクト作成")
    print("2. Google Sheets API + Google Drive API を有効化")
    print("3. サービスアカウント作成 → JSON鍵ダウンロード")
    print(f"4. {CRED_FILE} に配置")
    print("5. スプレッドシートをサービスアカウントのメールに共有（編集者）")
    print("\nセットアップ (Streamlit Cloud):")
    print("  .streamlit/secrets.toml に [gcp_service_account] セクションを追加")
    sys.exit(1)


def _parse_gsheet_id(url_or_id):
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_or_id)
    if m: return m.group(1)
    if re.match(r"^[a-zA-Z0-9_-]+$", url_or_id): return url_or_id
    print(f"✗ 不正なURL/ID: {url_or_id}"); sys.exit(1)


def create_gsheet_template():
    gc = _get_gc()
    print("スプレッドシート作成中...")
    sh = gc.create("勤務表_入力")

    # --- スタッフ一覧 ---
    ws1 = sh.sheet1
    ws1.update_title("スタッフ一覧")
    ws1.update("A1", [["名前", "Tier", "夜勤専従", "週勤務", "前月末",
                        "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可",
                        "夜勤研修", "研修夜勤回数", "", "列の説明"]])
    rows = [[n, t, d, w, p, "", "", "", "", "", "", "", "", ""] for n, t, d, w, p in SAMPLE_STAFF]
    descs = [
        "Tier: A=リーダー AB=サブリーダー B=メンバー1 C+=C既卒 C=メンバー2",
        "夜勤専従: ○=専従", "週勤務: 空欄=フルタイム 数字=週N日(日勤限定)",
        "前月末: 空欄=通常 夜=前月末夜勤 明=前月末明け",
        "夜勤Min/Max: 空欄=設定値 数字=個別指定",
        "連勤Max: 空欄=設定値 数字=個別の連続勤務上限",
        "勤務曜日: 空欄=全曜日 例:月火木",
        "祝日不可: ○=祝日は勤務しない",
        "土日不可: ○=土日は勤務しない",
        "夜勤研修: ○=研修中(3人目枠で夜勤配置)",
        "研修夜勤回数: 月間上限(空欄=制限なし)",
        f"希望は「勤務希望」シートに(最大{MAX_REQUEST_DAYS}日)",
    ]
    for i, r in enumerate(rows):
        r.append(descs[i] if i < len(descs) else "")
    ws1.update("A2", rows)
    ws1.format("A1:L1", {"textFormat": {"bold": True},
               "backgroundColor": {"red": 0.85, "green": 0.88, "blue": 0.95}})

    reqs_batch = [
        {"setDataValidation": {"range": {"sheetId": ws1.id,
         "startRowIndex": 1, "endRowIndex": 100, "startColumnIndex": 1, "endColumnIndex": 2},
         "rule": {"condition": {"type": "ONE_OF_LIST",
         "values": [{"userEnteredValue": v} for v in ["A", "AB", "B", "C+", "C"]]},
         "showCustomUi": True, "strict": True}}},
        {"setDataValidation": {"range": {"sheetId": ws1.id,
         "startRowIndex": 1, "endRowIndex": 100, "startColumnIndex": 2, "endColumnIndex": 3},
         "rule": {"condition": {"type": "ONE_OF_LIST",
         "values": [{"userEnteredValue": v} for v in ["○", ""]]},
         "showCustomUi": True}}},
        {"setDataValidation": {"range": {"sheetId": ws1.id,
         "startRowIndex": 1, "endRowIndex": 100, "startColumnIndex": 4, "endColumnIndex": 5},
         "rule": {"condition": {"type": "ONE_OF_LIST",
         "values": [{"userEnteredValue": v} for v in ["夜", "明", ""]]},
         "showCustomUi": True}}},
    ]
    sh.batch_update({"requests": reqs_batch})

    # --- 勤務希望 ---
    ws2 = sh.add_worksheet("勤務希望", rows=100, cols=33)
    ws2.update("A1", [["勤務希望入力"]])
    ws2.update("A2", [[f"日/夜/休/研/夜不/休暇/明休（希望は最大{MAX_REQUEST_DAYS}日 ※夜不・休暇・明休は制限外）"]])
    ws2.update("A4", [["名前"] + [str(d) for d in range(1, 32)]])
    ws2.update("A5", [[s[0]] for s in SAMPLE_STAFF])
    ws2.format("A4:AF4", {"textFormat": {"bold": True},
               "backgroundColor": {"red": 0.85, "green": 0.88, "blue": 0.95},
               "horizontalAlignment": "CENTER"})
    sh.batch_update({"requests": [
        {"setDataValidation": {"range": {"sheetId": ws2.id,
         "startRowIndex": 4, "endRowIndex": 100, "startColumnIndex": 1, "endColumnIndex": 32},
         "rule": {"condition": {"type": "ONE_OF_LIST",
         "values": [{"userEnteredValue": v} for v in ["日", "夜", "休", "研", "夜不", "休暇", "明休"]]},
         "showCustomUi": True}}}
    ]})

    # --- 設定 ---
    ws3 = sh.add_worksheet("設定", rows=20, cols=3)
    ws3.update("A1", [["勤務表設定"]])
    ws3.update("A3", [["項目", "値", "説明"]])
    ws3.update("A4", [[l, v, d] for l, v, d in SETTINGS_DEF])
    ws3.format("A3:C3", {"textFormat": {"bold": True},
               "backgroundColor": {"red": 0.85, "green": 0.88, "blue": 0.95}})

    sh.share("", perm_type="anyone", role="writer")
    print(f"\n✅ スプレッドシート作成: {sh.url}")
    print(f"   → 記入後: python3 shift_scheduler.py --gsheet '{sh.url}'")


def load_gsheet(url_or_id):
    gc = _get_gc()
    sid = _parse_gsheet_id(url_or_id)
    print("読み込み中...")
    sh = gc.open_by_key(sid)

    vals1 = sh.worksheet("スタッフ一覧").get_all_values()
    staff_list = _parse_staff_list(vals1[1:])  # skip header
    print(f"  スタッフ: {len(staff_list)}人")

    vals3 = sh.worksheet("設定").get_all_values()
    settings = _parse_settings(vals3[3:])  # row4以降

    year = settings.get("year") or 2026
    month = settings.get("month") or 5
    num_days = calendar.monthrange(year, month)[1]

    vals2 = sh.worksheet("勤務希望").get_all_values()
    staff_names = [s.name for s in staff_list]
    requests = _parse_requests(vals2[4:], staff_names, num_days)
    print(f"  希望: {len(requests)}人")

    return staff_list, requests, settings, sh


def write_gsheet_result(sh, results):
    """results: list of result dicts"""
    for result in results:
        _write_gsheet_one(sh, result)
    print(f"\n✅ スプレッドシート更新 ({len(results)}パターン): {sh.url}")


def _write_gsheet_one(sh, result):
    schedule = result["schedule"]
    names    = result["names"]
    tiers    = result["tiers"]
    num_days = result["num_days"]
    year     = result["year"]
    month    = result["month"]
    missed   = result.get("missed_requests", {})
    pat_num  = result.get("pattern_num", 1)

    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    first_wd = date(year, month, 1).weekday()

    sheet_title = f"パターン{pat_num}" if pat_num > 1 else "勤務表"
    try:
        sh.del_worksheet(sh.worksheet(sheet_title))
    except Exception:
        pass

    ws = sh.add_worksheet(sheet_title, rows=len(names) + 30, cols=num_days + 10)
    all_data = []
    all_data.append([f"{year}年{month}月 勤務表"])
    all_data.append([])
    all_data.append(["名前", "Tier"] + [str(d+1) for d in range(num_days)] + ["日", "夜", "明", "休", "研", "暇", "計"])
    all_data.append(["", ""] + [weekdays_jp[(first_wd+d)%7] for d in range(num_days)])

    current_tier = None
    for s in names:
        t = tiers[s]
        if t != current_tier:
            current_tier = t
            lbl = {"A": "A（リーダー）", "AB": "AB（サブリーダー）",
                   "B": "B（メンバー1）", "C+": "C+（C既卒）",
                   "C": "C（メンバー2）"}.get(t, t)
            all_data.append([f"── {lbl} ──"])
        c = {sh: schedule[s].count(sh) for sh in SHIFTS}
        all_data.append([s, t] + schedule[s] + [c[D], c[N], c[A], c[O], c[R], c[V], c[D]+c[N]])

    all_data.append([])
    all_data.append(["日別集計"])
    for label, shift in [("日勤計", D), ("夜勤計", N), ("明け計", A), ("休み計", O), ("研修計", R), ("休暇計", V)]:
        row = [label, ""] + [sum(1 for s in names if schedule[s][d] == shift)
                              for d in range(num_days)]
        all_data.append(row)

    all_data.append([])
    all_data.append(["夜勤ペア"])
    pair = ["メンバー", ""] + ["/".join(s for s in names if schedule[s][d] == N)
                                for d in range(num_days)]
    all_data.append(pair)

    if missed:
        all_data.append([])
        all_data.append(["⚠ 未達希望"])
        for s, days in missed.items():
            all_data.append([s, "", ", ".join(f"{d}日" for d in sorted(days))])

    ws.update("A1", all_data)

    # 書式（ヘッダー + 色付け）
    reqs = []
    reqs.append({"repeatCell": {
        "range": {"sheetId": ws.id, "startRowIndex": 2, "endRowIndex": 4,
                  "startColumnIndex": 0, "endColumnIndex": num_days + 7},
        "cell": {"userEnteredFormat": {"textFormat": {"bold": True},
                 "backgroundColor": {"red": 0.85, "green": 0.88, "blue": 0.95},
                 "horizontalAlignment": "CENTER"}},
        "fields": "userEnteredFormat(textFormat,backgroundColor,horizontalAlignment)"}})

    cmap = {D: {"red":1,"green":1,"blue":1}, N: {"red":0.27,"green":0.45,"blue":0.77},
            A: {"red":1,"green":0.95,"blue":0.8}, O: {"red":0.89,"green":0.94,"blue":0.85},
            R: {"red":0.91,"green":0.84,"blue":0.96},
            V: {"red":0.96,"green":0.80,"blue":0.80}}
    row_idx = 4
    current_tier = None
    for s in names:
        t = tiers[s]
        if t != current_tier:
            current_tier = t
            row_idx += 1
        for d in range(num_days):
            shift = schedule[s][d]
            if shift in cmap:
                fmt = {"backgroundColor": cmap[shift], "horizontalAlignment": "CENTER"}
                flds = "userEnteredFormat(backgroundColor,horizontalAlignment)"
                if shift == N:
                    fmt["textFormat"] = {"foregroundColorStyle":
                                         {"rgbColor": {"red":1,"green":1,"blue":1}}, "bold": True}
                    flds = "userEnteredFormat(backgroundColor,horizontalAlignment,textFormat)"
                reqs.append({"repeatCell": {
                    "range": {"sheetId": ws.id, "startRowIndex": row_idx,
                              "endRowIndex": row_idx + 1,
                              "startColumnIndex": d + 2, "endColumnIndex": d + 3},
                    "cell": {"userEnteredFormat": fmt}, "fields": flds}})
        row_idx += 1

    for i in range(0, len(reqs), 100):
        sh.batch_update({"requests": reqs[i:i+100]})


# ============================================================
# ソルバー
# ============================================================

def build_and_solve(staff_list, requests, settings, num_patterns=1,
                    night_hours=16, night_72h_mode="none", op_rules=None,
                    enabled_shifts=None):
    """勤務表を num_patterns パターン生成して返す (list of result dict)"""
    year     = settings.get("year") or 2026
    month    = settings.get("month") or 5
    num_days = calendar.monthrange(year, month)[1]
    days     = list(range(num_days))

    def S(k, d):
        v = settings.get(k)
        return v if v is not None else d
    min_day     = S("min_day_staff", 5)
    min_day_excl = S("min_day_staff_excl_new", 4)  # 新人除く日勤最低人数
    min_day_hd  = S("min_day_staff_hd", 8)   # 土日祝の日勤系合計下限
    max_day_hd  = S("max_day_staff_hd", 8)   # 土日祝の日勤系合計上限
    max_day_wd  = S("max_day_staff_wd", 14)  # 平日の日勤系合計上限
    night_count = S("night_staff_count", 2)
    max_n_reg   = S("max_night_regular", 5)
    pref_n_reg  = S("pref_night_regular", 4)
    max_n_ded   = S("max_night_dedicated", 10)
    pref_n_ded  = S("pref_night_dedicated", 9)
    max_consec  = S("max_consecutive", 5)
    pref_consec = S("pref_consecutive", 4)
    time_limit  = S("solver_time_limit", 120)
    # 祝日・公休自動取得
    jp_holidays, weekends, auto_public_off = _get_holidays_and_days_off(year, month)
    manual_holidays = _parse_holidays(settings.get("holidays", ""))
    holidays = jp_holidays | manual_holidays  # 自動+手動
    # 公休日数: 設定値があればそちらを使用、なければ自動計算
    po_override = settings.get("public_off_override")
    if po_override is not None and po_override != "":
        public_off = int(po_override)
    else:
        public_off = auto_public_off

    names     = [s.name for s in staff_list]
    classes   = {s.name: s.cls for s in staff_list}
    tiers     = classes  # 後方互換エイリアス
    weekly    = {s.name: s.weekly_days for s in staff_list}
    prev_m    = {s.name: s.prev_month for s in staff_list}
    night_min = {s.name: s.night_min for s in staff_list}
    night_max = {s.name: s.night_max for s in staff_list}
    can_late  = {s.name: s.can_late for s in staff_list}
    is_leader_map    = {s.name: s.is_leader for s in staff_list}
    is_er_leader_map = {s.name: s.is_er_leader for s in staff_list}
    dedicated = {s.name: False for s in staff_list}  # 3Eは専従なし（ICU互換）
    # 3E 設定値読み込み
    min_ward_wd = S("min_ward_wd", 4)
    min_hcu_wd  = S("min_hcu_wd",  2)
    min_er_wd   = S("min_er_wd",   3)
    min_ward_hd = S("min_ward_hd", 4)
    min_hcu_hd  = S("min_hcu_hd",  2)
    max_n_reg   = S("max_night", 5)
    pref_n_reg  = S("pref_night", 4)
    # デフォルト夜勤回数（フルタイム・夜勤禁止でないスタッフ）
    _default_min = max(0, pref_n_reg - 1)
    _default_max = pref_n_reg + 1
    for _s in staff_list:
        if _s.weekly_days is not None:
            continue
        if _s.night_max == 0:
            continue
        if night_min[_s.name] is None:
            night_min[_s.name] = _default_min
        if night_max[_s.name] is None:
            night_max[_s.name] = _default_max
    consec_max_ind = {s.name: s.consec_max for s in staff_list}
    work_days_map  = {s.name: s.work_days for s in staff_list}
    no_holiday_map = {s.name: s.no_holiday for s in staff_list}
    no_weekend_map = {s.name: s.no_weekend for s in staff_list}

    fulltime = [n for n in names if weekly[n] is None]
    parttime = [n for n in names if weekly[n] is not None]
    fwd = date(year, month, 1).weekday()  # 月初の曜日 (0=月)

    # クラス別プール（ユニット配置能力）
    er_staff    = [n for n in names if classes[n] == CLS_ER]   # ER/HCU/病棟配置可
    hcu_staff   = [n for n in names if classes[n] == CLS_HCU]  # HCU/病棟配置可
    wd_staff    = [n for n in names if classes[n] == CLS_WD]   # 病棟のみ

    # ERリーダー枠（平日ER必須1名）
    erl_staff   = [n for n in names if is_er_leader_map[n]]
    # リーダー枠（病棟・HCU・共リーダー）
    leader_pool = [n for n in names if is_leader_map[n]]
    # ER配置可 = クラス=ER可 のみ（HCU/病棟クラスはER配置不可）
    er_capable  = er_staff
    # HCU配置可 = ER可 + HCU可
    hcu_capable = er_staff + hcu_staff
    # 遅出可能スタッフ
    late_pool   = [n for n in names if can_late[n]]

    night_count = 4  # 3E: 夜勤は病棟2+HCU1+リーダー1=4名固定

    print(f"\n{'='*55}")
    print(f"  3E勤務表: {year}年{month}月 ({num_days}日)")
    print(f"{'='*55}")
    print(f"スタッフ: {len(names)}人 "
          f"(ER可:{len(er_staff)} HCU可:{len(hcu_staff)} 病棟可:{len(wd_staff)}) "
          f"／ リーダー可:{len(leader_pool)} ERリーダー可:{len(erl_staff)}")
    if parttime:
        pt_str = ", ".join(f"{n}(週{weekly[n]})" for n in parttime)
        print(f"パートタイム: {pt_str}")
    if late_pool:
        print(f"遅出可能: {', '.join(late_pool)}")
    carry = [(n, prev_m[n]) for n in names if prev_m[n]]
    if carry:
        print(f"前月繰越: {', '.join(f'{n}({v})' for n, v in carry)}")
    if po_override is not None and po_override != "":
        print(f"公休日数: {public_off}日 (手動設定)  ※自動計算={auto_public_off}日")
    else:
        print(f"公休日数: {public_off}日 (土日{len(weekends)}日 + 平日祝日{len(holidays)}日)")
    if holidays:
        hol_names = {d.day: name for d, name in jpholiday.month_holidays(year, month)}
        hol_str = ", ".join(f"{d}日({hol_names.get(d, '祝')})" for d in sorted(holidays))
        print(f"祝日: {hol_str}")
    if requests:
        print(f"希望: {len(requests)}人")
        for s, rq in requests.items():
            sn = {D:"日", N:"夜", O:"休", L:"遅", "夜不":"夜不", "休暇":"休暇", "明休":"明休", "遅希":"遅希"}
            print(f"  {s}: {', '.join(f'{d}日={sn.get(t,t)}' for d,t in sorted(rq.items()))}")
    if num_patterns > 1:
        print(f"\n生成パターン数: {num_patterns}")

    total_leader_slots = len(leader_pool) * max_n_reg
    if total_leader_slots < num_days:
        print(f"\n⚠ リーダー夜勤枠({total_leader_slots}) < {num_days}日")
    else:
        print(f"\n✓ リーダー夜勤枠({total_leader_slots}) >= {num_days}日")

    # ============================================================
    prob = pulp.LpProblem("NurseShift", pulp.LpMinimize)

    # --- 変数削減: スタッフ属性に応じて不要なシフト変数を除外 ---
    _excluded = set()  # (s, t) の組: このスタッフにこのシフトは不要

    # 3E: 使用しない拡張シフトを除外（デフォルト: R/I/E/ST/LD/SN 無効）
    _3e_inactive = [R, I, E, ST, LD, SN]
    if enabled_shifts is not None:
        _sym_to_shift = {"準": SN, "早": E, "遅": L, "長": LD, "短": ST}
        for sym, shift_type in _sym_to_shift.items():
            if sym not in enabled_shifts:
                for s in names:
                    _excluded.add((s, shift_type))
    else:
        # デフォルト: 3E で使わないシフトを全員から除外
        for t in _3e_inactive:
            for s in names:
                _excluded.add((s, t))

    # 遅出(L): can_late=False のスタッフは遅出不可
    for s in names:
        if not can_late.get(s):
            _excluded.add((s, L))

    # パートタイム: N 不可
    for s in parttime:
        _excluded.add((s, N))

    # STは全員除外（3Eでは時短シフトなし）
    for s in names:
        _excluded.add((s, ST))

    # 有効な(s,d,t)のみ変数を生成
    _valid_keys = [(s, d, t) for s in names for d in days for t in SHIFTS
                   if (s, t) not in _excluded]
    x_real = pulp.LpVariable.dicts("x", _valid_keys, cat=pulp.LpBinary)

    # 除外された変数はゼロ定数で代替（制約式をそのまま使えるように）
    _zero = pulp.LpAffineExpression()  # 常に0
    class _XProxy:
        """存在しないキーは0を返すプロキシ"""
        def __init__(self, real):
            self._r = real
        def __getitem__(self, key):
            return self._r.get(key, _zero)
    x = _XProxy(x_real)

    n_total = len(_valid_keys)
    n_saved = len(names) * num_days * len(SHIFTS) - n_total
    if n_saved > 0:
        print(f"  変数削減: {n_saved}個除外 → {n_total}個（{n_saved/(n_saved+n_total)*100:.0f}%削減）")

    # --- ハード制約 ---
    for s in names:
        for d in days:
            prob += pulp.lpSum(x[s, d, t] for t in SHIFTS) == 1
    # 夜(N)→明(A) は全員共通
    for s in names:
        for d in days[:-1]:
            prob += x[s, d, N] <= x[s, d+1, A]
    # 短夜勤(SN, 12h)→翌日: 明け or 休（日勤系は不可）
    for s in names:
        for d in days[:-1]:
            prob += x[s, d, SN] <= x[s, d+1, A] + x[s, d+1, O]
    # 長日勤(LD, 12h)→翌日: 休推奨（ソフト: 翌日勤務にペナルティで対応）
    # ハードには強制しないが、明け(A)は禁止（夜勤していないのでAは不自然）
    for s in names:
        for d in days[:-1]:
            prob += x[s, d, LD] + x[s, d+1, A] <= 1
    # 明→休 は通常スタッフ（専従は明→夜も可）
    for s in names:
        if dedicated.get(s):
            # 専従: 明の翌日は 夜 or 休 のみ（日勤・研修不可）
            for d in days[:-1]:
                prob += x[s, d, A] <= x[s, d+1, O] + x[s, d+1, N] + x[s, d+1, SN]
        else:
            # 通常: 明の翌日は休のみ
            for d in days[:-1]:
                prob += x[s, d, A] <= x[s, d+1, O]
    # 専従のみ: 夜明は連続2回まで（夜明夜明 OK、夜明夜明夜 NG）
    # → 5日連続で N,A,N,A,N は禁止 = 5日中のN合計 ≤ 2
    # ★前月繰越の夜勤も含めて判定する
    for s in fulltime:
        if not dedicated[s]:
            continue
        # 前月末の夜勤数を算出（prev_month="夜"→前日N, "明"→2日前N）
        prev_nights = {}  # {relative_day: 1} (day 0 = 月初1日)
        if prev_m[s] == "夜":
            prev_nights[-1] = 1  # 前月末日がN
        elif prev_m[s] == "明":
            prev_nights[-2] = 1  # 前月末前日がN
        # 月内の5日ウィンドウ
        for d in range(num_days - 4):
            prob += pulp.lpSum(x[s, d+i, N] for i in range(5)) <= 2
        # 前月繰越を含むウィンドウ（月初数日）
        if prev_nights:
            for start in range(-max(1, max(-k for k in prev_nights)), num_days - 4):
                if start >= 0:
                    break  # 月内ウィンドウは上で処理済み
                n_count = sum(prev_nights.get(start + i, 0) for i in range(5)
                              if (start + i) < 0)
                month_sum = pulp.lpSum(x[s, start+i, N] for i in range(5)
                                       if 0 <= start + i < num_days)
                prob += month_sum <= 2 - n_count
    # 専従のみ: 連続夜勤(夜明夜明)の後は最低1日休み（ハード制約）
    # 夜明夜明の後 = d,d+1,d+2,d+3 が N,A,N,A なら d+4 は O 必須
    # ★前月繰越も含む
    for s in fulltime:
        if not dedicated[s]:
            continue
        # 月内
        for d in range(num_days - 4):
            prob += x[s, d, N] + x[s, d+2, N] - 1 <= x[s, d+4, O]
        # 前月繰越: 前月末N + 月初のNで2連夜 → 休み必須
        if prev_m[s] == "夜":
            # 前月末N(day-1) + day1にN → day3はO必須
            if num_days > 3:
                prob += x[s, 1, N] <= x[s, 3, O]
        elif prev_m[s] == "明":
            # 前月末前日N(day-2) + day0にN → day2はO必須
            if num_days > 2:
                prob += x[s, 0, N] <= x[s, 2, O]

    # 休暇(V)日の事前計算（前月繰越・曜日制限で参照するため、ここで先に計算）
    vacation_days = {}  # {staff_name: set of day_indices}
    for s, reqs_s in requests.items():
        if s not in names:
            continue
        for day_num, shift_type in reqs_s.items():
            if shift_type == "休暇" and 1 <= day_num <= num_days:
                vacation_days.setdefault(s, set()).add(day_num - 1)

    # 前月繰越
    for s in names:
        s_vac_carry = vacation_days.get(s, set())
        if prev_m[s] == "夜":
            if 0 not in s_vac_carry:
                prob += x[s, 0, A] == 1
            # 通常スタッフは2日目=休、専従は夜or休
            if not dedicated.get(s) and num_days > 1 and 1 not in s_vac_carry:
                prob += x[s, 1, O] == 1
        elif prev_m[s] == "明":
            if dedicated.get(s):
                if 0 not in s_vac_carry:
                    prob += x[s, 0, A] == 0  # 専従: 明の翌日は夜or休
            else:
                if 0 not in s_vac_carry:
                    prob += x[s, 0, O] == 1  # 通常: 明の翌日は休
                    prob += x[s, 0, A] == 0
        else:
            if 0 not in s_vac_carry:
                prob += x[s, 0, A] == 0
        for d in days[1:]:
            prob += x[s, d, A] <= x[s, d-1, N]
    # 通常スタッフ: 休→夜(N)/短夜勤(SN) 禁止（専従は休→夜OK）
    for s in names:
        if dedicated.get(s):
            continue
        for d in days[:-1]:
            prob += x[s, d, O] + x[s, d+1, N] <= 1
            prob += x[s, d, O] + x[s, d+1, SN] <= 1
    # パートタイム: N 禁止は変数削減で処理済み
    # 3E では専従スタッフなし
    # 夜勤専従代替不要（旧: 希望がない限り日勤(D)なし）
    for s in fulltime:
        if dedicated[s]:
            day_reqs = {d for d, t in requests.get(s, {}).items() if t == D}
            for d in days:
                if (d + 1) not in day_reqs:
                    prob += x[s, d, D] == 0
    # 時短スタッフ: D/N/SN/LDは変数削減で除外済み。早出・遅出は許可。

    # ================================================================
    # 3E: 遅出制約
    # ================================================================
    # 遅出(L) は1日1名固定（can_late=Trueのスタッフのみ）
    late_short = {}
    for d in days:
        late_sum = pulp.lpSum(x[s, d, L] for s in late_pool) if late_pool else pulp.lpSum([])
        # ハード: 遅出は必ず1名
        if late_pool:
            prob += late_sum == 1
        # 遅出の翌日: 休(O) か 夜勤(N) のみ
        if d < num_days - 1:
            for s in late_pool:
                prob += x[s, d, L] + pulp.lpSum(
                    x[s, d+1, t] for t in SHIFTS if t not in [O, N]
                ) <= 1

    # ================================================================
    # 3E: ユニット割り当て変数
    # ================================================================
    # ud[s, d, unit]: 日勤(D)スタッフのユニット割り当て
    # un[s, d, unit]: 夜勤(N)スタッフのユニット割り当て
    ud_keys = [(s, d, u) for s in names for d in days for u in DAY_UNITS]
    un_keys = [(s, d, u) for s in names for d in days for u in NIGHT_UNITS]
    ud = pulp.LpVariable.dicts("ud", ud_keys, cat=pulp.LpBinary)
    un = pulp.LpVariable.dicts("un", un_keys, cat=pulp.LpBinary)

    # 日勤者はユニット割り当てを1つ持つ (遅出はユニット外・独立)
    for s in names:
        for d in days:
            prob += pulp.lpSum(ud[s, d, u] for u in DAY_UNITS) == x[s, d, D]

    # 夜勤者はユニット割り当てを1つ持つ
    for s in names:
        for d in days:
            prob += pulp.lpSum(un[s, d, u] for u in NIGHT_UNITS) == x[s, d, N]

    # ── ユニット配置制約（クラス + リーダーフラグ分離） ──
    # 1) リーダー枠: is_leader=True のみ配置可
    non_leader = [s for s in names if not is_leader_map[s]]
    for s in non_leader:
        for d in days:
            prob += ud[s, d, UNIT_LEAD] == 0
            prob += un[s, d, UNIT_LEAD] == 0

    # 2) ER配置: クラス=ER可 のみ（HCU可/病棟可はER不可）
    non_er = [s for s in names if classes[s] != CLS_ER]
    for s in non_er:
        for d in days:
            prob += ud[s, d, UNIT_ER] == 0

    # 3) HCU配置: クラス=ER可 or HCU可（病棟可はHCU不可）
    wd_only = [s for s in names if classes[s] == CLS_WD]
    for s in wd_only:
        for d in days:
            prob += ud[s, d, UNIT_HCU] == 0
            prob += un[s, d, UNIT_HCU] == 0

    # ================================================================
    # 3E: 日次配置人数制約（平日 vs 休日）
    # ================================================================
    # 休日判定: 土日 or 祝日
    rest_days_set = weekends | holidays  # 土日 + 祝日（day number 1-indexed）

    day_short = {}  # ソフト不足変数（コンソール表示用）
    for d in days:
        day_num = d + 1
        is_rest = day_num in rest_days_set

        ward_sum  = pulp.lpSum(ud[s, d, UNIT_WD]   for s in names)
        hcu_sum   = pulp.lpSum(ud[s, d, UNIT_HCU]  for s in names)
        er_sum    = pulp.lpSum(ud[s, d, UNIT_ER]   for s in names)
        lead_sum  = pulp.lpSum(ud[s, d, UNIT_LEAD] for s in names)

        # 共通リーダーは毎日1名（平日・休日とも）
        prob += lead_sum == 1

        if is_rest:
            # 休日: 病棟≥min_ward_hd, HCU≥min_hcu_hd, ER=0
            ds_w = pulp.LpVariable(f"ds_ward_hd_{d}", lowBound=0)
            prob += ds_w >= min_ward_hd - ward_sum
            day_short[("ward_hd", d)] = ds_w
            prob += ward_sum >= max(1, min_ward_hd - 1)

            ds_h = pulp.LpVariable(f"ds_hcu_hd_{d}", lowBound=0)
            prob += ds_h >= min_hcu_hd - hcu_sum
            day_short[("hcu_hd", d)] = ds_h
            prob += hcu_sum >= max(1, min_hcu_hd - 1)

            prob += er_sum == 0  # 休日はER配置なし

            # 休日日勤系合計: min_day_hd ≤ sum ≤ max_day_hd（下限/上限両方）
            rest_day_sum = pulp.lpSum(x[s, d, t] for s in names for t in DAY_SHIFTS)
            prob += rest_day_sum >= min_day_hd
            prob += rest_day_sum <= max_day_hd
        # 平日日勤系合計 ≤ max_day_wd（偏り防止）
        if not is_rest:
            wd_day_sum = pulp.lpSum(x[s, d, t] for s in names for t in DAY_SHIFTS)
            prob += wd_day_sum <= max_day_wd
        else:
            # 平日: 病棟≥min_ward_wd, HCU≥min_hcu_wd, ER≥min_er_wd
            ds_w = pulp.LpVariable(f"ds_ward_wd_{d}", lowBound=0)
            prob += ds_w >= min_ward_wd - ward_sum
            day_short[("ward_wd", d)] = ds_w
            prob += ward_sum >= max(1, min_ward_wd - 1)

            ds_h = pulp.LpVariable(f"ds_hcu_wd_{d}", lowBound=0)
            prob += ds_h >= min_hcu_wd - hcu_sum
            day_short[("hcu_wd", d)] = ds_h
            prob += hcu_sum >= max(1, min_hcu_wd - 1)

            ds_e = pulp.LpVariable(f"ds_er_wd_{d}", lowBound=0)
            prob += ds_e >= min_er_wd - er_sum
            day_short[("er_wd", d)] = ds_e
            prob += er_sum >= max(1, min_er_wd - 1)

            # ER: ERリーダー(CLS_ERL)を必ず1名以上配置
            prob += pulp.lpSum(ud[s, d, UNIT_ER] for s in erl_staff) >= 1

    # ================================================================
    # 3E: 夜勤ユニット制約（病棟2・HCU1・リーダー1 = 計4名）
    # ================================================================
    for d in days:
        prob += pulp.lpSum(x[s, d, N] for s in names) == night_count  # 夜勤4名固定
        prob += pulp.lpSum(un[s, d, UNIT_WD]   for s in names) == 2
        prob += pulp.lpSum(un[s, d, UNIT_HCU]  for s in names) == 1
        prob += pulp.lpSum(un[s, d, UNIT_LEAD] for s in names) == 1

    # 夜勤リーダーは leader_pool から
    non_night_leader = [s for s in names if s not in leader_pool]
    for s in non_night_leader:
        for d in days:
            prob += un[s, d, UNIT_LEAD] == 0

    print(f"  夜勤: 4名固定（病棟2+HCU1+リーダー1）")
    print(f"  遅出可能スタッフ: {len(late_pool)}名")
    print(f"  平日日勤: 病棟≥{min_ward_wd}, HCU≥{min_hcu_wd}, ER≥{min_er_wd}+ERリーダー1（合計上限{max_day_wd}名）")
    print(f"  休日日勤: 病棟≥{min_ward_hd}, HCU≥{min_hcu_hd}, ER=0（合計{min_day_hd}〜{max_day_hd}名）")

    # day_short_excl は3Eでは未使用（新人制度なし）
    day_short_excl = {}
    # ================================================================
    # 3E: 夜勤Min/Max
    # ================================================================
    ft_non_ded = fulltime  # 3Eでは専従なし
    ft_slots   = night_count * num_days
    if ft_non_ded:
        even_target = ft_slots / len(ft_non_ded)
    else:
        even_target = 0
    print(f"  夜勤配分: {ft_slots}枠 ÷ {len(ft_non_ded)}人 ≒ {even_target:.1f}回/人")

    night_min_miss = {}
    night_max_over = {}
    for s in fulltime:
        total_n = pulp.lpSum(x[s, d, N] for d in days)
        s_min = night_min[s]
        s_max = night_max[s] if night_max[s] is not None else max_n_reg
        if s_min is not None:
            prob += total_n >= s_min
        if night_max[s] is not None and night_max[s] == 0:
            prob += total_n == 0
        else:
            nm_over = pulp.LpVariable(f"nmax_over_{s}", lowBound=0)
            prob += nm_over >= total_n - s_max
            night_max_over[s] = nm_over

    night_72h_over = {}  # 3Eでは未使用（互換変数）
    ld_sn_pen = {}
    ld_consec_pen = {}

    # 3E: 研修(R)は希望(=R)で指定された日のみ許可。それ以外は禁止。
    _r_requested = set()
    for _s, _rs in requests.items():
        if _s not in names:
            continue
        for _d, _t in _rs.items():
            if 1 <= _d <= num_days and _t == R:
                _r_requested.add((_s, _d - 1))
    for s in names:
        for d in days:
            if (s, d) not in _r_requested:
                prob += x[s, d, R] == 0

    for s in parttime:
        for d in days:
            prob += x[s, d, N] == 0
            prob += x[s, d, SN] == 0   # パートは短夜勤も不可
        target = round(weekly[s] * num_days / 7)
        # 休暇日数を差し引く
        vac_count = len(vacation_days.get(s, set()))
        # 勤務可能日数を考慮（曜日制限・祝日不可・土日不可・休暇で減る）
        wd_set = work_days_map.get(s)
        no_hol = no_holiday_map.get(s, False)
        no_we = no_weekend_map.get(s, False)
        avail_days = 0
        for d in days:
            # 休暇日は勤務不可
            if d in vacation_days.get(s, set()):
                continue
            day_wd = (fwd + d) % 7
            is_hol = (d + 1) in holidays
            blocked = False
            if wd_set is not None and day_wd not in wd_set:
                blocked = True
            if no_hol and is_hol:
                blocked = True
            if no_we and day_wd >= 5:  # 土(5)日(6)
                blocked = True
            if not blocked:
                avail_days += 1
        actual_target = min(target, avail_days)
        if actual_target < target:
            reason = []
            if vac_count > 0:
                reason.append(f"休暇{vac_count}日")
            if wd_set is not None or no_hol or no_we:
                reason.append("曜日/祝日/土日制限")
            reason_str = "・".join(reason) if reason else "制限"
            print(f"  ⚠ {s}: 勤務可能{avail_days}日 < 週{weekly[s]}目標{target}日 → {actual_target}日に調整 ({reason_str})")
        # 3Eでは時短スタッフ未対応 → Dのみでカウント
        _pt_shifts = [D]
        prob += pulp.lpSum(x[s, d, t] for d in days for t in _pt_shifts) >= max(actual_target - 1, 0)
        prob += pulp.lpSum(x[s, d, t] for d in days for t in _pt_shifts) <= actual_target + 1
    # 勤務曜日制限 + 祝日不可 + 土日不可
    for s in names:
        wd_set = work_days_map.get(s)
        no_hol = no_holiday_map.get(s, False)
        no_we = no_weekend_map.get(s, False)
        if wd_set is not None or no_hol or no_we:
            s_vac = vacation_days.get(s, set())
            for d in days:
                if d in s_vac:
                    continue  # 休暇日は曜日制限を適用しない（V=1が優先）
                day_wd = (fwd + d) % 7  # 0=月..6=日
                is_hol = (d + 1) in holidays
                blocked = False
                if wd_set is not None and day_wd not in wd_set:
                    blocked = True  # 勤務曜日外
                if no_hol and is_hol:
                    blocked = True  # 祝日不可
                if no_we and day_wd >= 5:
                    blocked = True  # 土日不可
                if blocked:
                    # 勤務不可 → 休のみ
                    prob += x[s, d, O] == 1
    # 連勤制限（研も勤務日としてカウント）— 個別指定対応
    for s in names:
        s_consec = consec_max_ind.get(s)
        limit = s_consec if s_consec is not None else max_consec
        for d in range(num_days - limit):
            prob += pulp.lpSum(
                x[s, d+i, D] + x[s, d+i, N] + x[s, d+i, R] for i in range(limit + 1)
            ) <= limit
    # 3E: Tierペア制約なし（クラス制度ではペアリングルール不要）
    # 日勤リーダー制約はユニット割り当て（UNIT_LEAD）で既に対応済み

    # 休暇(V)は希望がある日のみ
    for s in names:
        vac_set = vacation_days.get(s, set())
        for d in days:
            if d not in vac_set:
                prob += x[s, d, V] == 0

    # 委員会(I)は3Eでは無効化（全員禁止）
    for s in names:
        for d in days:
            prob += x[s, d, I] == 0

    # 公休日数制約（フルタイム全スタッフ）
    public_off_alerts = []
    for s in fulltime:
        v_count = len(vacation_days.get(s, set()))
        work_possible = num_days - v_count
        target_off    = public_off
        min_work_needed = (night_min[s] or 0) * 2  # 夜勤Min×2（夜+明）
        max_off_possible = work_possible - min_work_needed
        if target_off > max_off_possible:
            public_off_alerts.append(
                f"  ⚠ {s}: 公休{target_off}日は確保不可"
                f"（勤務可能{work_possible}日, 最低勤務{min_work_needed}日"
                f" → 最大休{max_off_possible}日）")
        total_o = pulp.lpSum(x[s, d, O] for d in days)
        prob += total_o >= target_off
        prob += total_o <= target_off + 1

    if public_off_alerts:
        print(f"\n⚠ 公休日数アラート:")
        for alert in public_off_alerts:
            print(alert)
        print(f"  → Infeasibleの原因になる可能性があります。\n")

    # 勤務希望（ソフト制約）+ 夜不（ハード制約）
    req_miss = {}
    for s, s_reqs in requests.items():
        if s not in names:
            continue
        for day_num, shift_type in s_reqs.items():
            if 1 <= day_num <= num_days:
                d_idx = day_num - 1
                if shift_type == "夜不":
                    # 夜勤不可: ハード制約
                    prob += x[s, d_idx, N] == 0
                elif shift_type == "休暇":
                    # 休暇: 強制的に暇シフト
                    prob += x[s, d_idx, V] == 1
                elif shift_type == "明休":
                    # 明または休: ハード制約（明 or 休のどちらか）
                    prob += x[s, d_idx, A] + x[s, d_idx, O] >= 1
                elif shift_type == "遅希":
                    # 遅出希望: ソフト制約（can_late フラグが必要）
                    if can_late.get(s):
                        key = (s, d_idx)
                        req_miss[key] = pulp.LpVariable(f"rmiss_{s}_{d_idx}", cat=pulp.LpBinary)
                        prob += x[s, d_idx, L] + req_miss[key] >= 1
                    # can_late=False のスタッフは遅出不可のため無視
                elif shift_type == R:
                    # 研修: ハード制約（勤務日数に含むが日勤人数にはカウントしない）
                    prob += x[s, d_idx, R] == 1
                elif shift_type == I:
                    pass  # 3Eでは委員会なし、無視
                else:
                    key = (s, d_idx)
                    req_miss[key] = pulp.LpVariable(f"rmiss_{s}_{d_idx}", cat=pulp.LpBinary)
                    prob += x[s, d_idx, shift_type] + req_miss[key] >= 1

    # --- ソフト制約 ---
    # 3E: 夜勤リーダー欠如ペナルティ（leader_pool からリーダー夜勤が入るか）
    a_miss = pulp.LpVariable.dicts("a_miss", days, cat=pulp.LpBinary)
    if leader_pool:
        ldr_ft = [n for n in leader_pool if weekly[n] is None]
        for d in days:
            if ldr_ft:
                prob += pulp.lpSum(x[s, d, N] for s in ldr_ft) + a_miss[d] >= 1
            else:
                prob += a_miss[d] == 1
    else:
        for d in days:
            prob += a_miss[d] == 1

    # 夜勤均等配分
    night_dev = {}
    if ft_non_ded:
        n_max_var = pulp.LpVariable("n_eq_max", lowBound=0)
        n_min_var = pulp.LpVariable("n_eq_min", lowBound=0, upBound=num_days)
        for s in ft_non_ded:
            total = pulp.lpSum(x[s, d, N] for d in days)
            prob += n_max_var >= total
            prob += n_min_var <= total
    else:
        n_max_var = pulp.LpVariable("n_eq_max", lowBound=0)
        n_min_var = pulp.LpVariable("n_eq_min", lowBound=0)
        prob += n_max_var == 0
        prob += n_min_var == 0

    # 推奨連勤超過ペナルティ
    cp = {}
    idx = 0
    for s in names:
        for d in range(num_days - pref_consec):
            cp[idx] = pulp.LpVariable(f"cp_{idx}", cat=pulp.LpBinary)
            prob += pulp.lpSum(
                x[s, d+i, D] + x[s, d+i, N] + x[s, d+i, L] for i in range(pref_consec + 1)
            ) - pref_consec <= cp[idx]
            idx += 1

    max_off = pulp.LpVariable("max_off", lowBound=0)
    min_off = pulp.LpVariable("min_off", lowBound=0, upBound=num_days)
    for s in fulltime:
        if not dedicated[s]:
            total_off = pulp.lpSum(x[s, d, O] for d in days)
            prob += max_off >= total_off
            prob += min_off <= total_off

    # 専従: 夜勤後の単休にペナルティ → 2連休推奨（夜明夜明後は特に重要）
    ded_single_off = {}
    ded_idx = 0
    for s in [n for n in fulltime if dedicated[n]]:
        for d in days[1:-1]:
            # 前日が勤務(N or A) かつ 当日が休 かつ 翌日が勤務(N) → 単休
            ded_single_off[ded_idx] = pulp.LpVariable(f"dso_{ded_idx}", cat=pulp.LpBinary)
            # x[s,d,O]=1 かつ x[s,d-1,O]=0 かつ x[s,d+1,O]=0 → 単休
            prob += ded_single_off[ded_idx] >= x[s, d, O] - x[s, d-1, O] - x[s, d+1, O] - (1 - x[s, d, O])
            ded_idx += 1

    # 夜勤まんべんなく配置: 月を3分割し各区間の夜勤数の偏りをペナルティ
    night_spread = {}
    ns_idx = 0
    third = num_days // 3
    periods = [list(range(0, third)), list(range(third, 2*third)), list(range(2*third, num_days))]
    for s in fulltime:
        for i in range(len(periods)):
            for j in range(i+1, len(periods)):
                night_spread[ns_idx] = pulp.LpVariable(f"nsprd_{ns_idx}", lowBound=0)
                ni = pulp.lpSum(x[s, d, N] for d in periods[i])
                nj = pulp.lpSum(x[s, d, N] for d in periods[j])
                prob += night_spread[ns_idx] >= ni - nj
                prob += night_spread[ns_idx] >= nj - ni
                ns_idx += 1

    # ユニット配置不足ペナルティ
    unit_short_vars = list(day_short.values())

    # 目的関数 (3E版)
    # P0 夜勤リーダー欠      300
    # P1 希望未達            250
    # P0 ユニット人数不足    200
    # P2 夜勤Max超過          25
    # P3 夜勤均等             60
    # P4 夜勤分散              6
    # P5 推奨連勤超過          3
    # P6 公休均等              8
    obj = (
        300 * pulp.lpSum(a_miss[d] for d in days)
        + 250 * pulp.lpSum(req_miss[k] for k in req_miss)
        + 200 * pulp.lpSum(unit_short_vars)
        +  25 * pulp.lpSum(night_max_over[s] for s in night_max_over)
        +  60 * (n_max_var - n_min_var)
        +   6 * pulp.lpSum(night_spread[i] for i in night_spread)
        +   3 * pulp.lpSum(cp[i] for i in cp)
        +   8 * (max_off - min_off)
    )
    if night_min_miss:
        obj += 40 * pulp.lpSum(night_min_miss[s] for s in night_min_miss)
    prob += obj

    # ============================================================
    # 複数パターン生成ループ
    # ============================================================
    # HiGHS優先（Python API）、なければCBCフォールバック
    try:
        solver = pulp.HiGHS(timeLimit=time_limit, msg=False)
        print("  ソルバー: HiGHS")
    except Exception:
        solver = pulp.PULP_CBC_CMD(timeLimit=time_limit, msg=False)
        print("  ソルバー: CBC（HiGHS未対応のためフォールバック）")
    # 夜勤の総割当数（差分制約の閾値計算用）
    total_night_slots = night_count * num_days  # 例: 2×31=62
    min_diff = max(int(total_night_slots * 0.3), night_count * 2)  # 30%以上変える

    all_results = []
    prev_solutions = []  # 過去パターンの夜勤割当を記録

    for pat_idx in range(num_patterns):
        pat_num = pat_idx + 1
        print(f"\n--- パターン {pat_num}/{num_patterns} 求解中（最大{time_limit}秒）...")

        status = prob.solve(solver)
        print(f"    結果: {pulp.LpStatus[status]}")

        if status != pulp.constants.LpStatusOptimal:
            if pat_idx == 0:
                print("✗ 解なし。希望・設定を見直してください。")
                return None
            else:
                print(f"    パターン{pat_num}以降は生成できませんでした。")
                break

        # 結果取得
        schedule = {}
        for s in names:
            schedule[s] = []
            for d in days:
                for t in SHIFTS:
                    if pulp.value(x[s, d, t]) > 0.5:
                        schedule[s].append(t)
                        break
                else:
                    schedule[s].append("?")

        missed_requests = {}
        for (s, d_idx), var in req_miss.items():
            if pulp.value(var) > 0.5:
                missed_requests.setdefault(s, []).append(d_idx + 1)

        # コンソール出力
        _print_result(pat_num, schedule, names, tiers, weekly, parttime,
                      days, holidays, weekends, public_off, min_day, min_day_excl,
                      a_miss, missed_requests, day_short, day_short_excl, year, month)
        # 夜勤均等度表示（N+SN合計）
        nd_counts = {s: sum(schedule[s].count(t) for t in NIGHT_SHIFTS) for s in ft_non_ded}
        if nd_counts:
            nd_min_v = min(nd_counts.values())
            nd_max_v = max(nd_counts.values())
            print(f"  夜勤均等(N+準): {nd_min_v}〜{nd_max_v}回 (差{nd_max_v - nd_min_v})")

        # 3E: 新人なし
        new_hire_days = {}

        # ユニット割り当て結果の取得
        unit_day = {}   # {staff: [unit_per_day]} (D シフトのみ)
        unit_night = {} # {staff: [unit_per_night]} (N シフトのみ)
        for s in names:
            ud_row = []
            un_row = []
            for d in days:
                if schedule[s][d] == D:
                    assigned = next((u for u in DAY_UNITS if pulp.value(ud[s, d, u]) > 0.5), "?")
                else:
                    assigned = ""
                ud_row.append(assigned)
                if schedule[s][d] == N:
                    assigned_n = next((u for u in NIGHT_UNITS if pulp.value(un[s, d, u]) > 0.5), "?")
                else:
                    assigned_n = ""
                un_row.append(assigned_n)
            unit_day[s]   = ud_row
            unit_night[s] = un_row

        result = {
            "schedule": schedule, "names": names, "tiers": classes,
            "classes": classes,
            "num_days": num_days, "year": year, "month": month,
            "settings": settings, "requests": requests,
            "holidays": holidays, "weekends": weekends,
            "public_off": public_off, "weekly": weekly,
            "dedicated": {n: False for n in names},
            "short_time": {n: False for n in names},
            "missed_requests": missed_requests,
            "new_hire_days": {},
            "new_hire_map": {n: False for n in names},
            "new_hire_grad_map": {n: None for n in names},
            "unit_day": unit_day,
            "unit_night": unit_night,
            "is_leader_map": is_leader_map,
            "is_er_leader_map": is_er_leader_map,
            "pattern_num": pat_num,
        }
        all_results.append(result)

        # 次パターン用: 今の夜勤配置と一定数以上異なる制約を追加
        if pat_idx < num_patterns - 1:
            current_night = {}
            for s in fulltime:
                for d in days:
                    current_night[(s, d)] = 1 if schedule[s][d] == N else 0
            prev_solutions.append(current_night)
            for p_idx, prev_night in enumerate(prev_solutions):
                diff_expr = pulp.lpSum(
                    x[s, d, N] * (1 - prev_night[(s, d)])
                    + (1 - x[s, d, N]) * prev_night[(s, d)]
                    for s in fulltime for d in days
                )
                prob += diff_expr >= min_diff, f"differ_p{pat_idx+1}_from_p{p_idx+1}"

    return all_results


def _print_result(pat_num, schedule, names, tiers, weekly, parttime,
                  days, holidays, weekends, public_off, min_day, min_day_excl,
                  a_miss, missed, day_short, day_short_excl, year, month):
    """1パターンのコンソール出力（3E版）"""
    print(f"\n{'='*55}")
    print(f"  パターン {pat_num} - スタッフ別統計 (公休{public_off}日)")
    print(f"{'='*55}")
    hdr = f"{'名前':>8} {'クラス':>8} {'日':>3} {'夜':>3} {'遅':>3} {'明':>3} {'休':>3} {'暇':>3} {'公休':>4}"
    if parttime:
        hdr += f" {'週':>3}"
    print(hdr)
    print("-" * len(hdr))
    for s in names:
        c = {t: schedule[s].count(t) for t in SHIFTS}
        ko = c[O] + c[V]
        cls_abbr = tiers[s][:4] if tiers[s] else "?"
        line = (f"{s:>8} {cls_abbr:>8} {c[D]:>3} {c[N]:>3} {c[L]:>3} {c[A]:>3}"
                f" {c[O]:>3} {c[V]:>3} {ko:>4}")
        if parttime:
            w = weekly.get(s)
            line += f" {f'週{w}' if w else '':>4}"
        print(line)

    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    first_wd = date(year, month, 1).weekday()
    print(f"\n  日別夜勤4名（病棟2+HCU1+リーダー1）:")
    for d in days:
        wd = weekdays_jp[(first_wd + d) % 7]
        hol = "祝" if (d+1) in holidays else ("★" if (d+1) in weekends else "")
        nn = [s for s in names if schedule[s][d] == N]
        ns = "+".join(f"{s}({tiers[s]})" for s in nn)
        print(f"    {d+1:>2} {wd}{hol:1} [{ns}]")

    miss = [d+1 for d in days if pulp.value(a_miss[d]) > 0.5]
    print(f"\n  {'⚠ 夜勤リーダー欠: '+str(miss)+'日' if miss else '✓ 全日夜勤リーダー配置'}")
    if missed:
        print(f"  ⚠ 未達希望:")
        for s, ds in missed.items():
            print(f"    {s}: {', '.join(f'{d}日' for d in ds)}")
    else:
        print("  ✓ 全希望達成")


# ============================================================
# ローカルExcel出力
# ============================================================

def _write_one_sheet(wb, result, sheet_title):
    """1パターン分のシートを wb に追加"""
    schedule = result["schedule"]
    names    = result["names"]
    tiers    = result["tiers"]
    num_days = result["num_days"]
    year     = result["year"]
    month    = result["month"]
    settings = result["settings"]
    requests = result["requests"]
    holidays = result.get("holidays", set())
    weekends = result.get("weekends", set())
    weekly_d = result.get("weekly", {})
    missed   = result.get("missed_requests", {})
    new_hire_days = result.get("new_hire_days", {})  # {name: set of 1-indexed days}
    min_day  = settings.get("min_day_staff") or 5

    def _is_nh(s, d_idx):
        """0-indexed day: そのスタッフがその日新人扱いか"""
        return (d_idx + 1) in new_hire_days.get(s, set())

    ws = wb.create_sheet(title=sheet_title)
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()

    from datetime import datetime as _dt
    _gen_at = _dt.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value=f"【3E】{year}年{month}月 勤務表  (出力: {_gen_at})").font = FONT_T

    RD, RW, RS = 3, 4, 5
    for r in (RD, RW):
        ws.cell(row=r, column=1).fill = FH
        ws.cell(row=r, column=2).fill = FH
    ws.cell(row=RD, column=1, value="名前").font = FONT_H
    ws.cell(row=RD, column=1).fill = FH
    ws.cell(row=RD, column=2, value="クラス").font = FONT_H
    ws.cell(row=RD, column=2).fill = FH
    ws.cell(row=RD, column=2).alignment = CTR

    for d in range(num_days):
        col = d + 3
        wi = (fwd + d) % 7
        is_special = wi >= 5 or (d+1) in holidays
        cd = ws.cell(row=RD, column=col, value=d+1)
        cd.alignment = CTR; cd.border = BDR; cd.fill = FH
        cd.font = FONT_R if is_special else FONT_H
        lbl = wdj[wi] + ("祝" if (d+1) in holidays and wi < 5 else "")
        cw = ws.cell(row=RW, column=col, value=lbl)
        cw.alignment = CTR; cw.border = BDR; cw.fill = FH
        cw.font = FONT_RS if is_special else Font(size=9)

    sc = num_days + 3
    for i, label in enumerate(["日", "夜", "準", "明", "早", "遅", "長", "短", "休", "研", "暇", "夜計"]):
        c = ws.cell(row=RD, column=sc+i, value=label)
        c.font = FONT_H; c.alignment = CTR; c.fill = FH; c.border = BDR

    row = RS
    cur_cls = None
    cls_label = {CLS_ER: "ER可", CLS_HCU: "HCU可", CLS_WD: "病棟可"}
    for s in names:
        t = tiers[s]
        if t != cur_cls:
            cur_cls = t
            ws.cell(row=row, column=1,
                    value=f"── {cls_label.get(t, t)} ──").font = Font(bold=True, size=9, color="666666")
            row += 1
        ws.cell(row=row, column=1, value=s).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).fill = FT.get(t, FT[CLS_WD])
        ws.cell(row=row, column=1).border = BDR
        ws.cell(row=row, column=1).alignment = CTR
        ct = ws.cell(row=row, column=2, value=t)
        ct.font = FONT; ct.alignment = CTR
        ct.fill = FT.get(t, FT[CLS_WD]); ct.border = BDR

        rd = requests.get(s, {})  # {day_num: shift_type}
        md = set(missed.get(s, []))
        # 希望達成セル用: 斜線ハッチング風の塗り（背景色を濃くして区別）
        FS_REQ = {
            D: _PF(patternType="lightUp", fgColor="FFCC99", bgColor="FFFFFF"),
            N: _PF(patternType="lightUp", fgColor="FFFFFF", bgColor="4472C4"),
            A: _PF(patternType="lightUp", fgColor="FFCC99", bgColor="FFF2CC"),
            O: _PF(patternType="lightUp", fgColor="FFCC99", bgColor="E2EFDA"),
            R: _PF(patternType="lightUp", fgColor="FFCC99", bgColor="E8D5F5"),
            V: _PF(patternType="lightUp", fgColor="FFCC99", bgColor="F4CCCC"),
        }
        for d in range(num_days):
            shift = schedule[s][d]
            cell = ws.cell(row=row, column=d+3, value=shift)
            is_req = (d+1) in rd
            is_missed = (d+1) in md
            if is_missed:
                # 未達希望: 赤太枠 + 網掛け
                cell.fill = _PF(patternType="lightUp", fgColor="FF6666", bgColor=FS.get(shift, FS[O]).fgColor or "FFFFFF")
                cell.font = FONT_N if shift == N else FONT
                cell.border = Border(
                    left=_S(style="medium", color="FF0000"),
                    right=_S(style="medium", color="FF0000"),
                    top=_S(style="medium", color="FF0000"),
                    bottom=_S(style="medium", color="FF0000"))
            elif is_req:
                # 希望達成: オレンジ枠 + 網掛け
                cell.fill = FS_REQ.get(shift, FS.get(shift, FS[O]))
                cell.font = FONT_N if shift == N else FONT
                cell.border = BDR_REQ
            else:
                cell.fill = FS.get(shift, FS[O])
                cell.font = FONT_N if shift == N else FONT
                cell.border = BDR
            cell.alignment = CTR
            # 希望セルにコメント（希望内容を表示）
            if is_req:
                from openpyxl.comments import Comment
                req_label = {"夜不": "夜不", "休暇": "休暇"}.get(rd[d+1], rd[d+1])
                cell.comment = Comment(f"希望: {req_label}", "シフト作成")

        counts = {sh: schedule[s].count(sh) for sh in SHIFTS}
        night_total = counts[N] + counts[SN]
        for i, v in enumerate([counts[D], counts[N], counts[SN], counts[A],
                                counts[E], counts[L], counts[LD], counts[ST],
                                counts[O], counts[R], counts[V], night_total]):
            c = ws.cell(row=row, column=sc+i, value=v)
            c.alignment = CTR; c.border = BDR
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="日別集計").font = FONT_H
    row += 1

    # 新人が一人でもいれば「日勤計(新人除く)」行を追加
    has_new_hire = bool(new_hire_days)
    if has_new_hire:
        ws.cell(row=row, column=1, value="日勤計(新人除く)").font = FONT_H
        ws.cell(row=row, column=1).border = BDR
        for d in range(num_days):
            cnt = sum(1 for s in names
                      if schedule[s][d] in DAY_SHIFTS and not _is_nh(s, d))
            cell = ws.cell(row=row, column=d+3, value=cnt)
            cell.alignment = CTR; cell.border = BDR
            cell.font = Font(italic=True, color="555555")
        row += 1

    for tl2, ts in [("日勤計", D), ("夜勤計", N), ("短夜勤計", SN), ("明け計", A),
                    ("早出計", E), ("遅出計", L), ("長日勤計", LD), ("時短計", ST),
                    ("休み計", O), ("研修計", R), ("休暇計", V)]:
        ws.cell(row=row, column=1, value=tl2).font = FONT_H
        ws.cell(row=row, column=1).border = BDR
        for d in range(num_days):
            cnt = sum(1 for s in names if schedule[s][d] == ts)
            cell = ws.cell(row=row, column=d+3, value=cnt)
            cell.alignment = CTR; cell.border = BDR
            # 日勤系合計が最低人数を下回る場合に警告色
            if ts == D:
                day_total = sum(1 for s in names if schedule[s][d] in DAY_SHIFTS)
                if day_total < min_day:
                    cell.fill = _PF(start_color="FF9999", end_color="FF9999", fill_type="solid")
        row += 1

    # リーダー可 / ERリーダー可（勤務中人数）
    _ldr_map = result.get("is_leader_map", {})
    _erl_map = result.get("is_er_leader_map", {})
    _WORK = set(DAY_SHIFTS) | set(NIGHT_SHIFTS)
    for tl3, _map in [("リーダー可", _ldr_map), ("ERリーダー可", _erl_map)]:
        ws.cell(row=row, column=1, value=tl3).font = FONT_H
        ws.cell(row=row, column=1).border = BDR
        for d in range(num_days):
            cnt = sum(1 for s in names if _map.get(s) and schedule[s][d] in _WORK)
            cell = ws.cell(row=row, column=d+3, value=cnt)
            cell.alignment = CTR; cell.border = BDR
            cell.font = Font(bold=True, color="1F4E79")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="夜勤ペア").font = FONT_H
    row += 1
    for lbl, fn in [("メンバー", lambda s: s), ("Tier", lambda s: tiers[s])]:
        ws.cell(row=row, column=1, value=lbl).font = Font(bold=True, size=9)
        ws.cell(row=row, column=1).border = BDR
        for d in range(num_days):
            nm = [s for s in names if schedule[s][d] in NIGHT_SHIFTS]
            cell = ws.cell(row=row, column=d+3, value="/".join(fn(s) for s in nm))
            cell.alignment = CTR; cell.border = BDR; cell.font = Font(size=8)
            if lbl == "Tier" and not any(tiers[s] in (TIER_A, TIER_AB) for s in nm):
                cell.fill = _PF(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 5
    for d in range(num_days):
        ws.column_dimensions[get_column_letter(d+3)].width = 4.2
    for i in range(7):
        ws.column_dimensions[get_column_letter(sc+i)].width = 4.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0

    return ws


def export_excel(results):
    """results: list of result dicts → 1ファイルにパターン別シート"""
    if not results:
        return None
    first = results[0]
    year  = first["year"]
    month = first["month"]

    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシート削除

    for result in results:
        pat = result.get("pattern_num", 1)
        title = f"パターン{pat}" if len(results) > 1 else "勤務表"
        _write_one_sheet(wb, result, title)

    path = os.path.join(BASE_DIR, f"勤務表_{year}_{month:02d}.xlsx")
    wb.save(path)
    print(f"\n✅ Excel: {path} ({len(results)}パターン)")
    return path


# ============================================================
# メイン
# ============================================================

def main():
    p = argparse.ArgumentParser(description="勤務表自動作成 v5.0")
    p.add_argument("--init", action="store_true", help="Excelテンプレート生成")
    p.add_argument("--init-gsheet", action="store_true", help="Gスプレッドシートテンプレート生成")
    p.add_argument("--gsheet", metavar="URL", help="Gスプレッドシートから読み込み")
    p.add_argument("--patterns", type=int, default=1, metavar="N",
                   help="生成パターン数（デフォルト: 1）")
    args = p.parse_args()

    if args.init:
        create_template(); return
    if args.init_gsheet:
        create_gsheet_template(); return

    num_pat = max(1, args.patterns)

    if args.gsheet:
        staff, reqs, settings, sh = load_gsheet(args.gsheet)
        results = build_and_solve(staff, reqs, settings, num_patterns=num_pat)
        if not results: sys.exit(1)
        write_gsheet_result(sh, results)
        export_excel(results)
    else:
        staff, reqs, settings = load_input()
        results = build_and_solve(staff, reqs, settings, num_patterns=num_pat)
        if not results: sys.exit(1)
        export_excel(results)
    print("\n完了!")


if __name__ == "__main__":
    main()
