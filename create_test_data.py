#!/usr/bin/env python3
"""
実ICUシナリオのテストデータ生成（2026年5月）
ベース: 新人あり勤務表テンプレート_2026_05 (5).xlsx
反映した決定事項:
  - 島末 C → C+（C既卒）
  - 宮本 C → C+（C既卒）
  - 小倉 AB → B（再分類）
  - 加嶋  B（据え置き）
  - 重信 C → A + 時短ON + 夜勤Max=0 + 週3
  - 川本 C + 新人ON（2ヶ月前採用、研修夜勤も併用）
  - 橋本/岡崎/清水 C + 新人ON（完全新人・夜勤なし）
  - その他（満田/川島/吉川/堺/伊藤/花田 → A, 星野 → AB 他）は据え置き
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import calendar
from datetime import date
import jpholiday

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection as CellProtection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter

year, month = 2026, 5
num_days = calendar.monthrange(year, month)[1]  # 31

wb = Workbook()
thin = Side(style="thin")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
staff_hdr_fill = PatternFill("solid", fgColor="548235")
staff_hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF", size=11)

# ====== スタッフデータ ======
# 列: 名前, Tier, 夜勤専従, 時短, 週勤務, 前月末, 夜勤Min, 夜勤Max,
#     連勤Max, 勤務曜日, 祝日不可, 土日不可, 夜勤研修, 研修夜勤回数, 新人, 新人卒業日
staff_data = [
    # ── A（単独L可） ──
    ["満田",   "A",  None, None, None, "明", None, None, None, None, None, None, None, None, None, None],
    ["川島",   "A",  None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["吉川",   "A",  "○",  None, None, "夜", None, None, None, None, None, None, None, None, None, None],
    ["堺",     "A",  None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["伊藤",   "A",  None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["花田",   "A",  None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    # 重信: A + 時短ON + 夜勤Max=0 + 週3（日勤単独Lは可、夜勤は入らない）
    ["重信",   "A",  None, "○",  3,    None, None, 0,    None, None, None, None, None, None, None, None],
    # ── AB ──
    ["星野",   "AB", None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    # ── B（小倉を再分類, 加嶋も従来B） ──
    ["小倉",   "B",  None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["加嶋",   "B",  None, None, None, "明", None, None, None, None, None, None, None, None, None, None],
    # ── C+（既卒） ──
    ["島末",   "C+", None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    ["宮本",   "C+", None, None, None, "明", None, None, None, None, None, None, None, None, None, None],
    # ── C（通常） ──
    ["長谷川", "C",  None, None, None, "夜", None, None, None, None, None, None, None, None, None, None],
    # 南・三崎: 週3・月火木/月火水のみ・祝日/土日不可・夜勤なし
    ["南",     "C",  None, None, 3,    None, None, 0,    None, "月火木", "○", "○", None, None, None, None],
    ["三崎",   "C",  None, None, 3,    None, None, 0,    None, "月火水", "○", "○", None, None, None, None],
    # ── 新人（日勤頭数のみ、独立Cやリーダー判定から除外） ──
    # 川本: 2ヶ月前採用。夜勤研修2回で通常夜勤へ移行中
    ["川本",   "C",  None, None, None, None, 4,    4,    None, None, None, None, "○", 2, "○", None],
    # 完全新人（夜勤なし・研修なし・日勤のみ）
    ["橋本",   "C",  None, None, None, None, 0,    0,    None, None, None, None, None, None, "○", None],
    ["岡崎",   "C",  None, None, None, None, 0,    0,    None, None, None, None, None, None, "○", None],
    ["清水",   "C",  None, None, None, None, 0,    0,    None, None, None, None, None, None, "○", None],
]

# ====== 勤務希望データ (day: value) ======
# 元テンプレートから移植。名前に対応するもののみ記載。
requests = {
    "島末":   {20:"休", 21:"休", 22:"休", 23:"休", 28:"日"},
    "満田":   {3:"休", 4:"休", 5:"休", 11:"研", 17:"休", 25:"夜不", 27:"休", 28:"日", 31:"休"},
    "川島":   {13:"休", 14:"休", 15:"休", 16:"休", 20:"日"},
    "吉川":   {5:"休", 6:"休", 7:"休", 8:"休", 9:"休"},
    "堺":     {1:"休", 2:"休", 9:"休", 10:"休", 26:"日", 30:"夜"},
    "伊藤":   {9:"休", 12:"日", 23:"休", 26:"休", 27:"休", 28:"休", 29:"休"},
    "小倉":   {7:"夜不", 21:"日", 31:"休"},
    "星野":   {3:"休", 4:"休", 5:"休", 10:"休", 30:"研"},
    "加嶋":   {22:"日"},
    "花田":   {8:"休", 9:"休", 14:"日", 15:"日", 21:"日", 22:"日", 30:"夜"},
    "重信":   {2:"休", 3:"休", 4:"休", 6:"休", 10:"休", 12:"休", 16:"休", 17:"休", 24:"休", 31:"休"},
    "長谷川": {3:"休", 18:"休", 19:"休", 20:"休", 31:"休"},
    "川本":   {30:"研"},
    "宮本":   {4:"夜不", 5:"休", 7:"夜不", 10:"休", 11:"夜不", 14:"夜不", 17:"休", 18:"夜不",
               21:"夜不", 24:"休", 25:"夜不", 28:"夜不", 31:"休"},
    "南":     {4:"休", 5:"休", 6:"休", 8:"休"},
    "三崎":   {},
    "橋本":   {11:"研", 18:"研", 25:"研"},
    "岡崎":   {11:"研", 18:"研", 25:"研"},
    "清水":   {11:"研", 18:"研", 25:"研"},
}

num_staff = len(staff_data)
staff_headers = ["名前", "Tier", "夜勤専従", "時短", "週勤務", "前月末",
                 "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可",
                 "夜勤研修", "研修夜勤回数", "新人", "新人卒業日"]
n_staff_cols = len(staff_headers)
weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
first_wd = date(year, month, 1).weekday()
holidays = {d.day for d, _ in jpholiday.month_holidays(year, month)}

# ゼブラ
staff_fill_even = PatternFill("solid", fgColor="E2EFDA")
staff_fill_odd  = PatternFill("solid", fgColor="F5FAF0")
req_fill_even   = PatternFill("solid", fgColor="D6E4F0")
req_fill_odd    = PatternFill("solid", fgColor="EFF5FB")
ref_fill = PatternFill("solid", fgColor="F2F2F2")
ref_font = Font(color="888888", size=10)

# =============== Sheet 1: 設定 ===============
ws_s = wb.active
ws_s.title = "設定"
ws_s.column_dimensions["A"].width = 22
ws_s.column_dimensions["B"].width = 14
ws_s.column_dimensions["C"].width = 35

from shift_scheduler import SETTINGS_DEF, SETTINGS_KEYS
ws_s.cell(row=1, column=1, value="勤務表設定").font = Font(bold=True, size=14)
for c, txt in enumerate(["項目", "値", "説明"], 1):
    cell = ws_s.cell(row=3, column=c, value=txt)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = bdr
    cell.alignment = Alignment(horizontal="center")
for i, (label, default, desc) in enumerate(SETTINGS_DEF):
    r = 4 + i
    ws_s.cell(row=r, column=1, value=label).border = bdr
    ws_s.cell(row=r, column=2, value=default).border = bdr
    ws_s.cell(row=r, column=3, value=desc).border = bdr
    ws_s.cell(row=r, column=3).font = Font(color="888888", size=9)

# =============== Sheet 2: スタッフ情報 ===============
ws_staff = wb.create_sheet("スタッフ情報")
ws_staff.cell(row=1, column=1, value=f"👤 スタッフ情報 — {year}年{month}月").font = Font(bold=True, size=14)
ws_staff.cell(row=2, column=1, value="※ このシートが原本です。名前・Tierは勤務希望シートに自動反映されます。").font = Font(color="888888", size=9)

for c, txt in enumerate(staff_headers, 1):
    cell = ws_staff.cell(row=3, column=c, value=txt)
    cell.fill = staff_hdr_fill; cell.font = staff_hdr_font; cell.border = bdr
    cell.alignment = Alignment(horizontal="center")

for i, row_data in enumerate(staff_data):
    r = 4 + i
    _s_fill = staff_fill_even if i % 2 == 0 else staff_fill_odd
    for c, val in enumerate(row_data, 1):
        cell = ws_staff.cell(row=r, column=c)
        cell.value = val if val is not None and val != "" else None
        cell.border = bdr; cell.fill = _s_fill

ws_staff.column_dimensions["A"].width = 14
for i in range(1, n_staff_cols):
    ws_staff.column_dimensions[get_column_letter(i + 1)].width = 10
ws_staff.freeze_panes = "B4"

# 凡例エリア
legend_start_row = 4 + num_staff + 2
ws_staff.cell(row=legend_start_row, column=1, value="📖 Tier定義").font = Font(bold=True, size=11, color="548235")
tier_defs = [
    ("A",  "ベテラン・リーダー格（日勤/夜勤リーダー単独可）"),
    ("AB", "中堅・リーダー代行可（夜勤リーダー可）"),
    ("B",  "一人立ち済み（B+B, B+C族の夜勤ペア禁止）"),
    ("C+", "C既卒（C+C+, C++通常Cペア禁止／A/AB/B下で夜勤可）"),
    ("C",  "新人・経験浅い（必ずA/AB/Bと夜勤ペア）"),
]
for i, (tier, desc) in enumerate(tier_defs):
    ws_staff.cell(row=legend_start_row + 1 + i, column=1, value=tier).font = Font(bold=True)
    ws_staff.cell(row=legend_start_row + 1 + i, column=2, value=desc).font = Font(color="555555", size=9)

# =============== Sheet 3: 勤務希望 ===============
ws_req = wb.create_sheet("勤務希望")
ws_req.cell(row=1, column=1, value=f"📝 勤務希望 — {year}年{month}月").font = Font(bold=True, size=14)
ws_req.cell(row=2, column=1, value="名前・Tierはスタッフ情報シートから自動同期。日付セルにシフト記号を入力してください。").font = Font(color="888888", size=9)

req_headers_fixed = ["名前", "Tier"]
for c, txt in enumerate(req_headers_fixed, 1):
    cell = ws_req.cell(row=3, column=c, value=txt)
    cell.fill = staff_hdr_fill; cell.font = staff_hdr_font; cell.border = bdr
    cell.alignment = Alignment(horizontal="center")

day_start_col = 3
for d in range(1, num_days + 1):
    col = day_start_col + d - 1
    wd_name = weekdays_jp[(first_wd + d - 1) % 7]
    cell = ws_req.cell(row=3, column=col, value=f"{d}({wd_name})")
    cell.alignment = Alignment(horizontal="center"); cell.border = bdr
    wd_idx = (first_wd + d - 1) % 7
    if d in holidays:
        cell.fill = PatternFill("solid", fgColor="F4CCCC")
        cell.font = Font(bold=True, color="CC0000", size=9)
    elif wd_idx >= 5:
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF", size=9)
    else:
        cell.fill = PatternFill("solid", fgColor="BDD7EE")
        cell.font = Font(bold=True, color="1F4E79", size=9)

locked = CellProtection(locked=True)
unlocked = CellProtection(locked=False)

for i in range(num_staff):
    r = 4 + i
    staff_row = 4 + i
    name = staff_data[i][0]
    _r_fill = req_fill_even if i % 2 == 0 else req_fill_odd

    # 名前 — 数式(実運用) + 直接値(テスト用: data_only=True対応)
    cell_name = ws_req.cell(row=r, column=1)
    cell_name.value = name  # 直接値（テスト時のdata_only=True対応）
    cell_name.border = bdr; cell_name.fill = ref_fill; cell_name.font = ref_font
    cell_name.protection = locked

    # Tier — 直接値
    cell_tier = ws_req.cell(row=r, column=2)
    cell_tier.value = staff_data[i][1]  # Tier直接値
    cell_tier.border = bdr; cell_tier.fill = ref_fill; cell_tier.font = ref_font
    cell_tier.alignment = Alignment(horizontal="center")
    cell_tier.protection = locked

    # 勤務希望データ
    staff_reqs = requests.get(name, {})
    for d in range(1, num_days + 1):
        cell = ws_req.cell(row=r, column=day_start_col + d - 1)
        val = staff_reqs.get(d)
        cell.value = val if val else None
        cell.border = bdr; cell.fill = _r_fill
        cell.protection = unlocked

ws_req.column_dimensions["A"].width = 14
ws_req.column_dimensions["B"].width = 8
for d in range(1, num_days + 1):
    ws_req.column_dimensions[get_column_letter(day_start_col + d - 1)].width = 7
ws_req.freeze_panes = "C4"

# シート保護
ws_req.protection = SheetProtection(sheet=True, objects=True, scenarios=True,
                                     formatColumns=False, formatRows=False)

# 凡例
req_legend_row = 4 + num_staff + 2
ws_req.cell(row=req_legend_row, column=1, value="📖 シフト種別").font = Font(bold=True, size=11, color="1F4E79")
shift_legend = [
    ("日", "日勤 (8:00〜17:00)"), ("夜", "夜勤 (16:45〜翌9:00 / 16h)"),
    ("準", "短夜勤 (17:00〜翌5:00 / 12h)"), ("早", "早出 (7:00〜16:00)"),
    ("遅", "遅出 (12:00〜21:00)"), ("長", "長日勤 (8:45〜21:00 / 12h)"),
    ("短", "時短 (8:45〜16:00 / 6.25h)"), ("休", "公休"),
    ("研", "研修"), ("夜不", "この日は夜勤不可"),
    ("休暇", "有給休暇"), ("明休", "明または休"),
]
for i, (sym, desc) in enumerate(shift_legend):
    ws_req.cell(row=req_legend_row + 1 + i, column=1, value=sym).font = Font(bold=True, size=10)
    ws_req.cell(row=req_legend_row + 1 + i, column=2, value=desc).font = Font(color="555555", size=9)

# 保存
outpath = os.path.join(os.path.dirname(__file__), "テスト_勤務表_2026_05.xlsx")
wb.save(outpath)
print(f"✅ テストデータ作成完了: {outpath}")
print(f"   スタッフ: {num_staff}人")
print(f"   月: {year}年{month}月（{num_days}日）")

# 検証
wb2 = load_workbook(outpath)
print(f"   シート: {wb2.sheetnames}")
ws_check = wb2["スタッフ情報"]
tier_count = {}
for r in range(4, 4 + num_staff):
    t = ws_check.cell(row=r, column=2).value
    tier_count[t] = tier_count.get(t, 0) + 1
print(f"   Tier分布: {tier_count}")
