#!/usr/bin/env python3
"""
看護師勤務表自動作成 - Streamlit アプリ
Excel / Google Sheets 両対応
"""
import streamlit as st
import pandas as pd
import calendar
from datetime import date
from io import BytesIO

import jpholiday

from shift_scheduler import (
    Staff, build_and_solve,
    D, N, A, O, R, V, E, L, ST, LD, SN, I,
    SHIFTS, DAY_SHIFTS, NIGHT_SHIFTS,
    TIER_A, TIER_AB, TIER_B, TIER_CP, TIER_C, VALID_TIERS,
    _get_holidays_and_days_off, _write_one_sheet,
    SETTINGS_DEF, SETTINGS_KEYS,
    _parse_staff_list, _parse_requests, _parse_settings,
)

# ============================================================
# ページ設定
# ============================================================
st.set_page_config(
    page_title="ICU勤務表作成ツール",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    /* ── 全体フォント ── */
    html, body, [class*="css"] { font-family: 'Noto Sans JP', 'Hiragino Kaku Gothic ProN', sans-serif; }

    /* ── メトリックカード ── */
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #f8fbff 0%, #eef4fc 100%);
        border: 1px solid #d0e4f7;
        border-left: 4px solid #2563eb;
        padding: 14px 16px;
        border-radius: 10px;
        box-shadow: 0 1px 4px rgba(37,99,235,0.08);
    }
    div[data-testid="stMetric"] label { color: #4b6a8a; font-size: 0.78rem; font-weight: 600; letter-spacing: 0.03em; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { color: #1e3a5f; font-size: 1.6rem; font-weight: 700; }

    /* ── タブ ── */
    button[data-baseweb="tab"] { font-weight: 600; font-size: 0.88rem; }
    button[data-baseweb="tab"][aria-selected="true"] { color: #2563eb; border-bottom-color: #2563eb; }

    /* ── データフレーム ── */
    .stDataFrame { font-size: 12px; }
    .stDataFrame thead th { background: #1e3a5f !important; color: white !important; font-weight: 600; }

    /* ── ボタン ── */
    button[kind="primary"] { background: #2563eb; border-radius: 8px; font-weight: 600; }
    button[kind="primary"]:hover { background: #1d4ed8; }

    /* ── サイドバー ── */
    section[data-testid="stSidebar"] { background: #f0f5ff; }
    section[data-testid="stSidebar"] .stMarkdown h3 { color: #1e3a5f; font-size: 0.9rem; }

    /* ── 凡例バッジ ── */
    .shift-badge {
        display: inline-block; padding: 2px 10px; border-radius: 12px;
        font-size: 0.8rem; font-weight: 600; margin: 2px;
    }

    /* ── セクション区切り ── */
    hr { border: none; border-top: 2px solid #e2eaf5; margin: 1rem 0; }

    /* ── 警告・成功バナー強化 ── */
    div[data-testid="stAlert"] { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 定数・ヘルパー
# ============================================================
WEEKDAY_MAP = {"月": 0, "火": 1, "水": 2, "木": 3, "金": 4, "土": 5, "日": 6}
WEEKDAY_REV = {v: k for k, v in WEEKDAY_MAP.items()}

# ============================================================
# 日本看護協会 夜勤ガイドラインチェック
# ============================================================
def check_nursing_guidelines(schedule, names, tiers, r_dedicated, night_hours=16):
    """
    日本看護協会「夜勤・交代制勤務に関するガイドライン（2013年）」+
    労基法準拠チェック
    Args:
        night_hours: 1夜勤あたりの時間数（二交代=16h, 三交代=8h）
    """
    violations = []
    warnings = []
    ok_list = []

    for s in names:
        shifts = schedule[s]
        is_dedicated = r_dedicated.get(s, False)
        issues = []
        night_count = shifts.count(N)   # 通常夜勤
        sn_count = shifts.count(SN)     # 短夜勤
        total_night = night_count + sn_count

        # ── 1. 月の夜勤系回数（8回以内）────────────────────────────
        if not is_dedicated:
            if total_night > 8:
                issues.append({"rule": "夜勤回数", "level": "violation",
                                "detail": f"{total_night}回（N:{night_count}+準:{sn_count}）上限8回超過"})
            elif total_night == 8:
                issues.append({"rule": "夜勤回数", "level": "warning",
                                "detail": f"{total_night}回（N:{night_count}+準:{sn_count}）上限8回ちょうど"})

        # ── 2. 月夜勤時間72時間以内（日本看護協会ガイドライン）──
        # N×night_hours + SN×12h
        if not is_dedicated and night_hours > 0:
            total_hours = night_count * night_hours + sn_count * 12
            if total_hours > 72:
                issues.append({"rule": "72時間規制", "level": "violation",
                                "detail": f"月夜勤{total_hours}h（72h超過 / N:{night_count}×{night_hours}h+準:{sn_count}×12h）"})
            elif total_hours >= 64:
                issues.append({"rule": "72時間規制", "level": "warning",
                                "detail": f"月夜勤{total_hours}h（72h上限に近い）"})

        # ── 3. 連続夜勤（2回以内）───────────────────────────────
        i = 0; consec = 0; max_consec = 0
        while i < len(shifts):
            if shifts[i] == N:
                consec += 1; max_consec = max(max_consec, consec); i += 2
            else:
                consec = 0; i += 1
        if max_consec > 2:
            issues.append({"rule": "連続夜勤", "level": "violation",
                           "detail": f"最大{max_consec}連続（上限2回超過）"})

        # ── 4. 勤務間インターバル（11時間以上）──────────────────
        iv = [f"{d+2}日" for d in range(len(shifts)-1) if shifts[d]==A and shifts[d+1]==D]
        if iv:
            issues.append({"rule": "インターバル", "level": "violation",
                           "detail": f"明け翌日に日勤: {', '.join(iv)}（11時間未満の疑い）"})

        # ── 5. 夜勤翌日の明け確認 ──────────────────────────────
        for d in range(len(shifts)-1):
            if shifts[d] == N and shifts[d+1] != A:
                issues.append({"rule": "夜勤後の明け", "level": "violation",
                               "detail": f"{d+1}日: 夜勤翌日が明けでない（{shifts[d+1]}）"})

        # ── 集計 ─────────────────────────────────────────────
        total_hours = night_count * night_hours + sn_count * 12
        for item in issues:
            rec = {"名前": s, "Tier": tiers[s],
                   "夜勤回数": f"N:{night_count}+準:{sn_count}",
                   "月夜勤時間": f"{total_hours}h", **item}
            (violations if item["level"] == "violation" else warnings).append(rec)
        if not issues:
            ok_list.append(s)

    return violations, warnings, ok_list

def check_skill_pairing(schedule, names, tiers, num_days, year, month):
    """
    夜勤ペアのスキルバランスをチェックする。
    NG条件: 夜勤ペアが全員 C（新人のみ）の日がある。
    推奨: 各夜勤ペアにA/AB以上が1名以上含まれる。
    Returns:
        bad_days: list of dict（NG日の情報）
        warn_days: list of dict（注意日）
        ok_days: int
    """
    tier_rank = {TIER_A: 5, TIER_AB: 4, TIER_B: 3, TIER_CP: 2, TIER_C: 1}
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()
    bad_days = []
    warn_days = []
    ok_days = 0

    for d in range(num_days):
        night_staff = [s for s in names if schedule[s][d] in NIGHT_SHIFTS]
        if not night_staff:
            ok_days += 1
            continue
        wd = wdj[(fwd + d) % 7]
        day_label = f"{d+1}日({wd})"
        tier_list = [tiers.get(s, TIER_C) for s in night_staff]
        max_rank = max(tier_rank.get(t, 1) for t in tier_list)
        members_str = " + ".join(f"{s}({tiers.get(s,'?')})" for s in night_staff)

        if max_rank == 1:  # 全員C（新人のみ）
            bad_days.append({
                "日": day_label, "夜勤メンバー": members_str,
                "問題": "🚨 全員通常C",
            })
        elif max_rank == 2:  # 最高がC+
            bad_days.append({
                "日": day_label, "夜勤メンバー": members_str,
                "問題": "🚨 最高がC+（A/AB/B不在）",
            })
        elif max_rank == 3:  # 最高がB
            warn_days.append({
                "日": day_label, "夜勤メンバー": members_str,
                "問題": "⚠ A/AB不在（Bが最高）",
            })
        else:
            ok_days += 1

    return bad_days, warn_days, ok_days


# ユニット種別ごとの配置基準定義
# (ratio_val, check_night, display_ratio, basis_note)
# ※ 各管理料にクラス（1〜Nまで）が存在する場合でも、
#    看護師配置比率はクラス共通。クラス差は医師在院要件・施設基準の厳格さ。
UNIT_STANDARDS = {
    "ICU（特定集中治療室管理料 1〜4）":    (2,  True,  "2:1",
        "特定集中治療室管理料1〜4 / 看護配置は全クラス共通 2:1 / "
        "クラス差: 専任常勤医師の在院要件・集中治療実績基準"),
    "HCU（ハイケアユニット入院医療管理料 1〜2）": (4, True, "4:1",
        "ハイケアユニット入院医療管理料1〜2 / 看護配置は全クラス共通 4:1 / "
        "クラス差: 専任医師配置・施設要件の厳格さ"),
    "NICU（新生児集中治療室管理料 1〜2）": (3,  True,  "3:1",
        "新生児集中治療室管理料1〜2 / 看護配置は全クラス共通 3:1 / "
        "クラス差: 治療実績・専従医師要件"),
    "GCU（新生児治療回復室入院医療管理料）": (6, True, "6:1",
        "新生児治療回復室入院医療管理料 / 6:1（クラス区分なし）"),
    "SCU（脳卒中ケアユニット入院医療管理料 1〜2）": (3, True, "3:1",
        "脳卒中ケアユニット入院医療管理料1〜2 / 看護配置は全クラス共通 3:1 / "
        "クラス差: 専任医師の常時在院か否か"),
    "PICU（小児特定集中治療室管理料 1〜2）": (2, True, "2:1",
        "小児特定集中治療室管理料1〜2 / 看護配置は全クラス共通 2:1 / "
        "クラス差: 専任常勤医師の在院要件"),
    "一般病棟（急性期一般入院料1） 7:1":   (7,  False, "7:1",
        "急性期一般入院料1 / 7:1（日勤帯）"),
    "一般病棟（急性期一般入院料4） 10:1":  (10, False, "10:1",
        "急性期一般入院料4 / 10:1（日勤帯）"),
    "一般病棟（急性期一般入院料6） 13:1":  (13, False, "13:1",
        "急性期一般入院料6 / 13:1（日勤帯）"),
    "一般病棟（地域一般入院基本料） 15:1": (15, False, "15:1",
        "地域一般入院基本料 / 15:1（日勤帯）"),
    "回復期リハビリテーション病棟":         (13, False, "13:1",
        "回復期リハビリテーション病棟入院料 / 13:1（日勤帯）"),
    "地域包括ケア病棟":                     (13, False, "13:1",
        "地域包括ケア病棟入院料 / 13:1（日勤帯）"),
}


def check_staffing_ratio(schedule, names, r_dedicated, r_weekly,
                          num_days, bed_count, ratio_val, year, month,
                          check_night=False):
    """
    日別の人員数が配置基準を満たすか確認する。
    Args:
        check_night: True の場合、夜勤帯（夜勤中＝N）も同一基準でチェック（ICU系）
    必要人数 = ceil(病床数 / 配置比)
    Returns:
        shortfalls: list of dict（不足日の情報）
        ok_days: 基準を満たす日数
        required: 必要人数
    """
    required = max(1, -(-bed_count // ratio_val))  # ceiling division
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()
    shortfalls = []
    ok_days = 0
    for d in range(num_days):
        wd = wdj[(fwd + d) % 7]
        day_staff = sum(1 for s in names if schedule[s][d] in DAY_SHIFTS)
        issues = []
        if day_staff < required:
            issues.append(f"日勤系{day_staff}人（必要{required}人）")
        if check_night:
            night_staff = sum(1 for s in names if schedule[s][d] in NIGHT_SHIFTS)
            if night_staff < required:
                issues.append(f"夜勤系{night_staff}人（必要{required}人）")
        if issues:
            shortfalls.append({
                "日": f"{d+1}日({wd})",
                "日勤人数": day_staff,
                **({"夜勤人数": sum(1 for s in names if schedule[s][d] == N)} if check_night else {}),
                "必要人数": required,
                "不足内容": " / ".join(issues),
            })
        else:
            ok_days += 1
    return shortfalls, ok_days, required


SHIFT_DISPLAY = {
    D: "日", N: "夜", A: "明", O: "休", R: "研", V: "暇",
    E: "早", L: "遅", ST: "短", LD: "長", SN: "準", I: "委",
    "夜不": "夜不", "休暇": "休暇", "明休": "明休",
}
SHIFT_REVERSE = {
    "日": D, "夜": N, "明": A, "休": O, "研": R, "暇": V,
    "早": E, "遅": L, "短": ST, "長": LD, "準": SN, "委": I,
    "夜不": "夜不", "休暇": "休暇", "明休": "明休",
}
# シフト説明（ツールチップ用）
SHIFT_DESC = {
    D: "日勤 (8:00〜17:00)",
    N: "夜勤 (16:45〜翌9:00 / 16h)",
    A: "明け（夜勤明け）",
    O: "公休",
    R: "研修",
    V: "休暇",
    E: "早出 (7:00〜16:00 / 8h)",
    L: "遅出 (12:00〜21:00 / 8h)",
    ST: "時短 (8:45〜16:00 / 6.25h) 育児・介護",
    LD: "長日勤 (8:45〜21:00 / 12h)",
    SN: "短夜勤 (17:00〜翌5:00 / 12h)",
}


def _staff_to_df(staff_list):
    """Staffリスト → DataFrame"""
    rows = []
    for s in staff_list:
        wd_str = ""
        if s.work_days:
            wd_str = "".join(WEEKDAY_REV[d] for d in sorted(s.work_days))
        rows.append({
            "名前": s.name, "Tier": s.tier,
            "夜勤専従": s.dedicated,
            "時短": s.short_time,
            "週勤務": s.weekly_days,
            "前月末": s.prev_month or "",
            "夜勤Min": s.night_min,
            "夜勤Max": s.night_max,
            "連勤Max": s.consec_max,
            "勤務曜日": wd_str,
            "祝日不可": s.no_holiday,
            "土日不可": s.no_weekend,
            "夜勤研修": s.night_training,
            "研修夜勤回数": s.night_training_max,
            "新人": s.new_hire,
            "新人卒業日": s.new_hire_graduation_day,
        })
    return pd.DataFrame(rows) if rows else _default_staff()


def _reqs_to_df(reqs_dict, staff_list, num_days):
    """希望dict → DataFrame"""
    req_rows = []
    for name, rq in reqs_dict.items():
        row = {"名前": name}
        for d in range(1, num_days + 1):
            v = rq.get(d, "")
            row[str(d)] = SHIFT_DISPLAY.get(v, v) if v else ""
        req_rows.append(row)
    existing = {r["名前"] for r in req_rows}
    for s in staff_list:
        if s.name not in existing:
            row = {"名前": s.name}
            for d in range(1, num_days + 1):
                row[str(d)] = ""
            req_rows.append(row)
    return pd.DataFrame(req_rows)


def _render_load_preview(staff_df, requests_df=None):
    """読み込んだスタッフ情報の要約プレビューを表示。
    属性ごとにグループ化して「意図通りに読めているか」を1秒で確認できるようにする。
    論理エラーや要確認項目も警告として表示。
    requests_df を渡すと委員会予定等の希望ベースの情報も表示。
    """
    if staff_df is None or staff_df.empty:
        return
    # 有効な名前行のみ
    df = staff_df.dropna(subset=["名前"]).copy()
    df = df[df["名前"].astype(str).str.strip() != ""]
    if df.empty:
        return

    n_total = len(df)

    # Tier分布
    tier_counts = df["Tier"].fillna("").astype(str).str.strip().value_counts()
    tier_order = ["A", "AB", "B", "C+", "C"]
    tier_str = " / ".join(
        f"{t}:{int(tier_counts.get(t, 0))}" for t in tier_order if tier_counts.get(t, 0) > 0
    )
    other_tiers = [t for t in tier_counts.index if t not in tier_order and t]
    if other_tiers:
        tier_str += " / " + " / ".join(f"{t}:{int(tier_counts[t])}" for t in other_tiers)

    # ヘルパ: Trueっぽい値判定
    def _is_true(v):
        if isinstance(v, bool):
            return v
        s = str(v).strip().lower()
        return s in ("true", "1", "◯", "○", "yes")

    # 各カテゴリ抽出
    def _names_where(mask):
        return df.loc[mask, "名前"].astype(str).tolist()

    # 各カテゴリを単一ループで抽出（pandas.applyの挙動ブレを回避）
    def _is_zero(v):
        try:
            return v is not None and v != "" and float(v) == 0
        except (ValueError, TypeError):
            return False
    def _num_or_none(v):
        try:
            if v is None or v == "":
                return None
            return int(float(v))
        except (ValueError, TypeError):
            return None

    no_night = []
    new_hires = []
    nh_details = []
    nh_with_night_list = []
    pt_list = []
    dedicated = []
    short_time = []
    training = []
    prev_month = []
    day_restrict = []

    for _, row in df.iterrows():
        name = str(row.get("名前") or "").strip()
        if not name:
            continue
        tier = str(row.get("Tier") or "").strip()
        nmax = row.get("夜勤Max")
        wk_n = _num_or_none(row.get("週勤務"))
        is_new = _is_true(row.get("新人"))
        is_ded = _is_true(row.get("夜勤専従"))
        is_st = _is_true(row.get("時短"))
        is_tr = _is_true(row.get("夜勤研修"))
        pm = str(row.get("前月末") or "").strip()

        if _is_zero(nmax):
            no_night.append(name)
        if is_new:
            new_hires.append(name)
            grad = _num_or_none(row.get("新人卒業日"))
            nh_details.append(f"{name}({grad}日まで)" if grad is not None else f"{name}(月末まで)")
            # 新人で夜勤可能性あり（Max未指定 or >0）
            if nmax is None or nmax == "" or (not _is_zero(nmax)):
                nh_with_night_list.append(name)
        if wk_n is not None and wk_n > 0:
            pt_list.append(f"{name}(週{wk_n})")
        if is_ded:
            dedicated.append(name)
        if is_st:
            short_time.append(name)
        if is_tr:
            tm = _num_or_none(row.get("研修夜勤回数"))
            training.append(f"{name}(計{tm}回)" if tm is not None else f"{name}(回数未指定)")
        if pm in ("夜", "明"):
            prev_month.append(f"{name}({pm})")
        parts = []
        wd = str(row.get("勤務曜日") or "").strip()
        if wd:
            parts.append(wd)
        if _is_true(row.get("祝日不可")):
            parts.append("祝日×")
        if _is_true(row.get("土日不可")):
            parts.append("土日×")
        if parts:
            day_restrict.append(f"{name}({'/'.join(parts)})")

    # === 論理エラー検出 ===
    warnings = []
    errors = []
    for _, row in df.iterrows():
        name = row["名前"]
        tier = str(row.get("Tier") or "").strip()
        # Tierチェック
        if tier not in ("A", "AB", "B", "C+", "C"):
            errors.append(f"**{name}**: Tier '{tier}' が不正（A/AB/B/C+/C のいずれかが必要）")
        # 時短 x 週勤務
        if _is_true(row.get("時短")):
            wk = row.get("週勤務")
            if wk is None or wk == "":
                warnings.append(f"**{name}**: 時短=◯ ですが週勤務が未指定（時短は通常パートタイム運用）")
        # 夜勤研修 x 研修夜勤回数
        if _is_true(row.get("夜勤研修")):
            tm = row.get("研修夜勤回数")
            if tm is None or tm == "":
                warnings.append(f"**{name}**: 夜勤研修=◯ ですが研修夜勤回数が未指定")
        # 夜勤専従 x Tier
        if _is_true(row.get("夜勤専従")) and tier in ("C", "C+"):
            warnings.append(f"**{name}**: 夜勤専従=◯ ですが Tier={tier}（通常Aリーダー資格者が担当）")
        # 夜勤Min > 夜勤Max
        nmin = row.get("夜勤Min")
        nmax = row.get("夜勤Max")
        try:
            if (nmin is not None and nmin != "" and nmax is not None and nmax != ""
                    and float(nmin) > float(nmax)):
                errors.append(f"**{name}**: 夜勤Min({int(float(nmin))}) > 夜勤Max({int(float(nmax))}) は不可能")
        except (ValueError, TypeError):
            pass
        # 新人が夜勤する設定（情報として表示）
        # ※ 新人でも夜勤可の運用あり → エラーではなく"確認事項"

    # 新人で夜勤する可能性がある人（上のループで集計済み）
    nh_with_night = nh_with_night_list

    # === レンダリング ===
    with st.expander(f"📋 読み込み結果サマリー（{n_total}人）— クリックして確認", expanded=True):
        st.markdown(f"**🏷️ Tier構成**: {tier_str}")

        def _safe_join(lst):
            return ", ".join(str(x) for x in lst)

        cols = st.columns(2)
        with cols[0]:
            if no_night:
                st.markdown(f"**🚫 夜勤しない人（Max=0）** [{len(no_night)}人]  \n{_safe_join(no_night)}")
            if new_hires:
                st.markdown(f"**🎓 新人** [{len(new_hires)}人]  \n{_safe_join(nh_details)}")
            if pt_list:
                st.markdown(f"**⏰ パートタイム** [{len(pt_list)}人]  \n{_safe_join(pt_list)}")
            if short_time:
                st.markdown(f"**🪶 時短** [{len(short_time)}人]  \n{_safe_join(short_time)}")
        with cols[1]:
            if dedicated:
                st.markdown(f"**🌙 夜勤専従** [{len(dedicated)}人]  \n{_safe_join(dedicated)}")
            if training:
                st.markdown(f"**🏃 夜勤研修** [{len(training)}人]  \n{_safe_join(training)}")
            if prev_month:
                st.markdown(f"**🎯 前月繰越** [{len(prev_month)}人]  \n{_safe_join(prev_month)}")
            if day_restrict:
                st.markdown(f"**📆 勤務制限** [{len(day_restrict)}人]  \n{_safe_join(day_restrict)}")

        # 警告・エラー
        if errors:
            st.error("❌ **入力エラー**（このままでは動作不正）")
            for e in errors:
                st.markdown(f"- {e}")
        if warnings:
            st.warning("⚠️ **要確認**")
            for w in warnings:
                st.markdown(f"- {w}")
        if nh_with_night:
            st.info(
                f"ℹ️ **確認事項**: 新人で夜勤する設定の人がいます: {_safe_join(nh_with_night)}"
                f"（新人夜勤時はA必須のハード制約が自動適用されます。夜勤させない場合は夜勤Max=0に設定）"
            )
        if not errors and not warnings and not nh_with_night:
            st.success("✅ 論理エラー・要確認事項なし")

        # 委員会予定の集計（希望データがあれば表示）
        if requests_df is not None and not requests_df.empty:
            committee_items = []  # list of (name, day)
            day_cols = [c for c in requests_df.columns if str(c).isdigit()]
            for _, rq_row in requests_df.iterrows():
                name_v = str(rq_row.get("名前") or "").strip()
                if not name_v:
                    continue
                for dc in day_cols:
                    v = str(rq_row.get(dc) or "").strip()
                    if v == "委":
                        committee_items.append((name_v, int(dc)))
            if committee_items:
                # 日付ごとに集計
                from collections import defaultdict
                by_day = defaultdict(list)
                for nm, dd in committee_items:
                    by_day[dd].append(nm)
                day_summary = ", ".join(
                    f"{d}日({'/'.join(by_day[d])})" for d in sorted(by_day.keys())
                )
                st.info(
                    f"🏛️ **委員会予定** [{len(committee_items)}件]: {day_summary}"
                    f"（各日A/AB同席バックアップが自動で確保されます）"
                )


_SETTINGS_WIDGET_MAP = {
    "year": "inp_year", "month": "inp_month",
    "min_day_staff": "inp_min_day", "min_day_staff_excl_new": "inp_min_day_excl",
    "night_staff_count": "inp_night_count",
    "max_night_regular": "inp_max_n_reg", "pref_night_regular": "inp_pref_n_reg",
    "max_night_dedicated": "inp_max_n_ded", "pref_night_dedicated": "inp_pref_n_ded",
    "max_consecutive": "inp_max_consec", "pref_consecutive": "inp_pref_consec",
    "solver_time_limit": "inp_time_limit",
}

def _apply_settings(gs_settings):
    """読み込んだ設定を _pending_* キーに保存（ウィジェット描画前に反映させるため）"""
    for src, dst in _SETTINGS_WIDGET_MAP.items():
        v = gs_settings.get(src)
        if v is not None:
            st.session_state[f"_pending_{dst}"] = int(v)
    po = gs_settings.get("public_off_override")
    if po is not None and po != "":
        st.session_state["_pending_inp_po_mode"] = "手動指定"
        st.session_state["_pending_inp_po_val"] = int(po)
    else:
        st.session_state["_pending_inp_po_mode"] = "自動（土日祝）"


def _default_staff():
    return pd.DataFrame({
        "名前": ["スタッフA", "スタッフB", "スタッフC", "スタッフD", "スタッフE"],
        "Tier": ["A", "A", "AB", "C", "C"],
        "夜勤専従": [False, False, False, False, False],
        "時短": [False, False, False, False, False],
        "週勤務": [None, None, None, None, None],
        "前月末": ["", "", "", "", ""],
        "夜勤Min": [None, None, None, None, None],
        "夜勤Max": [None, None, None, None, None],
        "連勤Max": [None, None, None, None, None],
        "勤務曜日": ["", "", "", "", ""],
        "祝日不可": [False, False, False, False, False],
        "土日不可": [False, False, False, False, False],
        "夜勤研修": [False, False, False, False, False],
        "研修夜勤回数": [None, None, None, None, None],
        "新人": [False, False, False, False, False],
        "新人卒業日": [None, None, None, None, None],
    })


# ============================================================
# 様式9帳票生成（看護職員夜勤・交代制勤務に関する実態調査）
# ============================================================
def _generate_youshiki9_excel(schedule, names, tiers, r_dedicated, r_weekly,
                               year, month, night_hours=16,
                               facility_name="", ward_name=""):
    """
    日本看護協会「様式9」帳票を生成。
    夜勤・交代制勤務に関する実態調査 月次提出用フォーム。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "様式9"

    thin = Side(style="thin")
    medium = Side(style="medium")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_bold = Border(left=medium, right=medium, top=medium, bottom=medium)
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    sub_fill = PatternFill("solid", fgColor="D9E2F3")
    sub_font = Font(bold=True, size=9)
    warn_fill = PatternFill("solid", fgColor="F4CCCC")
    ok_fill   = PatternFill("solid", fgColor="E2EFDA")

    shift_system_label = "二交代制" if night_hours == 16 else "三交代制"

    # 列幅設定
    col_widths = {"A": 4, "B": 16, "C": 8, "D": 8, "E": 10, "F": 10,
                  "G": 10, "H": 10, "I": 10, "J": 14, "K": 10}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── タイトル ──
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "看護職員夜勤・交代制勤務に関する実態調査（様式9）"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center")

    # ── 施設情報 ──
    ws.merge_cells("A2:C2")
    ws["A2"] = "調査年月"
    ws["A2"].font = sub_font
    ws.merge_cells("D2:E2")
    ws["D2"] = f"{year}年{month}月"
    ws["D2"].font = Font(bold=True, size=11)

    ws.merge_cells("A3:C3")
    ws["A3"] = "施設名"
    ws["A3"].font = sub_font
    ws.merge_cells("D3:G3")
    ws["D3"] = facility_name or "（施設名）"

    ws.merge_cells("A4:C4")
    ws["A4"] = "病棟・部署名"
    ws["A4"].font = sub_font
    ws.merge_cells("D4:G4")
    ws["D4"] = ward_name or "（病棟・部署名）"

    ws.merge_cells("H2:I2")
    ws["H2"] = "交代制区分"
    ws["H2"].font = sub_font
    ws.merge_cells("J2:K2")
    ws["J2"] = shift_system_label
    ws["J2"].font = Font(bold=True)

    ws.merge_cells("H3:I3")
    ws["H3"] = "夜勤時間/回"
    ws["H3"].font = sub_font
    ws.merge_cells("J3:K3")
    ws["J3"] = f"{night_hours}時間"
    ws["J3"].font = Font(bold=True)

    ws.merge_cells("H4:I4")
    ws["H4"] = "72時間規制 上限"
    ws["H4"].font = sub_font
    ws.merge_cells("J4:K4")
    ws["J4"] = f"{72 // night_hours}回/月"
    ws["J4"].font = Font(bold=True)

    # ── ヘッダー行 (row 6) ──
    headers = ["No", "氏名", "Tier", "雇用区分", "夜勤専従",
               "月夜勤回数", "月夜勤時間(h)", "72h規制", "連続夜勤(最大)",
               "インターバル違反", "判定"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=c, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = bdr

    ws.row_dimensions[6].height = 28

    # ── データ行 ──
    def _max_consecutive_nights(shifts):
        i = 0; consec = 0; mx = 0
        while i < len(shifts):
            if shifts[i] == N:
                consec += 1; mx = max(mx, consec); i += 2
            else:
                consec = 0; i += 1
        return mx

    def _has_interval_violation(shifts):
        return any(shifts[d] == A and shifts[d+1] == D for d in range(len(shifts)-1))

    total_night_count = 0
    total_night_hours = 0
    violation_count = 0

    for idx, s in enumerate(names, 1):
        shifts = schedule[s]
        is_dedicated = r_dedicated.get(s, False)
        is_parttime = r_weekly.get(s) is not None
        night_cnt = shifts.count(N)
        total_h = night_cnt * night_hours
        max_consec = _max_consecutive_nights(shifts)
        iv_violation = _has_interval_violation(shifts)

        # 72h規制判定
        if is_dedicated:
            h72_label = "専従（対象外）"
            h72_ok = True
        elif total_h > 72:
            h72_label = f"🚨 {total_h}h 超過"
            h72_ok = False
        elif total_h >= 64:
            h72_label = f"⚠ {total_h}h"
            h72_ok = True
        else:
            h72_label = f"✅ {total_h}h"
            h72_ok = True

        # インターバル
        iv_label = "🚨 あり" if iv_violation else "✅ なし"

        # 総合判定
        is_violation = (not is_dedicated and total_h > 72) or iv_violation or max_consec > 2
        judgment = "🚨 要対応" if is_violation else "✅ 適合"

        row_data = [
            idx, s, tiers.get(s, "—"),
            "パート" if is_parttime else "常勤",
            "専従" if is_dedicated else "非専従",
            night_cnt, total_h, h72_label,
            max_consec, iv_label, judgment
        ]
        r = 6 + idx
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = bdr
            cell.alignment = Alignment(horizontal="center")
            if c == 2:
                cell.alignment = Alignment(horizontal="left")
            if is_violation:
                cell.fill = warn_fill
            elif idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFF")

        if not is_dedicated:
            total_night_count += night_cnt
            total_night_hours += total_h
        if is_violation:
            violation_count += 1

    # ── 集計行 ──
    summary_row = 6 + len(names) + 1
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    ws[f"A{summary_row}"] = "集計"
    ws[f"A{summary_row}"].font = Font(bold=True)
    ws[f"A{summary_row}"].fill = sub_fill

    ws.cell(row=summary_row, column=5, value=f"対象: {len(names)}名").font = Font(size=9)
    ws.cell(row=summary_row, column=6, value=f"合計: {total_night_count}回").font = Font(size=9, bold=True)
    ws.cell(row=summary_row, column=7, value=f"合計: {total_night_hours}h").font = Font(size=9, bold=True)
    ws.cell(row=summary_row, column=11,
            value=f"🚨 {violation_count}件" if violation_count else "✅ 全員適合")
    ws.cell(row=summary_row, column=11).font = Font(bold=True,
            color="CC0000" if violation_count else "006600")

    for c in range(1, 12):
        ws.cell(row=summary_row, column=c).border = bdr
        ws.cell(row=summary_row, column=c).fill = sub_fill

    # ── 注釈 ──
    note_row = summary_row + 2
    ws.merge_cells(f"A{note_row}:K{note_row}")
    ws[f"A{note_row}"] = (
        "【根拠】日本看護協会「夜勤・交代制勤務に関するガイドライン」(2013年) ／ "
        "夜勤回数: 月8回以内 ／ 72時間規制: 月夜勤時間72時間以内 ／ "
        "連続夜勤: 2サイクル以内 ／ 勤務間隔: 11時間以上"
    )
    ws[f"A{note_row}"].font = Font(size=8, color="888888", italic=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# Excelテンプレート生成
# ============================================================
def _generate_template_excel(year, month, num_staff=20):
    """入力用Excelテンプレートを生成（スタッフ情報 + 勤務希望の2シート構成）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    num_days = calendar.monthrange(year, month)[1]
    thin = Side(style="thin")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    # スタッフ情報エリア: 薄緑
    staff_hdr_fill = PatternFill("solid", fgColor="548235")
    staff_hdr_font = Font(bold=True, color="FFFFFF", size=10)
    staff_cell_fill = PatternFill("solid", fgColor="E2EFDA")
    # 勤務希望エリア: 青
    req_hdr_font = Font(bold=True, color="FFFFFF", size=9)
    req_cell_fill = PatternFill("solid", fgColor="D6E4F0")

    # --- Sheet 1: 設定 ---
    ws_s = wb.active
    ws_s.title = "設定"
    ws_s.column_dimensions["A"].width = 22
    ws_s.column_dimensions["B"].width = 14
    ws_s.column_dimensions["C"].width = 35
    # ヘッダー
    for c, txt in enumerate(["項目", "値", "説明"], 1):
        cell = ws_s.cell(row=3, column=c, value=txt)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.border = bdr
        cell.alignment = Alignment(horizontal="center")
    ws_s.cell(row=1, column=1, value="勤務表設定").font = Font(bold=True, size=14)
    # データ
    for i, (label, default, desc) in enumerate(SETTINGS_DEF):
        r = 4 + i
        ws_s.cell(row=r, column=1, value=label).border = bdr
        ws_s.cell(row=r, column=2, value=default).border = bdr
        ws_s.cell(row=r, column=3, value=desc).border = bdr
        ws_s.cell(row=r, column=3).font = Font(color="888888", size=9)

    # コンプライアンス・運用条件（参考情報 — アプリ側のサイドバーで設定）
    comp_start = 4 + len(SETTINGS_DEF) + 2
    ws_s.cell(row=comp_start, column=1, value="コンプライアンス・運用条件（参考）").font = Font(
        bold=True, size=12, color="C00000")
    ws_s.cell(row=comp_start + 1, column=1, value="※以下はアプリのサイドバーで設定します（Excel上では参考情報）").font = Font(
        color="888888", size=9)
    comp_items = [
        ("72時間規制", "strict / soft / none",
         "strict=必ず準拠, soft=なるべく準拠(ペナルティ), none=チェックのみ"),
        ("長日勤→翌日短夜勤", "strict / soft / none",
         "LD(12h)の翌日にSN(12h)を入れるか"),
        ("長日勤の連続禁止", "strict / soft / none",
         "LD(12h)を2日連続で割り当てるか"),
        ("夜勤時間数", "16",
         "1夜勤あたりの時間数（二交代=16h）"),
        ("ユニット種別", "ICU",
         "ICU/HCU/NICU/GCU/SCU/PICU/一般病棟等"),
        ("病床数", "10",
         "人員配置基準の計算に使用"),
    ]
    for c, txt in enumerate(["項目", "設定値", "説明"], 1):
        cell = ws_s.cell(row=comp_start + 2, column=c, value=txt)
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center")
    for i, (label, default, desc) in enumerate(comp_items):
        r = comp_start + 3 + i
        ws_s.cell(row=r, column=1, value=label).border = bdr
        ws_s.cell(row=r, column=2, value=default).border = bdr
        ws_s.cell(row=r, column=2).font = Font(color="666666")
        ws_s.cell(row=r, column=3, value=desc).border = bdr
        ws_s.cell(row=r, column=3).font = Font(color="888888", size=9)

    # === 勤務種別設定セクション ===
    shift_type_start = comp_start + 3 + len(comp_items) + 2
    ws_s.cell(row=shift_type_start, column=1, value="勤務種別設定").font = Font(
        bold=True, size=12, color="1F4E79")
    ws_s.cell(row=shift_type_start + 1, column=1,
              value="※「使用」列に○を入れると、その勤務種別がスケジュールに使用されます").font = Font(
        color="888888", size=9)
    ws_s.column_dimensions["D"].width = 12
    for c, txt in enumerate(["記号", "名称", "時間帯", "使用"], 1):
        cell = ws_s.cell(row=shift_type_start + 2, column=c, value=txt)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center")
    _shift_type_rows = [
        ("日", "日勤", "8:00〜17:00", "○"),
        ("夜", "夜勤", "16:45〜翌9:00 (16h)", "○"),
        ("準", "短夜勤", "17:00〜翌5:00 (12h)", ""),   # ICU二交代では不使用
        ("早", "早出", "7:00〜16:00", ""),               # ICU二交代では不使用
        ("遅", "遅出", "12:00〜21:00", ""),              # ICU二交代では不使用
        ("長", "長日勤", "8:45〜21:00 (12h)", ""),       # ICU二交代では不使用
        ("短", "時短", "8:45〜16:00 (6.25h)", "○"),     # 時短スタッフ用
        ("休", "公休", "—", "○"),
        ("研", "研修", "—", "○"),
        ("委", "委員会(勤務中に離席/A・AB同席必須)", "—", "○"),
        ("夜不", "夜勤不可(希望専用)", "—", "○"),
        ("休暇", "有給休暇(希望専用)", "—", "○"),
        ("明休", "明または休(希望専用)", "—", "○"),
    ]
    for i, (sym, name, hours, use) in enumerate(_shift_type_rows):
        r = shift_type_start + 3 + i
        ws_s.cell(row=r, column=1, value=sym).border = bdr
        ws_s.cell(row=r, column=1).font = Font(bold=True)
        ws_s.cell(row=r, column=2, value=name).border = bdr
        ws_s.cell(row=r, column=3, value=hours).border = bdr
        ws_s.cell(row=r, column=3).font = Font(color="888888", size=9)
        ws_s.cell(row=r, column=4, value=use).border = bdr
        ws_s.cell(row=r, column=4).alignment = Alignment(horizontal="center")

    # === 共通データ ===
    staff_headers = ["名前", "Tier", "夜勤専従", "時短", "週勤務", "前月末",
                     "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可",
                     "夜勤研修", "研修夜勤回数", "新人", "新人卒業日"]
    n_staff_cols = len(staff_headers)
    weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]
    first_wd = date(year, month, 1).weekday()
    holidays = {d.day for d, _ in jpholiday.month_holidays(year, month)}

    samples_name = ["山田太郎", "佐藤花子", "鈴木一郎"]
    samples_tier = ["A", "AB", "C"]
    samples_extra = [
        [None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        [None, None, "3", None, None, None, None, "月水金", None, None, None, None, None, None],
    ]
    total_rows = max(num_staff, len(samples_name))

    # ゼブラストライプ用の交互色
    staff_fill_even = PatternFill("solid", fgColor="E2EFDA")
    staff_fill_odd  = PatternFill("solid", fgColor="F5FAF0")
    req_fill_even   = PatternFill("solid", fgColor="D6E4F0")
    req_fill_odd    = PatternFill("solid", fgColor="EFF5FB")

    # ================================================================
    # Sheet 2: スタッフ情報（原本）
    # ================================================================
    ws_staff = wb.create_sheet("スタッフ情報")
    ws_staff.cell(row=1, column=1, value=f"👤 スタッフ情報 — {year}年{month}月").font = Font(bold=True, size=14)
    ws_staff.cell(row=2, column=1, value="※ このシートが原本です。名前・Tierは勤務希望シートに自動反映されます。").font = Font(color="888888", size=9)

    # ヘッダー (row 3)
    for c, txt in enumerate(staff_headers, 1):
        cell = ws_staff.cell(row=3, column=c, value=txt)
        cell.fill = staff_hdr_fill; cell.font = staff_hdr_font; cell.border = bdr
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for i in range(total_rows):
        r = 4 + i
        _s_fill = staff_fill_even if i % 2 == 0 else staff_fill_odd
        # 名前
        name_val = samples_name[i] if i < len(samples_name) else f"スタッフ{i + 1}"
        cell = ws_staff.cell(row=r, column=1, value=name_val)
        cell.border = bdr; cell.fill = _s_fill
        # Tier
        tier_val = samples_tier[i] if i < len(samples_tier) else None
        cell = ws_staff.cell(row=r, column=2)
        cell.value = tier_val; cell.border = bdr; cell.fill = _s_fill
        # 残りのカラム (夜勤専従〜祝日不可)
        for c_idx in range(3, n_staff_cols + 1):
            extra_val = None
            if i < len(samples_extra):
                extra_val = samples_extra[i][c_idx - 3]
            cell = ws_staff.cell(row=r, column=c_idx)
            cell.value = extra_val if extra_val is not None and extra_val != "" and extra_val != 0 else None
            cell.border = bdr; cell.fill = _s_fill

    # 列幅
    ws_staff.column_dimensions["A"].width = 14
    for i in range(1, n_staff_cols):
        ws_staff.column_dimensions[get_column_letter(i + 1)].width = 10
    ws_staff.freeze_panes = "B4"

    # --- 凡例エリア ---
    legend_start_row = 4 + total_rows + 2
    ws_staff.cell(row=legend_start_row, column=1, value="📖 Tier定義").font = Font(bold=True, size=11, color="548235")
    tier_defs = [
        ("A", "ベテラン・リーダー格（日勤/夜勤リーダー単独可）"),
        ("AB", "中堅・リーダー代行可（夜勤リーダー可）"),
        ("B", "一人立ち済み（B+B, B+C族の夜勤ペアは禁止）"),
        ("C+", "C既卒（C+C+, C++Cペア禁止／A/AB/B下で夜勤可）"),
        ("C", "新人・経験浅い（必ずA/AB/Bと夜勤ペア）"),
    ]
    for i, (tier, desc) in enumerate(tier_defs):
        ws_staff.cell(row=legend_start_row + 1 + i, column=1, value=tier).font = Font(bold=True)
        ws_staff.cell(row=legend_start_row + 1 + i, column=2, value=desc).font = Font(color="555555", size=9)
    ws_staff.cell(row=legend_start_row + 7, column=1, value="📖 カラム説明").font = Font(bold=True, size=11, color="548235")
    col_defs = [
        ("夜勤専従", "ONで夜勤/明け/休のみのパターン"),
        ("時短", "ONで時短(ST)/早出/遅出/休のみ"),
        ("週勤務", "パートの週勤務日数（空欄=フルタイム）"),
        ("前月末", "前月末の勤務状態（夜/明）"),
        ("夜勤Min/Max", "個別の夜勤回数制限（空欄=全体設定）"),
        ("連勤Max", "最大連続勤務日数（空欄=全体設定）"),
        ("勤務曜日", "特定曜日のみ勤務（例:月水金）"),
        ("祝日不可", "ONで祝日に勤務を入れない"),
        ("土日不可", "ONで土日に勤務を入れない"),
        ("夜勤研修", "ONで夜勤研修中（通常2人+研修1人=MAX3人）"),
        ("研修夜勤回数", "月間の研修夜勤上限（空欄=制限なし）"),
        ("新人", "ONで新人扱い（日勤頭数には含むがリーダー・独立C判定から除外）"),
        ("新人卒業日", "その日まで新人、翌日から通常運用（空欄=月末まで新人）"),
    ]
    for i, (col_name, desc) in enumerate(col_defs):
        ws_staff.cell(row=legend_start_row + 8 + i, column=1, value=col_name).font = Font(bold=True, size=9)
        ws_staff.cell(row=legend_start_row + 8 + i, column=2, value=desc).font = Font(color="555555", size=9)

    # ================================================================
    # Sheet 3: 勤務希望（名前・Tierはスタッフ情報から数式参照）
    # ================================================================
    ws_req = wb.create_sheet("勤務希望")
    ws_req.cell(row=1, column=1, value=f"📝 勤務希望 — {year}年{month}月").font = Font(bold=True, size=14)
    ws_req.cell(row=2, column=1, value="名前・Tierはスタッフ情報シートから自動同期。日付セルにシフト記号を入力してください。").font = Font(color="888888", size=9)

    # ヘッダー行 (row 3): 名前 | Tier | 1(月) | 2(火) | ...
    # row 3 ヘッダー
    req_headers_fixed = ["名前", "Tier"]
    for c, txt in enumerate(req_headers_fixed, 1):
        cell = ws_req.cell(row=3, column=c, value=txt)
        cell.fill = staff_hdr_fill; cell.font = staff_hdr_font; cell.border = bdr
        cell.alignment = Alignment(horizontal="center")

    day_start_col = len(req_headers_fixed) + 1  # 3列目から日付
    for d in range(1, num_days + 1):
        col = day_start_col + d - 1
        wd_name = weekdays_jp[(first_wd + d - 1) % 7]
        cell = ws_req.cell(row=3, column=col, value=f"{d}({wd_name})")
        cell.alignment = Alignment(horizontal="center")
        cell.border = bdr
        wd_idx = (first_wd + d - 1) % 7
        if d in holidays:
            cell.fill = PatternFill("solid", fgColor="F4CCCC")
            cell.font = Font(bold=True, color="CC0000", size=9)
        elif wd_idx >= 5:  # 土日
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(bold=True, color="FFFFFF", size=9)
        else:  # 平日
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
            cell.font = Font(bold=True, color="1F4E79", size=9)

    # データ行: 名前 = =スタッフ情報!A4, Tier = =スタッフ情報!B4
    ref_fill = PatternFill("solid", fgColor="F2F2F2")  # 参照セルはグレー背景
    ref_font = Font(color="888888", size=10)
    for i in range(total_rows):
        r = 4 + i
        staff_row = 4 + i  # スタッフ情報シートの対応行
        _r_fill = req_fill_even if i % 2 == 0 else req_fill_odd

        # 名前 — 数式で参照（空セルなら空白表示、編集不可グレー背景）
        cell_name = ws_req.cell(row=r, column=1)
        cell_name.value = f'=IF(スタッフ情報!A{staff_row}="","",スタッフ情報!A{staff_row})'
        cell_name.border = bdr; cell_name.fill = ref_fill; cell_name.font = ref_font
        # Tier — 数式で参照（空セルなら空白表示）
        cell_tier = ws_req.cell(row=r, column=2)
        cell_tier.value = f'=IF(スタッフ情報!B{staff_row}="","",スタッフ情報!B{staff_row})'
        cell_tier.border = bdr; cell_tier.fill = ref_fill; cell_tier.font = ref_font
        cell_tier.alignment = Alignment(horizontal="center")
        # 勤務希望セル — 空欄
        for d in range(1, num_days + 1):
            cell = ws_req.cell(row=r, column=day_start_col + d - 1)
            cell.value = None
            cell.border = bdr; cell.fill = _r_fill

    # 列幅
    ws_req.column_dimensions["A"].width = 14
    ws_req.column_dimensions["B"].width = 8
    for d in range(1, num_days + 1):
        ws_req.column_dimensions[get_column_letter(day_start_col + d - 1)].width = 7
    ws_req.freeze_panes = "C4"  # 名前+Tier固定、日付スクロール

    # --- シフト種別凡例 ---
    req_legend_row = 4 + total_rows + 2
    ws_req.cell(row=req_legend_row, column=1, value="📖 シフト種別").font = Font(bold=True, size=11, color="1F4E79")
    shift_legend = [
        ("日", "日勤 (8:00〜17:00)"), ("夜", "夜勤 (16:45〜翌9:00 / 16h)"),
        ("準", "短夜勤 (17:00〜翌5:00 / 12h)"), ("早", "早出 (7:00〜16:00)"),
        ("遅", "遅出 (12:00〜21:00)"), ("長", "長日勤 (8:45〜21:00 / 12h)"),
        ("短", "時短 (8:45〜16:00 / 6.25h)"), ("休", "公休"),
        ("研", "研修"), ("委", "委員会 (A/AB同席必須)"),
        ("夜不", "この日は夜勤不可"),
        ("休暇", "有給休暇"), ("明休", "明または休"),
    ]
    for i, (sym, desc) in enumerate(shift_legend):
        ws_req.cell(row=req_legend_row + 1 + i, column=1, value=sym).font = Font(bold=True, size=10)
        ws_req.cell(row=req_legend_row + 1 + i, column=2, value=desc).font = Font(color="555555", size=9)

    # シート保護: 勤務希望シートの名前・Tier列を保護（数式を壊さないように）
    from openpyxl.worksheet.protection import SheetProtection
    ws_req.protection = SheetProtection(sheet=True, objects=True, scenarios=True,
                                         formatColumns=False, formatRows=False)
    # 名前・Tier列はロック、日付セルはロック解除
    from openpyxl.styles import Protection as CellProtection
    locked = CellProtection(locked=True)
    unlocked = CellProtection(locked=False)
    for i in range(total_rows):
        r = 4 + i
        ws_req.cell(row=r, column=1).protection = locked
        ws_req.cell(row=r, column=2).protection = locked
        for d in range(1, num_days + 1):
            ws_req.cell(row=r, column=day_start_col + d - 1).protection = unlocked

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# Excelアップロード解析
# ============================================================
def _parse_uploaded_excel(uploaded_file, year, month):
    """アップロードされたExcelを解析してスタッフ・希望・設定を返す"""
    from openpyxl import load_workbook
    wb = load_workbook(uploaded_file, data_only=True)
    num_days = calendar.monthrange(year, month)[1]

    # 設定シート
    gs_settings = {}
    if "設定" in wb.sheetnames:
        ws = wb["設定"]
        rows = []
        for r in range(4, 4 + len(SETTINGS_KEYS)):
            row_vals = [("" if ws.cell(row=r, column=c).value is None else ws.cell(row=r, column=c).value) for c in range(1, 4)]
            rows.append(row_vals)
        gs_settings = _parse_settings(rows)

        # --- 勤務種別設定セクションの読み取り ---
        _all_shift_syms = {"日", "夜", "準", "早", "遅", "長", "短", "休", "研", "委", "夜不", "休暇", "明休"}
        _found_shift_section = False
        _enabled = []
        for r in range(4 + len(SETTINGS_KEYS), ws.max_row + 1):
            sym_val = ws.cell(row=r, column=1).value
            if sym_val is not None:
                sym_str = str(sym_val).strip()
                if sym_str in _all_shift_syms:
                    _found_shift_section = True
                    use_val = ws.cell(row=r, column=4).value
                    if use_val and str(use_val).strip() == "○":
                        _enabled.append(sym_str)
        if _found_shift_section:
            gs_settings["enabled_shifts"] = _enabled
        else:
            # 旧テンプレート: 全種別有効
            gs_settings["enabled_shifts"] = list(_all_shift_syms)

    # --- テンプレ形式の自動検出（3種対応） ---
    # 1. 新形式: 「スタッフ情報」+「勤務希望」（2シート分離、数式参照）
    # 2. 旧統合形式: 「スタッフ・勤務希望」（1シート統合）
    # 3. 旧分離形式: 「スタッフ一覧」+「勤務希望」
    staff_list = []
    reqs = {}
    n_staff_cols = 16  # 名前〜新人卒業日

    if "スタッフ情報" in wb.sheetnames:
        # === 新形式（スタッフ情報 + 勤務希望 2シート） ===
        ws_si = wb["スタッフ情報"]
        staff_rows = []
        for r in range(4, ws_si.max_row + 1):  # row 4 からデータ
            name_val = ws_si.cell(row=r, column=1).value
            if not name_val or not str(name_val).strip():
                continue
            s_vals = [("" if ws_si.cell(row=r, column=c).value is None else ws_si.cell(row=r, column=c).value) for c in range(1, n_staff_cols + 1)]
            staff_rows.append(s_vals)
        if staff_rows:
            staff_list = _parse_staff_list(staff_rows)
        staff_names = [s.name for s in staff_list]

        if "勤務希望" in wb.sheetnames:
            ws_rq = wb["勤務希望"]
            req_rows = []
            day_start = 3  # 勤務希望シートは 列C(3) から日付
            for r in range(4, ws_rq.max_row + 1):  # row 4 からデータ
                # 名前列は数式（=スタッフ情報!A4）→ data_only=Trueで未計算だとNone
                # → スタッフ情報シートから取得済みの名前を行番号で対応させる
                row_idx = r - 4  # 0-based index
                name_val = ws_rq.cell(row=r, column=1).value
                # 数式未計算の場合、スタッフ情報の名前リストからfallback
                if (not name_val or not str(name_val).strip()) and row_idx < len(staff_names):
                    name_val = staff_names[row_idx]
                if not name_val or not str(name_val).strip():
                    continue
                r_vals = [str(name_val).strip()]
                for d in range(1, num_days + 1):
                    _v = ws_rq.cell(row=r, column=day_start + d - 1).value
                    r_vals.append("" if _v is None else _v)
                req_rows.append(r_vals)
            if req_rows:
                reqs = _parse_requests(req_rows, staff_names, num_days)

    elif "スタッフ・勤務希望" in wb.sheetnames:
        # === 旧統合形式 ===
        ws = wb["スタッフ・勤務希望"]
        staff_rows = []
        req_rows = []
        for r in range(5, ws.max_row + 1):
            name_val = ws.cell(row=r, column=1).value
            if not name_val or not str(name_val).strip():
                continue
            s_vals = [("" if ws.cell(row=r, column=c).value is None else ws.cell(row=r, column=c).value) for c in range(1, n_staff_cols + 2)]
            staff_rows.append(s_vals)
            day_start = n_staff_cols + 1
            r_vals = [str(name_val).strip()]
            for d in range(1, num_days + 1):
                _v = ws.cell(row=r, column=day_start + d - 1).value
                r_vals.append("" if _v is None else _v)
            req_rows.append(r_vals)
        if staff_rows:
            staff_list = _parse_staff_list(staff_rows)
        staff_names = [s.name for s in staff_list]
        if req_rows:
            reqs = _parse_requests(req_rows, staff_names, num_days)
    else:
        # === 旧分離形式 ===
        if "スタッフ一覧" in wb.sheetnames:
            ws = wb["スタッフ一覧"]
            staff_rows = []
            for r in range(2, ws.max_row + 1):
                row_vals = [("" if ws.cell(row=r, column=c).value is None else ws.cell(row=r, column=c).value) for c in range(1, 13)]
                if str(row_vals[0]).strip():
                    staff_rows.append(row_vals)
            staff_list = _parse_staff_list(staff_rows)
        if "勤務希望" in wb.sheetnames:
            ws = wb["勤務希望"]
            staff_names = [s.name for s in staff_list]
            req_rows = []
            for r in range(5, ws.max_row + 1):
                row_vals = [("" if ws.cell(row=r, column=c).value is None else ws.cell(row=r, column=c).value) for c in range(1, num_days + 2)]
                if str(row_vals[0]).strip():
                    req_rows.append(row_vals)
            reqs = _parse_requests(req_rows, staff_names, num_days)

    return staff_list, reqs, gs_settings


# ============================================================
# セッション初期化
# ============================================================
if "staff_df" not in st.session_state:
    st.session_state.staff_df = _default_staff()
if "requests_df" not in st.session_state:
    st.session_state.requests_df = None
if "results" not in st.session_state:
    st.session_state.results = None
if "data_loaded" not in st.session_state:
    st.session_state.data_loaded = False

# ============================================================
# pending設定をウィジェットキーに反映（ウィジェット描画前に実行）
# ============================================================
for _src, _dst in _SETTINGS_WIDGET_MAP.items():
    _pk = f"_pending_{_dst}"
    if _pk in st.session_state:
        st.session_state[_dst] = st.session_state.pop(_pk)
for _pk, _wk in [("_pending_inp_po_mode", "inp_po_mode"), ("_pending_inp_po_val", "inp_po_val")]:
    if _pk in st.session_state:
        st.session_state[_wk] = st.session_state.pop(_pk)

# ============================================================
# サイドバー: 設定
# ============================================================
st.sidebar.title("勤務表設定")

col_y, col_m = st.sidebar.columns(2)
year = col_y.number_input("対象年", 2024, 2030, 2026, key="inp_year")
month = col_m.number_input("対象月", 1, 12, 5, key="inp_month")

num_days = calendar.monthrange(year, month)[1]
holidays_auto, weekends_auto, auto_public_off = _get_holidays_and_days_off(year, month)

st.sidebar.markdown("---")
st.sidebar.subheader("勤務条件")

public_off_mode = st.sidebar.radio("公休日数", ["自動（土日祝）", "手動指定"],
                                    horizontal=True, key="inp_po_mode")
if public_off_mode == "自動（土日祝）":
    public_off = auto_public_off
    st.sidebar.info(f"公休 {public_off}日（土日{len(weekends_auto)} + 祝{len(holidays_auto)}）")
    po_override = None
else:
    public_off = st.sidebar.number_input("公休日数", 0, num_days, auto_public_off, key="inp_po_val")
    po_override = public_off

min_day = st.sidebar.number_input("日勤最低人数（全体）", 1, 20, 5, key="inp_min_day",
                                  help="新人込みの1日あたり最低日勤人数")
min_day_excl = st.sidebar.number_input("日勤最低人数（新人除く）", 0, 20, 4, key="inp_min_day_excl",
                                        help="新人を除いた経験者の1日あたり最低人数")
night_count = st.sidebar.number_input("夜勤人数/日", 1, 5, 2, key="inp_night_count")

st.sidebar.markdown("---")
st.sidebar.subheader("夜勤設定")
col1, col2 = st.sidebar.columns(2)
max_n_reg = col1.number_input("通常 上限", 1, 15, 5, key="inp_max_n_reg")
pref_n_reg = col2.number_input("通常 推奨", 1, 15, 4, key="inp_pref_n_reg")
col3, col4 = st.sidebar.columns(2)
max_n_ded = col3.number_input("専従 上限", 1, 20, 10, key="inp_max_n_ded")
pref_n_ded = col4.number_input("専従 推奨", 1, 20, 9, key="inp_pref_n_ded")

st.sidebar.markdown("---")
st.sidebar.subheader("連勤設定")
col5, col6 = st.sidebar.columns(2)
max_consec = col5.number_input("最大連勤", 1, 10, 5, key="inp_max_consec")
pref_consec = col6.number_input("推奨連勤", 1, 10, 4, key="inp_pref_consec")

time_limit = st.sidebar.number_input("計算時間上限（秒）", 10, 600, 120, key="inp_time_limit")
num_patterns = st.sidebar.number_input("生成パターン数", 1, 5, 3, key="inp_num_patterns")

st.sidebar.markdown("---")
st.sidebar.subheader("⚖ コンプライアンス設定")
# ICU機能制限版: 二交代固定
shift_system = "二交代（16h夜勤）"
st.sidebar.info("🏥 二交代制（16h夜勤）固定")
night_hours = 16

_72h_limit = 72 // night_hours
night_72h_mode_label = st.sidebar.radio(
    f"72時間規制（≦{_72h_limit}回/月）",
    ["🚫 必ず準拠（ハード制約）", "⚠ なるべく準拠（ソフト制約）", "✅ 許容（チェックのみ）"],
    index=2,
    key="inp_72h_mode",
)
_72h_mode_map = {
    "🚫 必ず準拠（ハード制約）": "strict",
    "⚠ なるべく準拠（ソフト制約）": "soft",
    "✅ 許容（チェックのみ）": "none",
}
night_72h_mode = _72h_mode_map[night_72h_mode_label]
if night_72h_mode == "strict":
    st.sidebar.caption(f"✅ {night_hours}h×夜勤回数 を必ず72h以内に収めます（上限{_72h_limit}回）")
elif night_72h_mode == "soft":
    st.sidebar.caption(f"⚠ 72h超過を最小化しますが、達成できない場合は許容します")
else:
    st.sidebar.caption(f"✅ 制約なし（結果画面でのみ警告表示）")

# ICU機能制限版: LD/SN不使用のため運用ルールは固定
op_rules = {
    "ld_sn": "strict",         # LD/SN自体が無効なので影響なし
    "ld_consecutive": "strict",
}

st.sidebar.markdown("---")
st.sidebar.subheader("🔄 使用シフト種別")
# ICU二交代制デフォルト: 日/夜/短(時短)/休/研 + 希望専用(夜不/休暇/明休)
_all_shift_symbols = ["日", "夜", "準", "早", "遅", "長", "短", "休", "研", "委", "夜不", "休暇", "明休"]
_icu_default = ["日", "夜", "短", "休", "研", "委", "夜不", "休暇", "明休"]
_default_enabled = st.session_state.get("enabled_shifts", _icu_default)
enabled_shifts = st.sidebar.multiselect(
    "有効なシフト種別",
    options=_all_shift_symbols,
    default=[s for s in _default_enabled if s in _all_shift_symbols],
    key="inp_enabled_shifts",
    help="ICU二交代制: 日勤・夜勤が基本。時短は時短スタッフ用。"
)
st.session_state.enabled_shifts = enabled_shifts

st.sidebar.markdown("---")
st.sidebar.subheader("🏥 人員配置基準")
# ICU機能制限版: ICU固定
unit_type = "ICU（特定集中治療室管理料 1〜4）"
st.sidebar.info(f"🏥 {unit_type}")
_ratio_val, _check_night, _ratio_label, _basis_label = UNIT_STANDARDS[unit_type]
bed_count = st.sidebar.number_input(
    "病床数（床）", min_value=1, max_value=200, value=4,
    key="inp_bed_count"
)
_required_nurses = max(1, -(-bed_count // _ratio_val))
_scope = "日勤・夜勤（24時間）" if _check_night else "日勤帯"
st.sidebar.caption(
    f"基準: {_ratio_label} ／ {_scope}\n"
    f"必要人数: {bed_count}床 ÷ {_ratio_val} = **{_required_nurses}人以上**"
)
st.sidebar.caption(f"根拠: {_basis_label}")
nurse_ratio = _ratio_label  # 表示用

settings = {
    "year": year, "month": month,
    "public_off_override": po_override,
    "min_day_staff": min_day, "min_day_staff_excl_new": min_day_excl,
    "night_staff_count": night_count,
    "max_night_regular": max_n_reg, "pref_night_regular": pref_n_reg,
    "max_night_dedicated": max_n_ded, "pref_night_dedicated": pref_n_ded,
    "max_consecutive": max_consec, "pref_consecutive": pref_consec,
    "solver_time_limit": time_limit, "holidays": "",
}

# ============================================================
# メインエリア
# ============================================================
st.title("🏥 ICU勤務表自動作成ツール（二交代制）")

tab0, tab1, tab3, tab4, tab5 = st.tabs(
    ["📂 データ入力", "📋 スタッフ・勤務希望", "⚡ 生成", "📊 結果", "🏠 ダッシュボード"])

# ============================================================
# Tab 0: データ入力（テンプレ出力 / アップロード / スプシ）
# ============================================================
with tab0:
    st.subheader("データの入力方法を選んでください")

    col_left, col_right = st.columns(2)

    # --- 左カラム: テンプレートダウンロード ---
    with col_left:
        st.markdown("### 📥 テンプレートをダウンロード")
        st.markdown("""
        1. スタッフ人数を入力してテンプレートをダウンロード
        2. **スタッフ情報・勤務希望** を記入
        3. 右の「データ読み込み」で読み込み
        """)
        tmpl_staff_count = st.number_input(
            "テンプレートのスタッフ人数", min_value=1, max_value=100,
            value=15, step=1, key="tmpl_staff_count")
        template_bytes = _generate_template_excel(year, month, num_staff=tmpl_staff_count)
        st.download_button(
            label=f"📄 Excelテンプレート（{year}年{month}月・{tmpl_staff_count}人）",
            data=template_bytes,
            file_name=f"勤務表テンプレート_{year}_{month:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        # --- Googleスプレッドシートへテンプレート出力 ---
        if st.button("📊 Googleスプレッドシートにテンプレート作成", use_container_width=True, key="btn_gsheet_tmpl"):
            try:
                from shift_scheduler import _get_gc as _get_gsheet_client
                gc = _get_gsheet_client()
                title = f"勤務表テンプレート_{year}_{month:02d}"
                sh = gc.create(title)
                # --- 設定シート ---
                ws_s = sh.sheet1
                ws_s.update_title("設定")
                s_rows = [["項目", "値", "説明"]]
                for label, default, desc in SETTINGS_DEF:
                    s_rows.append([label, default, desc])
                ws_s.update(s_rows, "A1")
                # --- スタッフ一覧シート ---
                ws_st = sh.add_worksheet("スタッフ一覧", rows=30, cols=20)
                st_rows = [["名前", "Tier", "夜勤専従", "時短", "週勤務", "前月末",
                            "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可",
                            "夜勤研修", "研修夜勤回数", "新人", "新人卒業日", "", "Tier定義", "説明"]]
                st_rows.append(["山田太郎", "A", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                                "", "A", "ベテラン・リーダー格"])
                st_rows.append(["佐藤花子", "AB", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                                "", "AB", "中堅・リーダー代行可"])
                st_rows.append(["鈴木一郎", "C", "", "", "3", "", "", "", "", "月水金", "", "", "", "", "", "",
                                "", "B", "一人立ち済み"])
                st_rows.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                                "", "C", "新人・経験浅い"])
                ws_st.update(st_rows, "A1")
                # --- 勤務希望シート ---
                _num_days = calendar.monthrange(year, month)[1]
                ws_r = sh.add_worksheet("勤務希望", rows=30, cols=_num_days + 40)
                _wdj = ["月", "火", "水", "木", "金", "土", "日"]
                _fwd = date(year, month, 1).weekday()
                r_hdr = ["名前"] + [f"{d}({_wdj[(_fwd+d-1)%7]})" for d in range(1, _num_days+1)]
                # 凡例列
                _ref_col = _num_days + 2
                r_hdr += [""] + ["記号", "説明"]
                ws_r.update([r_hdr], "A1")
                legend = [
                    ["日", "日勤(8:00-17:00)"], ["夜", "夜勤(16:45-翌9:00)"],
                    ["準", "短夜勤(17:00-翌5:00)"], ["早", "早出(7:00-16:00)"],
                    ["遅", "遅出(12:00-21:00)"], ["長", "長日勤(8:45-21:00)"],
                    ["短", "時短(8:45-16:00)"], ["休", "公休"],
                    ["研", "研修"], ["委", "委員会(A/AB同席必須)"],
                    ["夜不", "夜勤不可"], ["休暇", "有給休暇"],
                    ["明休", "明または休"],
                ]
                import gspread.utils as _gu
                _start = _gu.rowcol_to_a1(2, _num_days + 3)
                ws_r.update([[s, d] for s, d in legend], _start)

                url = sh.url
                st.success(f"✅ Googleスプレッドシートを作成しました")
                st.markdown(f"[📊 スプレッドシートを開く]({url})")
            except ImportError:
                st.error("gspread/google-authが必要です: pip install gspread google-auth")
            except Exception as e:
                st.error(f"作成エラー: {e}\n\n認証設定を確認してください（.streamlit/secrets.toml または credentials.json）")

    # --- 右カラム: データ読み込み ---
    with col_right:
        st.markdown("### 📤 データを読み込む")
        import_method = st.radio("読み込み方法", ["Excel ファイル", "Google スプレッドシート"],
                                  horizontal=True, key="import_method")

        # ローカルデフォルトファイルが存在する場合に表示
        import os as _os
        _default_excel = _os.path.join(_os.path.dirname(__file__), "勤務表_入力.xlsx")
        if _os.path.exists(_default_excel):
            if st.button("📁 ローカルファイルを読み込む（勤務表_入力.xlsx）",
                         use_container_width=True, key="btn_load_local"):
                try:
                    with st.spinner("読み込み中..."):
                        from shift_scheduler import load_input as _load_local
                        staff_list, reqs, gs_settings = _load_local()
                    if gs_settings:
                        _apply_settings(gs_settings)
                    st.session_state.staff_df = _staff_to_df(staff_list)
                    st.session_state.requests_df = _reqs_to_df(reqs, staff_list, num_days)
                    st.session_state.data_loaded = True
                    st.success(f"✅ {len(staff_list)}人のスタッフを読み込みました")
                    st.rerun()
                except Exception as e:
                    st.error(f"読み込みエラー: {e}")
            st.divider()

        if import_method == "Excel ファイル":
            uploaded = st.file_uploader("Excelファイルを選択", type=["xlsx", "xlsm"],
                                         key="excel_upload")
            if uploaded is not None:
                if st.button("📤 Excelから読み込み", type="primary", use_container_width=True,
                             key="btn_load_excel"):
                    try:
                        with st.spinner("読み込み中..."):
                            staff_list, reqs, gs_settings = _parse_uploaded_excel(
                                uploaded, year, month)
                        if gs_settings:
                            _apply_settings(gs_settings)
                        st.session_state.enabled_shifts = gs_settings.get(
                            "enabled_shifts",
                            ["日", "夜", "短", "休", "研", "委", "夜不", "休暇", "明休"])
                        st.session_state.staff_df = _staff_to_df(staff_list)
                        st.session_state.requests_df = _reqs_to_df(reqs, staff_list, num_days)
                        st.session_state.data_loaded = True
                        st.success(f"✅ {len(staff_list)}人のスタッフと希望を読み込みました")
                        st.rerun()
                    except Exception as e:
                        st.error(f"読み込みエラー: {e}")

        else:  # Google スプレッドシート
            gsheet_id = st.text_input(
                "スプレッドシートID or URL",
                value="1mezi6NHOQZj0VR_TzmRb9UdkFGrFiWg7qJmvOrJopgA",
                key="gsheet_input")
            if st.button("📤 スプレッドシートから読み込み", type="primary",
                         use_container_width=True, key="btn_load_gsheet"):
                try:
                    with st.spinner("読み込み中..."):
                        from shift_scheduler import load_gsheet as _load_gs
                        staff_list, reqs_dict, gs_settings, _ = _load_gs(gsheet_id)
                    if gs_settings:
                        _apply_settings(gs_settings)
                    st.session_state.staff_df = _staff_to_df(staff_list)
                    st.session_state.requests_df = _reqs_to_df(reqs_dict, staff_list, num_days)
                    st.session_state.data_loaded = True
                    st.success(f"✅ {len(staff_list)}人のスタッフと希望を読み込みました")
                    st.rerun()
                except Exception as e:
                    st.error(f"読み込みエラー: {e}")

    # 読み込み状況
    st.markdown("---")
    if st.session_state.data_loaded:
        n_staff = len([n for n in st.session_state.staff_df["名前"].dropna() if str(n).strip()])
        st.success(f"✅ データ読み込み済み（{n_staff}人）→ 「スタッフ・勤務希望」タブで確認・編集できます")
        # 読み込み結果サマリー（Layer 2+3: 意図通り読めてるか1秒で確認）
        _render_load_preview(st.session_state.staff_df, st.session_state.get("requests_df"))
    else:
        st.info("💡 テンプレートをDLして記入 → アップロード、またはスプレッドシートから読み込んでください")

# ============================================================
# Tab 1: スタッフ・勤務希望（統合タブ）
# ============================================================
with tab1:
    st.subheader(f"スタッフ・勤務希望 — {year}年{month}月（{num_days}日間）")

    # --- スタッフ枠数設定 ---
    current_count = len(st.session_state.staff_df.dropna(subset=["名前"]))
    col_cnt, col_btn, col_spacer = st.columns([1, 1, 2])
    with col_cnt:
        target_count = st.number_input("スタッフ人数", min_value=1, max_value=100,
                                        value=max(current_count, 5), step=1, key="staff_total_count")
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button(f"✅ {target_count}人枠に確定", key="btn_set_staff_count"):
            current_df = st.session_state.staff_df
            valid_rows = current_df.dropna(subset=["名前"])
            valid_rows = valid_rows[valid_rows["名前"].str.strip() != ""]
            existing = len(valid_rows)
            if target_count > existing:
                new_rows = []
                for i in range(target_count - existing):
                    new_rows.append({
                        "名前": f"スタッフ{existing + i + 1}",
                        "Tier": "C", "夜勤専従": False, "時短": False,
                        "週勤務": None, "前月末": "", "夜勤Min": None,
                        "夜勤Max": None, "連勤Max": None, "勤務曜日": "", "祝日不可": False,
                        "土日不可": False, "夜勤研修": False, "研修夜勤回数": None,
                        "新人": False, "新人卒業日": None,
                    })
                st.session_state.staff_df = pd.concat(
                    [valid_rows, pd.DataFrame(new_rows)], ignore_index=True)
            elif target_count < existing:
                st.session_state.staff_df = valid_rows.head(target_count).reset_index(drop=True)
            st.rerun()

    # --- 表示モード切替 ---
    view_mode = st.radio(
        "表示モード",
        ["👤 スタッフ情報", "📝 勤務希望", "👤+📝 すべて"],
        horizontal=True, index=2, key="view_mode",
    )

    # --- スタッフ情報カラム選択（表示/非表示） ---
    _all_staff_detail_cols = ["Tier", "夜勤専従", "時短", "週勤務", "前月末",
                              "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可",
                              "夜勤研修", "研修夜勤回数", "新人", "新人卒業日"]
    if view_mode in ("👤 スタッフ情報", "👤+📝 すべて"):
        with st.expander("⚙ 表示カラム選択", expanded=False):
            _visible_staff_cols = st.multiselect(
                "表示するスタッフ情報カラム",
                _all_staff_detail_cols,
                default=_all_staff_detail_cols,
                key="staff_col_toggle",
            )
    else:
        _visible_staff_cols = _all_staff_detail_cols

    # --- 定義パネル（表示モードに応じて出し分け） ---
    if view_mode in ("👤 スタッフ情報", "👤+📝 すべて"):
        with st.expander("📖 Tier（スキルランク）の定義", expanded=False):
            st.markdown("""
| Tier | 説明 | 夜勤の役割 |
|------|------|-----------|
| **A** | ベテラン・リーダー格 | 日勤/夜勤リーダー単独可（必ずA/ABが1人以上） |
| **AB** | 中堅・リーダー代行可 | Aが不在時に夜勤リーダー代行 |
| **B** | 一人立ち済み | B+B/B+C族の夜勤ペアは**禁止**（ハード） |
| **C+** | C既卒（経験数ヶ月〜数年） | C+C+/C++通常Cペア禁止、A/AB下で夜勤可 |
| **C** | 新人・経験浅い | 必ずA/AB/Bと夜勤ペア（C+Cペア禁止） |

**各カラムの説明:**
- **夜勤専従**: ONにすると夜勤・明け・休のみの勤務パターンになります
- **時短**: ONにすると時短(ST)・早出(E)・遅出(L)・休のみ。夜勤/長日勤は入りません
- **週勤務**: パートタイムの場合に週あたりの勤務日数を指定（空欄=フルタイム）
- **前月末**: 前月最終日の勤務（夜=夜勤中/明=夜勤明け）→翌月初の制約に反映
- **夜勤Min/Max**: 個別の夜勤回数制限（空欄=全体設定に従う）
- **連勤Max**: 最大連続勤務日数（空欄=全体設定に従う）
- **勤務曜日**: パート等で特定曜日のみ勤務する場合（例: 月水金）
- **祝日不可**: ONにすると祝日に勤務を入れません
- **土日不可**: ONにすると土日に勤務を入れません
- **夜勤研修**: ONにすると通常夜勤2人に加え3人目（研修枠）として夜勤配置。研修者同士は同日に入りません
- **研修夜勤回数**: 月間の研修夜勤上限回数（空欄=制限なし）
- **新人**: ONで新人扱い。日勤配置基準の頭数にはカウントされますが、リーダー判定・独立C判定（B+Cペア制約）からは除外されます。夜勤を禁止したい場合は「夜勤Max=0」を併設
- **新人卒業日**: その日まで新人扱い、翌日から通常運用（空欄=月末まで新人）。月中から独り立ちさせる場合に使用
""")

    if view_mode in ("📝 勤務希望", "👤+📝 すべて"):
        with st.expander("📖 シフト種別の定義", expanded=False):
            st.markdown("""
| 記号 | 正式名 | 時間帯 | 備考 |
|------|--------|--------|------|
| **日** | 日勤 | 8:00〜17:00 (8h) | 標準勤務 |
| **夜** | 夜勤 | 16:45〜翌9:00 (16h) | 翌日は自動で「明け」 |
| **準** | 短夜勤 | 17:00〜翌5:00 (12h) | 夜勤の短縮版 |
| **早** | 早出 | 7:00〜16:00 (8h) | 日勤の早番 |
| **遅** | 遅出 | 12:00〜21:00 (8h) | 日勤の遅番 |
| **長** | 長日勤 | 8:45〜21:00 (12h) | 日勤の延長版 |
| **短** | 時短 | 8:45〜16:00 (6.25h) | 育児・介護向け |
| **休** | 公休 | — | 通常の休日 |
| **研** | 研修 | — | 研修日（勤務扱い） |
| **夜不** | 夜勤不可 | — | この日は夜勤に入れないで |
| **休暇** | 有給休暇 | — | 休暇申請済み |
| **明休** | 明または休 | — | その日を明または休に |

**注意:** 夜勤専従スタッフは「早・遅」を選択しないでください（自動で除外されます）
""")

    # --- 日付ヘッダー準備 ---
    wdj = ["月", "火", "水", "木", "金", "土", "日"]
    fwd = date(year, month, 1).weekday()
    header_map = {}
    for d in range(1, num_days + 1):
        wd = wdj[(fwd + d - 1) % 7]
        is_hol = d in holidays_auto
        is_we = d in weekends_auto
        suffix = "祝" if is_hol else ("★" if is_we else "")
        header_map[str(d)] = f"{d}({wd}{suffix})"

    # --- スタッフ名の同期（staff_df → requests_df） ---
    staff_names = [n for n in st.session_state.staff_df["名前"].dropna().tolist() if str(n).strip()]
    if st.session_state.requests_df is None or set(st.session_state.requests_df["名前"].tolist()) != set(staff_names):
        if st.session_state.requests_df is not None:
            old = st.session_state.requests_df.set_index("名前")
            rows = []
            for name in staff_names:
                if name in old.index:
                    rows.append(old.loc[name].to_dict() | {"名前": name})
                else:
                    row = {"名前": name}
                    for d in range(1, num_days + 1):
                        row[str(d)] = ""
                    rows.append(row)
            st.session_state.requests_df = pd.DataFrame(rows)
        else:
            rows = []
            for name in staff_names:
                row = {"名前": name}
                for d in range(1, num_days + 1):
                    row[str(d)] = ""
                rows.append(row)
            st.session_state.requests_df = pd.DataFrame(rows)

    # --- 統合DataFrame構築 ---
    _staff_cols = ["Tier", "夜勤専従", "時短", "週勤務", "前月末",
                   "夜勤Min", "夜勤Max", "連勤Max", "勤務曜日", "祝日不可", "土日不可",
                   "夜勤研修", "研修夜勤回数", "新人", "新人卒業日"]
    _day_cols = [str(d) for d in range(1, num_days + 1)]

    # staff_df と requests_df を名前で結合
    _sdf = st.session_state.staff_df.copy()
    _rdf = st.session_state.requests_df.copy()
    # 名前の順序はstaff_dfを優先
    combined_df = _sdf.set_index("名前").reindex(staff_names)
    _rdf_indexed = _rdf.set_index("名前").reindex(staff_names)
    for col in _day_cols:
        if col in _rdf_indexed.columns:
            combined_df[col] = _rdf_indexed[col].values
        else:
            combined_df[col] = ""
    combined_df = combined_df.reset_index()

    # --- 表示モードに応じたカラム選択（トグル反映） ---
    if view_mode == "👤 スタッフ情報":
        show_cols = ["名前"] + _visible_staff_cols
    elif view_mode == "📝 勤務希望":
        # 勤務希望: 名前 + Tier（読み取り専用）+ 日付列
        show_cols = ["名前", "Tier"] + _day_cols
    else:
        show_cols = ["名前"] + _visible_staff_cols + _day_cols

    # --- column_config 構築 ---
    _col_config_defs = {
        "Tier": st.column_config.SelectboxColumn("Tier", options=["A", "AB", "B", "C+", "C"], width="small"),
        "夜勤専従": st.column_config.CheckboxColumn("夜勤専従", width="small"),
        "時短": st.column_config.CheckboxColumn("時短", width="small"),
        "週勤務": st.column_config.NumberColumn("週勤務", min_value=1, max_value=7, step=1, width="small"),
        "前月末": st.column_config.SelectboxColumn("前月末", options=["", "夜", "明"], width="small"),
        "夜勤Min": st.column_config.NumberColumn("夜勤Min", min_value=0, max_value=15, step=1, width="small"),
        "夜勤Max": st.column_config.NumberColumn("夜勤Max", min_value=0, max_value=15, step=1, width="small"),
        "連勤Max": st.column_config.NumberColumn("連勤Max", min_value=1, max_value=10, step=1, width="small"),
        "勤務曜日": st.column_config.TextColumn("勤務曜日", width="small", help="例: 月火木"),
        "祝日不可": st.column_config.CheckboxColumn("祝日不可", width="small"),
        "土日不可": st.column_config.CheckboxColumn("土日不可", width="small"),
        "夜勤研修": st.column_config.CheckboxColumn("夜勤研修", width="small"),
        "研修夜勤回数": st.column_config.NumberColumn("研修夜勤回数", min_value=1, max_value=15, step=1, width="small"),
        "新人": st.column_config.CheckboxColumn("新人", width="small", help="ONで新人扱い（頭数のみカウント・リーダー/ペア判定外）"),
        "新人卒業日": st.column_config.NumberColumn("新人卒業日", min_value=1, max_value=31, step=1, width="small", help="その日まで新人。空欄=月末まで新人"),
    }
    col_config = {}
    # 勤務希望モードでは名前・Tierを読み取り専用に
    if view_mode == "📝 勤務希望":
        col_config["名前"] = st.column_config.TextColumn("名前", width="medium", disabled=True)
        col_config["Tier"] = st.column_config.TextColumn("Tier", width="small", disabled=True)
    else:
        col_config["名前"] = st.column_config.TextColumn("名前", width="medium")
    # 選択されたスタッフ情報カラムのみ追加（勤務希望モード以外）
    if view_mode != "📝 勤務希望":
        for c in _visible_staff_cols:
            if c in _col_config_defs:
                col_config[c] = _col_config_defs[c]
    # 勤務希望カラム
    if view_mode != "👤 スタッフ情報":
        _enabled_set = set(st.session_state.get("enabled_shifts",
                            ["日", "夜", "準", "早", "遅", "長", "短", "休", "研", "委", "夜不", "休暇", "明休"]))
        shift_options = [""] + [s for s in ["日", "夜", "準", "早", "遅", "長", "短", "休", "研", "委", "夜不", "休暇", "明休"]
                                if s in _enabled_set]
        for d in range(1, num_days + 1):
            col_config[str(d)] = st.column_config.SelectboxColumn(
                header_map[str(d)], options=shift_options, width="small")

    # --- data_editor（統合） ---
    edited_combined = st.data_editor(
        combined_df[show_cols],
        num_rows="dynamic" if view_mode != "📝 勤務希望" else "fixed",
        column_config=col_config,
        use_container_width=True,
        key="combined_editor",
        hide_index=True,
    )

    # --- 編集結果を元のDFに書き戻す ---
    # スタッフ情報: 常に全カラムを保持し、表示中のカラムのみ上書き
    _full_staff = st.session_state.staff_df.copy()
    if view_mode == "📝 勤務希望":
        # 勤務希望モードではスタッフ属性は一切変更しない
        edited_staff = _full_staff
    else:
        # 表示中のカラムだけ edited_combined から取得し、非表示カラムは元の値を保持
        _visible_cols_in_edit = [c for c in _visible_staff_cols if c in edited_combined.columns]
        # 行数が変わった場合（行追加・削除）に対応: edited_combined の名前ベースでマージ
        edited_staff = _full_staff.copy()
        # edited_combined の行数に合わせてリサイズ
        _edit_names = edited_combined["名前"].tolist() if "名前" in edited_combined.columns else []
        _new_rows = []
        for idx in range(len(edited_combined)):
            row = {}
            row["名前"] = edited_combined.iloc[idx]["名前"] if "名前" in edited_combined.columns else ""
            # 表示中カラムは編集値を使用
            for c in _visible_cols_in_edit:
                row[c] = edited_combined.iloc[idx][c]
            # 非表示カラムは既存staff_dfの同名行から取得（なければデフォルト）
            _name = row.get("名前", "")
            _match = _full_staff[_full_staff["名前"] == _name] if _name and str(_name).strip() else pd.DataFrame()
            for c in _all_staff_detail_cols:
                if c not in _visible_cols_in_edit:
                    if len(_match) > 0:
                        row[c] = _match.iloc[0][c] if c in _match.columns else None
                    else:
                        row[c] = None
            _new_rows.append(row)
        edited_staff = pd.DataFrame(_new_rows, columns=["名前"] + _all_staff_detail_cols)
    st.session_state.staff_df = edited_staff

    # 勤務希望
    _req_out_cols = [c for c in (["名前"] + _day_cols) if c in edited_combined.columns]
    if len(_req_out_cols) > 1:  # 名前+日付列がある
        edited_reqs = edited_combined[_req_out_cols].copy()
        st.session_state.requests_df = edited_reqs
    else:
        edited_reqs = st.session_state.requests_df

    if holidays_auto:
        hol_names = {d.day: name for d, name in jpholiday.month_holidays(year, month)}
        hol_str = "、".join(f"{d}日({hol_names.get(d, '祝')})" for d in sorted(holidays_auto))
        st.info(f"🗓 {year}年{month}月の祝日: {hol_str}")

# ============================================================
# Tab 3: 生成
# ============================================================
with tab3:
    st.subheader("勤務表生成")

    col_a, col_b, col_c = st.columns(3)
    valid_staff = edited_staff.dropna(subset=["名前"])
    valid_staff = valid_staff[valid_staff["名前"].str.strip() != ""]
    ft_count = len(valid_staff[valid_staff["週勤務"].isna()]) if "週勤務" in valid_staff.columns else len(valid_staff)
    pt_count = len(valid_staff[valid_staff["週勤務"].notna()]) if "週勤務" in valid_staff.columns else 0
    ded_count = len(valid_staff[valid_staff["夜勤専従"] == True]) if "夜勤専従" in valid_staff.columns else 0

    col_a.metric("スタッフ数", f"{len(valid_staff)}人")
    col_b.metric("フルタイム / パート", f"{ft_count} / {pt_count}")
    col_c.metric("夜勤専従", f"{ded_count}人")

    st.markdown("---")

    if st.button("🚀 勤務表を生成", type="primary", use_container_width=True):
        staff_list = []
        for _, row in valid_staff.iterrows():
            name = str(row["名前"]).strip()
            tier = str(row["Tier"]).strip() if pd.notna(row["Tier"]) else "C"
            if tier not in VALID_TIERS:
                st.warning(f"⚠ {name}: Tier '{tier}' 不正 → スキップ")
                continue
            ded = bool(row.get("夜勤専従", False))
            weekly = int(row["週勤務"]) if pd.notna(row.get("週勤務")) else None
            prev = str(row.get("前月末", "")).strip()
            if prev not in ("夜", "明", ""):
                prev = ""
            n_min = int(row["夜勤Min"]) if pd.notna(row.get("夜勤Min")) else None
            n_max = int(row["夜勤Max"]) if pd.notna(row.get("夜勤Max")) else None
            c_max = int(row["連勤Max"]) if pd.notna(row.get("連勤Max")) else None
            wd_str = str(row.get("勤務曜日", "")).strip()
            work_days = None
            if wd_str:
                work_days = set()
                for ch in wd_str:
                    if ch in WEEKDAY_MAP:
                        work_days.add(WEEKDAY_MAP[ch])
                if not work_days:
                    work_days = None
            no_hol = bool(row.get("祝日不可", False))
            short_t = bool(row.get("時短", False))
            no_we = bool(row.get("土日不可", False))
            night_tr = bool(row.get("夜勤研修", False))
            nt_max = int(row["研修夜勤回数"]) if pd.notna(row.get("研修夜勤回数")) else None
            new_h = bool(row.get("新人", False))
            nh_grad = int(row["新人卒業日"]) if pd.notna(row.get("新人卒業日")) else None
            staff_list.append(Staff(name, tier, ded, weekly, prev,
                                     n_min, n_max, c_max, work_days, no_hol, short_t,
                                     no_we, night_tr, nt_max, new_h, nh_grad))

        if not staff_list:
            st.error("スタッフが0人です")
        else:
            reqs_dict = {}
            dedicated_names = {s.name for s in staff_list if s.dedicated}
            for _, row in edited_reqs.iterrows():
                name = str(row["名前"]).strip()
                if not name:
                    continue
                rq = {}
                for d in range(1, num_days + 1):
                    val = str(row.get(str(d), "")).strip()
                    if val and val in SHIFT_REVERSE:
                        rq[d] = SHIFT_REVERSE[val]
                    elif val and val in (D, N, O, R):
                        rq[d] = val
                # 夜勤専従: 早出・遅出の希望を除外
                if name in dedicated_names:
                    dropped = [d for d, v in rq.items() if v in (E, L)]
                    if dropped:
                        st.warning(f"⚠ {name}（専従）: {len(dropped)}日分の早出/遅出希望を無視しました")
                        for d in dropped:
                            del rq[d]
                if rq:
                    reqs_dict[name] = rq

            with st.spinner(f"最適化計算中... （最大{time_limit}秒 × {num_patterns}パターン）"):
                import io, contextlib
                console = io.StringIO()
                with contextlib.redirect_stdout(console):
                    results = build_and_solve(staff_list, reqs_dict, settings,
                                              num_patterns=num_patterns,
                                              night_hours=night_hours,
                                              night_72h_mode=night_72h_mode,
                                              op_rules=op_rules,
                                              enabled_shifts=st.session_state.get("enabled_shifts"))

            console_output = console.getvalue()

            if results:
                st.session_state.results = results
                st.session_state.console_output = console_output
                st.success(f"✅ {len(results)}パターン生成完了！「結果」タブで確認できます。")

                buf = BytesIO()
                from openpyxl import Workbook
                wb = Workbook()
                wb.remove(wb.active)
                for res in results:
                    pat = res.get("pattern_num", 1)
                    title = f"パターン{pat}" if len(results) > 1 else "勤務表"
                    _write_one_sheet(wb, res, title)
                wb.save(buf)
                buf.seek(0)
                st.session_state.excel_bytes = buf.getvalue()
            else:
                st.error("❌ 解なし（Infeasible）。設定・希望を見直してください。")
                with st.expander("ソルバーログ"):
                    st.code(console_output)

# ============================================================
# Tab 4: 結果表示
# ============================================================
with tab4:
    if st.session_state.results is None:
        st.info("「生成」タブで勤務表を作成してください。")
    else:
        results = st.session_state.results

        if len(results) > 1:
            pat_idx = st.selectbox("パターン選択",
                                    range(len(results)),
                                    format_func=lambda i: f"パターン {i+1}")
        else:
            pat_idx = 0

        result = results[pat_idx]
        schedule = result["schedule"]
        names = result["names"]
        tiers = result["tiers"]
        r_num_days = result["num_days"]
        r_year = result["year"]
        r_month = result["month"]
        r_holidays = result.get("holidays", set())
        r_weekends = result.get("weekends", set())
        r_public_off = result.get("public_off", 13)
        r_weekly = result.get("weekly", {})
        r_dedicated = result.get("dedicated", {})
        missed = result.get("missed_requests", {})

        st.subheader(f"📊 {r_year}年{r_month}月 勤務表 — パターン {pat_idx + 1}")

        # ── 出力ボタン（上部に目立つように配置） ──
        _dl_top1, _dl_top2, _dl_top3 = st.columns([2, 2, 3])
        if "excel_bytes" in st.session_state:
            _dl_top1.download_button(
                label="📥 Excel出力",
                data=st.session_state.excel_bytes,
                file_name=f"勤務表_{r_year}_{r_month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="dl_excel_top",
            )
        # 手動編集後の再出力ボタン
        if _dl_top2.button("🔄 最新版をExcel化", use_container_width=True, key="regen_excel_top"):
            from openpyxl import Workbook as _WB
            _buf = BytesIO()
            _wb = _WB()
            _wb.remove(_wb.active)
            for _res in results:
                _pat = _res.get("pattern_num", 1)
                _title = f"パターン{_pat}" if len(results) > 1 else "勤務表"
                _write_one_sheet(_wb, _res, _title)
            _wb.save(_buf)
            _buf.seek(0)
            st.session_state.excel_bytes = _buf.getvalue()
            st.success("✅ 最新のスケジュールでExcelを再生成しました")
            st.rerun()
        # Googleスプレッドシート出力
        if _dl_top3.button("📊 Googleスプレッドシート出力", use_container_width=True, key="btn_gsheet_result"):
            try:
                from shift_scheduler import _get_gc as _get_gsheet_client, _write_gsheet_one
                with st.spinner("Googleスプレッドシートに出力中..."):
                    gc = _get_gsheet_client()
                    title = f"勤務表_{r_year}_{r_month:02d}"
                    sh = gc.create(title)
                    # デフォルトSheet1を削除
                    try:
                        sh.del_worksheet(sh.sheet1)
                    except Exception:
                        pass
                    for _res in results:
                        _write_gsheet_one(sh, _res)
                    # 誰でもアクセス可能に（閲覧のみ）
                    sh.share("", perm_type="anyone", role="writer")
                    _gsheet_url = sh.url
                st.session_state.gsheet_result_url = _gsheet_url
                st.success(f"✅ Googleスプレッドシートに出力しました！")
                st.rerun()
            except ImportError:
                st.error("gspread / google-auth が必要です: `pip install gspread google-auth`")
            except Exception as e:
                st.error(f"エラー: {e}")

        if "gsheet_result_url" in st.session_state:
            st.markdown(f"🔗 **[Googleスプレッドシートを開く]({st.session_state.gsheet_result_url})**")

        # 夜勤系合計（N + SN）
        night_counts = {s: sum(schedule[s].count(t) for t in NIGHT_SHIFTS) for s in names}
        nc_regular = [v for s, v in night_counts.items()
                      if r_weekly.get(s) is None and not r_dedicated.get(s, False)]
        total_missed = sum(len(v) for v in missed.values())
        violations_pre, _, _ = check_nursing_guidelines(
            schedule, names, tiers, r_dedicated, night_hours)
        bad_pairs_pre, _, _ = check_skill_pairing(
            schedule, names, tiers, r_num_days, r_year, r_month)
        shortfalls_pre, _, _ = check_staffing_ratio(
            schedule, names, r_dedicated, r_weekly,
            r_num_days, bed_count, _ratio_val, r_year, r_month,
            check_night=_check_night)

        col1, col2, col3, col4, col5, col6 = st.columns(6)
        col1.metric("公休日数", f"{r_public_off}日")
        col2.metric("夜勤均等", f"{min(nc_regular)}〜{max(nc_regular)}回" if nc_regular else "—")
        col3.metric("未達希望", f"{total_missed}件" if total_missed else "0件 ✓",
                    delta="要確認" if total_missed else None, delta_color="inverse")
        col4.metric("ガイドライン", "✅ 適合" if not violations_pre else f"🚨 {len(violations_pre)}件",
                    delta="要対応" if violations_pre else None, delta_color="inverse")
        col5.metric("スキルペア", "✅ 適正" if not bad_pairs_pre else f"🚨 {len(bad_pairs_pre)}日",
                    delta="要対応" if bad_pairs_pre else None, delta_color="inverse")
        col6.metric("配置基準", "✅ 適合" if not shortfalls_pre else f"🚨 {len(shortfalls_pre)}日",
                    delta="要対応" if shortfalls_pre else None, delta_color="inverse")

        wdj = ["月", "火", "水", "木", "金", "土", "日"]
        fwd = date(r_year, r_month, 1).weekday()

        shift_colors = {
            D: "#FFFFFF",   # 日勤: 白
            N: "#4472C4",   # 夜勤: 青
            A: "#FFF2CC",   # 明け: 薄黄
            O: "#E2EFDA",   # 休み: 薄緑
            R: "#E8D5F5",   # 研修: 薄紫
            V: "#F4CCCC",   # 休暇: 薄赤
            E:  "#E6F3FF",  # 早出: 水色
            L:  "#FFF0E0",  # 遅出: 薄橙
            ST: "#F0FFF0",  # 時短: 薄緑系
            LD: "#FFF5CC",  # 長日勤: 薄黄緑
            SN: "#2E75B6",  # 短夜勤: 濃青
        }
        shift_text_colors = {N: "#FFFFFF", SN: "#FFFFFF"}

        day_cols = [f"{d+1}" for d in range(r_num_days)]
        table_data = []
        for s in names:
            row = {"名前": s, "Tier": tiers[s]}
            for d in range(r_num_days):
                row[day_cols[d]] = schedule[s][d]
            counts = {t: schedule[s].count(t) for t in SHIFTS}
            row["日"] = counts[D]
            row["夜"] = counts[N]
            row["明"] = counts[A]
            # 有効なシフトのみ集計列を追加
            _en = set(st.session_state.get("enabled_shifts", []))
            if "準" in _en: row["準"] = counts[SN]
            if "早" in _en: row["早"] = counts[E]
            if "遅" in _en: row["遅"] = counts[L]
            if "長" in _en: row["長"] = counts[LD]
            if "短" in _en: row["短"] = counts[ST]
            row["休"] = counts[O]
            row["研"] = counts[R]
            row["暇"] = counts[V]
            row["公休"] = counts[O] + counts[V]
            table_data.append(row)

        df = pd.DataFrame(table_data)

        def color_shifts(val):
            if val in shift_colors:
                bg = shift_colors[val]
                fg = shift_text_colors.get(val, "#000000")
                return f"background-color: {bg}; color: {fg}; text-align: center; font-weight: {'bold' if val == N else 'normal'}"
            return "text-align: center"

        # ── シフト表示モード切替 ─────────────────────────────
        view_mode = st.radio("表示モード", ["👁 確認", "✏️ 手動編集"],
                             horizontal=True, key=f"view_mode_{pat_idx}")

        if view_mode == "👁 確認":
            styled = df.style.map(color_shifts, subset=day_cols)
            st.dataframe(styled, use_container_width=True, height=600, hide_index=True)
        else:
            st.caption("⚠ セルを直接編集できます。変更後「変更を保存」を押してください。")
            _en = set(st.session_state.get("enabled_shifts", []))
            _sym_map = {"準": SN, "早": E, "遅": L, "長": LD, "短": ST, "委": I}
            shift_options_edit = [D, N, A, O, R, V] + [v for k, v in _sym_map.items() if k in _en]
            col_cfg_edit = {
                "名前": st.column_config.TextColumn("名前", disabled=True, width="small"),
                "Tier": st.column_config.TextColumn("Tier", disabled=True, width="small"),
            }
            for dc in day_cols:
                col_cfg_edit[dc] = st.column_config.SelectboxColumn(
                    dc, options=shift_options_edit, width="small")
            _summary_cols = ["日", "夜", "明"]
            if "準" in _en: _summary_cols.append("準")
            if "早" in _en: _summary_cols.append("早")
            if "遅" in _en: _summary_cols.append("遅")
            if "長" in _en: _summary_cols.append("長")
            if "短" in _en: _summary_cols.append("短")
            _summary_cols += ["休", "研", "暇", "公休"]
            for col in _summary_cols:
                if col in df.columns:
                    col_cfg_edit[col] = st.column_config.NumberColumn(col, disabled=True, width="small")

            edited_df = st.data_editor(
                df, column_config=col_cfg_edit,
                use_container_width=True, height=600, hide_index=True,
                key=f"manual_edit_{pat_idx}"
            )
            if st.button("💾 変更を保存", key=f"save_edit_{pat_idx}"):
                # edited_df の内容をresultsに反映
                new_schedule = {}
                for _, row in edited_df.iterrows():
                    sname = row["名前"]
                    new_schedule[sname] = [row[dc] for dc in day_cols]
                st.session_state.results[pat_idx]["schedule"] = new_schedule
                st.success("✅ 変更を保存しました。ページを再読み込みすると反映されます。")
                st.rerun()

        # ── 夜勤回数分布チャート ─────────────────────────────
        with st.expander("📊 スタッフ別 夜勤・勤務統計", expanded=False):
            _en = set(st.session_state.get("enabled_shifts", []))
            stat_rows = []
            for s in names:
                sh = schedule[s]
                n_cnt = sh.count(N)
                sn_cnt = sh.count(SN)
                total_h = n_cnt * night_hours + sn_cnt * 12
                _row = {
                    "名前": s, "Tier": tiers[s],
                    "夜勤": n_cnt,
                    "月夜勤時間": f"{total_h}h",
                    "72h": "🚨" if total_h > 72 else ("⚠" if total_h >= 64 else "✅"),
                    "日勤": sh.count(D),
                    "休み": sh.count(O) + sh.count(V),
                }
                if "準" in _en: _row["短夜勤"] = sn_cnt
                if "早" in _en: _row["早出"] = sh.count(E)
                if "遅" in _en: _row["遅出"] = sh.count(L)
                if "長" in _en: _row["長日勤"] = sh.count(LD)
                if "短" in _en: _row["時短"] = sh.count(ST)
                stat_rows.append(_row)
            stat_df = pd.DataFrame(stat_rows)
            st.dataframe(stat_df, use_container_width=True, hide_index=True)

            st.markdown("**夜勤回数分布**")
            chart_data = pd.DataFrame({
                "夜勤回数": [night_counts[s] for s in names]
            }, index=names)
            st.bar_chart(chart_data, height=200)

        with st.expander("🌙 日別夜勤ペア", expanded=False):
            pair_data = []
            for d in range(r_num_days):
                wd = wdj[(fwd + d) % 7]
                hol = "祝" if (d+1) in r_holidays else ("★" if (d+1) in r_weekends else "")
                nn_n  = [s for s in names if schedule[s][d] == N]
                nn_sn = [s for s in names if schedule[s][d] == SN]
                members = (
                    [f"{s}({tiers[s]})" for s in nn_n] +
                    [f"{s}({tiers[s]})[準]" for s in nn_sn]
                )
                pair_data.append({
                    "日": f"{d+1}日({wd}{hol})",
                    "夜勤メンバー": " + ".join(members) if members else "—",
                })
            st.dataframe(pd.DataFrame(pair_data), use_container_width=True, hide_index=True)

        if missed:
            with st.expander("⚠ 未達希望", expanded=True):
                for s, ds in missed.items():
                    st.warning(f"{s}: {', '.join(f'{d}日' for d in ds)}")
        else:
            st.success("✓ 全希望達成")

        # ── 日本看護協会 + 労基法ガイドラインチェック ─────────
        violations, warnings_gl, ok_list = check_nursing_guidelines(
            schedule, names, tiers, r_dedicated, night_hours
        )
        gl_label = "✅ ガイドライン適合" if not violations else f"🚨 違反 {len(violations)}件"
        with st.expander(f"⚖ 夜勤ガイドライン & 労基法チェック — {gl_label}", expanded=bool(violations)):
            st.caption(f"根拠: 日本看護協会「夜勤・交代制勤務に関するガイドライン」(2013年) ／ 交代制: {shift_system}")
            gl_col1, gl_col2, gl_col3 = st.columns(3)
            gl_col1.metric("🚨 違反", f"{len(violations)}件",
                           delta="要対応" if violations else None, delta_color="inverse")
            gl_col2.metric("⚠ 注意", f"{len(warnings_gl)}件")
            gl_col3.metric("✅ 適合", f"{len(ok_list)}名")

            st.markdown("##### チェック項目")
            rules_info = [
                ("夜勤回数", "月8回以内（夜勤専従者は対象外）"),
                ("72時間規制", f"月夜勤時間72時間以内（{night_hours}h×回数）"),
                ("連続夜勤", "連続夜勤2サイクル以内"),
                ("インターバル", "勤務間隔11時間以上（明け翌日に日勤がないか）"),
                ("夜勤後の明け", "夜勤翌日は必ず明けシフト"),
            ]
            for rule, desc in rules_info:
                v_cnt = sum(1 for x in violations if x["rule"] == rule)
                w_cnt = sum(1 for x in warnings_gl if x["rule"] == rule)
                icon = "🚨" if v_cnt else ("⚠" if w_cnt else "✅")
                st.markdown(f"- {icon} **{rule}**: {desc}"
                            + (f" → **{v_cnt}件違反**" if v_cnt else "")
                            + (f" → {w_cnt}件注意" if w_cnt else ""))

            if violations:
                st.markdown("##### 🚨 違反一覧")
                vcols = ["名前", "Tier", "夜勤回数", "月夜勤時間", "rule", "detail"]
                st.dataframe(
                    pd.DataFrame(violations)[[c for c in vcols if c in pd.DataFrame(violations).columns]]
                    .rename(columns={"rule": "項目", "detail": "内容"}),
                    use_container_width=True, hide_index=True
                )

            if warnings_gl:
                st.markdown("##### ⚠ 注意一覧")
                st.dataframe(
                    pd.DataFrame(warnings_gl)[["名前", "Tier", "夜勤回数", "月夜勤時間", "rule", "detail"]]
                    .rename(columns={"rule": "項目", "detail": "内容"}),
                    use_container_width=True, hide_index=True
                )

            if ok_list:
                st.markdown(f"##### ✅ 全項目適合: {', '.join(ok_list)}")

        # ── スキルペアチェック ───────────────────────────────
        bad_pairs, warn_pairs, ok_pair_days = check_skill_pairing(
            schedule, names, tiers, r_num_days, r_year, r_month)
        pair_label = ("✅ 全日適正" if not bad_pairs and not warn_pairs
                      else f"🚨 NG {len(bad_pairs)}日" if bad_pairs
                      else f"⚠ 注意 {len(warn_pairs)}日")
        with st.expander(f"👥 夜勤スキルペアチェック — {pair_label}",
                         expanded=bool(bad_pairs)):
            st.caption("A/AB（ベテラン）が各夜勤ペアに1名以上含まれるかチェックします。")
            sp1, sp2, sp3 = st.columns(3)
            sp1.metric("🚨 全員新人（C）日", f"{len(bad_pairs)}日",
                       delta="要対応" if bad_pairs else None, delta_color="inverse")
            sp2.metric("⚠ A/AB不在日", f"{len(warn_pairs)}日")
            sp3.metric("✅ 適正ペア日", f"{ok_pair_days}日")

            tier_legend = "Tier: **A**=エキスパート / **AB**=上級 / **B**=中堅 / **C**=新人"
            st.caption(tier_legend)

            if bad_days := bad_pairs:
                st.markdown("##### 🚨 問題あり（新人のみ夜勤）")
                st.dataframe(pd.DataFrame(bad_days), use_container_width=True, hide_index=True)
            if warn_days := warn_pairs:
                st.markdown("##### ⚠ 注意（A/AB不在）")
                st.dataframe(pd.DataFrame(warn_days), use_container_width=True, hide_index=True)
            if not bad_pairs and not warn_pairs:
                st.success("全夜勤ペアにA/AB（ベテラン）が含まれています。")

        # ── 人員配置基準チェック ─────────────────────────────
        shortfalls, ok_days, req_nurses = check_staffing_ratio(
            schedule, names, r_dedicated, r_weekly,
            r_num_days, bed_count, _ratio_val, r_year, r_month,
            check_night=_check_night
        )
        ratio_label = "✅ 基準適合" if not shortfalls else f"🚨 不足 {len(shortfalls)}日"
        scope_label = "日勤・夜勤（24時間）" if _check_night else "日勤帯"
        with st.expander(
            f"🏥 人員配置基準チェック — {unit_type}（{nurse_ratio}） — {ratio_label}",
            expanded=bool(shortfalls)
        ):
            st.caption(f"根拠: {_basis_label}")
            st.caption(f"配置基準: {nurse_ratio} ／ 適用範囲: {scope_label} ／ 病床数: {bed_count}床 ／ 必要人数: {req_nurses}人以上")
            rc1, rc2, rc3 = st.columns(3)
            rc1.metric("🚨 不足日数", f"{len(shortfalls)}日",
                       delta="要対応" if shortfalls else None, delta_color="inverse")
            rc2.metric("✅ 基準達成", f"{ok_days}日")
            rc3.metric("必要人数", f"{req_nurses}人/{scope_label}")
            if shortfalls:
                st.dataframe(pd.DataFrame(shortfalls), use_container_width=True, hide_index=True)
            else:
                st.success(f"全{r_num_days}日、{scope_label} {req_nurses}人以上を確保しています。")

        with st.expander("📈 日別集計", expanded=False):
            _en = set(st.session_state.get("enabled_shifts", []))
            summary_data = {"日付": [f"{d+1}" for d in range(r_num_days)]}
            _day_summary_items = [("日勤", D), ("夜勤", N)]
            if "準" in _en: _day_summary_items.append(("短夜勤", SN))
            _day_summary_items.append(("明け", A))
            if "早" in _en: _day_summary_items.append(("早出", E))
            if "遅" in _en: _day_summary_items.append(("遅出", L))
            if "長" in _en: _day_summary_items.append(("長日勤", LD))
            if "短" in _en: _day_summary_items.append(("時短", ST))
            _day_summary_items += [("休み", O), ("研修", R)]
            for label, shift in _day_summary_items:
                summary_data[label] = [sum(1 for s in names if schedule[s][d] == shift)
                                        for d in range(r_num_days)]
            # 新人を除いた日勤系合計（新人がいる場合のみ）
            _nh_days = result.get("new_hire_days", {})
            if _nh_days:
                def _is_nh(s, d):
                    return (d + 1) in _nh_days.get(s, set())
                summary_data["日勤系計(新人除く)"] = [
                    sum(1 for s in names if schedule[s][d] in DAY_SHIFTS and not _is_nh(s, d))
                    for d in range(r_num_days)
                ]
            # 日勤系合計列を追加
            summary_data["日勤系計"] = [
                sum(1 for s in names if schedule[s][d] in DAY_SHIFTS)
                for d in range(r_num_days)
            ]
            st.dataframe(pd.DataFrame(summary_data), use_container_width=True, hide_index=True)

        with st.expander("🔧 ソルバーログ"):
            st.code(st.session_state.get("console_output", ""))

        st.markdown("---")
        dl_col1, dl_col2 = st.columns(2)

        if "excel_bytes" in st.session_state:
            dl_col1.download_button(
                label="📥 勤務表 Excelダウンロード",
                data=st.session_state.excel_bytes,
                file_name=f"勤務表_{r_year}_{r_month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        # ── 様式9帳票 ──
        with st.expander("📋 様式9帳票（夜勤・交代制勤務実態調査）を出力", expanded=False):
            st.caption("日本看護協会・都道府県看護協会への月次提出用フォームです。")
            f9_col1, f9_col2 = st.columns(2)
            f9_facility = f9_col1.text_input("施設名", placeholder="○○病院", key=f"f9_facility_{pat_idx}")
            f9_ward = f9_col2.text_input("病棟・部署名", placeholder="ICU・CCU", key=f"f9_ward_{pat_idx}")

            youshiki9_bytes = _generate_youshiki9_excel(
                schedule, names, tiers, r_dedicated, r_weekly,
                r_year, r_month, night_hours,
                facility_name=f9_facility, ward_name=f9_ward
            )
            dl_col2.download_button(
                label="📋 様式9帳票ダウンロード",
                data=youshiki9_bytes,
                file_name=f"様式9_{r_year}_{r_month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"dl_youshiki9_{pat_idx}",
            )

# ============================================================
# Tab 5: ダッシュボード
# ============================================================
with tab5:
    st.subheader("🏠 ダッシュボード — 月次KPI概要")

    if st.session_state.results is None:
        st.info("「生成」タブで勤務表を作成すると、ここにKPIが表示されます。")
    else:
        # 最新パターンの結果を使用
        _res = st.session_state.results[0]
        _schedule = _res["schedule"]
        _names = _res["names"]
        _tiers = _res["tiers"]
        _r_num_days = _res["num_days"]
        _r_year = _res["year"]
        _r_month = _res["month"]
        _r_dedicated = _res.get("dedicated", {})
        _r_weekly = _res.get("weekly", {})
        _r_public_off = _res.get("public_off", 0)
        _missed = _res.get("missed_requests", {})

        st.caption(f"{_r_year}年{_r_month}月 ／ スタッフ {len(_names)}名 ／ {_r_num_days}日間")

        # ── KPIカード ──
        _v, _w, _ok = check_nursing_guidelines(_schedule, _names, _tiers, _r_dedicated, night_hours)
        _bp, _wp, _okp = check_skill_pairing(_schedule, _names, _tiers, _r_num_days, _r_year, _r_month)
        _sf, _okd, _req = check_staffing_ratio(
            _schedule, _names, _r_dedicated, _r_weekly,
            _r_num_days, bed_count, _ratio_val, _r_year, _r_month,
            check_night=_check_night)
        _total_missed = sum(len(v) for v in _missed.values())
        _nc = {s: _schedule[s].count(N) for s in _names}
        _nc_reg = [v for s, v in _nc.items()
                   if _r_weekly.get(s) is None and not _r_dedicated.get(s, False)]

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("スタッフ数", f"{len(_names)}人",
                  delta=f"うち専従 {sum(1 for s in _names if _r_dedicated.get(s))}人")
        k2.metric("公休日数", f"{_r_public_off}日")
        k3.metric("夜勤回数（均等）",
                  f"{min(_nc_reg)}〜{max(_nc_reg)}回" if _nc_reg else "—")
        k4.metric("未達希望", f"{_total_missed}件" if _total_missed else "0件 ✓",
                  delta="要確認" if _total_missed else None, delta_color="inverse")

        st.markdown("---")
        k5, k6, k7, k8 = st.columns(4)
        k5.metric("ガイドライン違反", f"{len(_v)}件",
                  delta="🚨 要対応" if _v else "✅ 適合", delta_color="inverse" if _v else "off")
        k6.metric("スキルペアNG日", f"{len(_bp)}日",
                  delta="🚨 要対応" if _bp else "✅ 適正", delta_color="inverse" if _bp else "off")
        k7.metric("配置基準不足日", f"{len(_sf)}日",
                  delta="🚨 要対応" if _sf else "✅ 適合", delta_color="inverse" if _sf else "off")
        k8.metric("コンプライアンス",
                  "✅ 全適合" if not _v and not _bp and not _sf else "🚨 要対応",
                  delta=None)

        st.markdown("---")
        dash_col1, dash_col2 = st.columns(2)

        # ── Tier分布 ──
        with dash_col1:
            st.markdown("##### 👥 スタッフ Tier分布")
            tier_counts = {}
            for s in _names:
                t = _tiers.get(s, "C")
                tier_counts[t] = tier_counts.get(t, 0) + 1
            tier_df = pd.DataFrame([
                {"Tier": t, "人数": tier_counts.get(t, 0)}
                for t in ["A", "AB", "B", "C+", "C"]
            ])
            st.dataframe(tier_df, use_container_width=True, hide_index=True)
            st.bar_chart(tier_df.set_index("Tier"), height=180)

        # ── 夜勤回数分布 ──
        with dash_col2:
            st.markdown("##### 🌙 夜勤回数分布（通常スタッフ）")
            reg_names = [s for s in _names
                         if _r_weekly.get(s) is None and not _r_dedicated.get(s, False)]
            if reg_names:
                nc_df = pd.DataFrame({
                    "夜勤回数": [_nc[s] for s in reg_names]
                }, index=reg_names)
                st.bar_chart(nc_df, height=180)
            else:
                st.info("通常スタッフなし")

        st.markdown("---")
        # ── 日別シフト構成 ──
        st.markdown("##### 📅 日別 シフト人数推移")
        daily_df = pd.DataFrame({
            "日勤": [sum(1 for s in _names if _schedule[s][d] == D) for d in range(_r_num_days)],
            "夜勤": [sum(1 for s in _names if _schedule[s][d] == N) for d in range(_r_num_days)],
            "休み": [sum(1 for s in _names if _schedule[s][d] in (O, V)) for d in range(_r_num_days)],
        }, index=[f"{d+1}日" for d in range(_r_num_days)])
        st.area_chart(daily_df, height=220)

        # ── 問題サマリー ──
        issues_any = bool(_v or _bp or _sf or _total_missed)
        if issues_any:
            st.markdown("---")
            st.markdown("##### 🚨 要対応サマリー")
            if _v:
                st.error(f"ガイドライン違反: {len(_v)}件 → 「結果」タブの⚖ガイドラインチェックを確認")
            if _bp:
                st.error(f"新人のみ夜勤: {len(_bp)}日 → 「結果」タブの👥スキルペアチェックを確認")
            if _sf:
                st.warning(f"配置基準不足: {len(_sf)}日 → 「結果」タブの🏥人員配置基準チェックを確認")
            if _total_missed:
                st.warning(f"未達希望: {_total_missed}件 → 「結果」タブの⚠未達希望を確認")
        else:
            st.success("✅ 全チェック項目に問題ありません。勤務表は提出可能な状態です。")
